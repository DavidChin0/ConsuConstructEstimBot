"""
EstimBot - Update BD Quantities from Schedules CSV
INPUT:  schedules_YYYYMMDD.csv  (step6_quantities)
        BaseDatosOpus2026.xlsx
OUTPUT: BaseDatosOpus2026_Updated_YYYYMMDD_HHMMSS.xlsx
"""

import csv
import re
import shutil
import os
import openpyxl
from collections import defaultdict
from datetime import datetime

# ── PATHS ─────────────────────────────────────────────────────────────────
ROOT_DIR = os.environ.get(
    "ESTIMBOT_PYREVIT_ROOT",
    os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
)
STEP6_DIR = os.environ.get(
    "ESTIMBOT_PYREVIT_SCHEDULES_DIR",
    os.path.join(ROOT_DIR, "step6_quantities")
)
BD_PATH = os.environ.get(
    "ESTIMBOT_PYREVIT_BD_PATH",
    os.path.join(ROOT_DIR, "BaseDatosOpus2026.xlsx")
)
OUTPUT_DIR = os.environ.get(
    "ESTIMBOT_PYREVIT_OUTPUT_DIR",
    ROOT_DIR
)

# ── AUTO-SELECT LATEST SCHEDULES CSV ──────────────────────────────────────
csv_files = sorted([f for f in os.listdir(STEP6_DIR) if f.startswith("schedules_") and f.endswith(".csv")])
if not csv_files:
    raise FileNotFoundError(f"No schedules CSV found in {STEP6_DIR}")
CSV_PATH = os.path.join(STEP6_DIR, csv_files[-1])

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
OUT_PATH  = os.path.join(OUTPUT_DIR, f"BaseDatosOpus2026_Updated_{timestamp}.xlsx")

print(f"CSV:    {CSV_PATH}")
print(f"BD:     {BD_PATH}")
print(f"Output: {OUT_PATH}")

# ── 1. PARSE CSV → SUM QUANTITIES BY KEYNOTE ──────────────────────────────
totals = defaultdict(float)
headers = []
keynote_col = qty_col = None

with open(CSV_PATH, encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    for row in reader:
        if not row:
            continue
        if row[0].startswith("###"):
            headers = []; keynote_col = qty_col = None
            continue
        if not headers:
            headers = row
            for i, h in enumerate(headers):
                hl = h.strip().lower()
                if 'keynote' in hl:
                    keynote_col = i
                if hl in ('count', 'length', 'area', 'volume', 'value'):
                    qty_col = i
            if qty_col is None and len(headers) >= 3:
                qty_col = 2
            continue
        if keynote_col is None:
            continue
        keynote = row[keynote_col].strip() if keynote_col < len(row) else ""
        if not keynote:
            continue
        qty_raw = row[qty_col].strip() if qty_col and qty_col < len(row) else ""
        qty_clean = re.sub(r'[^\d\.\,]', '', qty_raw).replace(',', '.')
        try:
            qty = float(qty_clean)
        except:
            qty = 1.0
        totals[keynote] += qty

print(f"\nUnique keynotes in CSV: {len(totals)}")

# ── 2. COPY BD AND UPDATE COL D ───────────────────────────────────────────
shutil.copy2(BD_PATH, OUT_PATH)
wb = openpyxl.load_workbook(OUT_PATH)
ws = wb.active

matched = zeroed = 0
unmatched_csv = []

for row in ws.iter_rows(min_row=3):
    key = str(row[0].value).strip() if row[0].value else ""
    if not key:
        continue
    if key in totals:
        row[3].value = round(totals[key], 4)
        matched += 1
    else:
        row[3].value = 0
        zeroed += 1

wb.save(OUT_PATH)

# ── 3. REPORT CSV KEYNOTES NOT FOUND IN BD ────────────────────────────────
bd_keys = set()
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[0]:
        bd_keys.add(str(row[0]).strip())
unmatched_csv = [k for k in totals if k not in bd_keys]

print(f"Matched+updated:    {matched}")
print(f"Zeroed (no match):  {zeroed}")
if unmatched_csv:
    print(f"\nCSV keynotes NOT in BD ({len(unmatched_csv)}):")
    for k in unmatched_csv:
        print(f"  {k}  →  {round(totals[k], 4)}")
print(f"\nDone: {OUT_PATH}")
