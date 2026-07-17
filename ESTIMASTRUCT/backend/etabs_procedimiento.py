# -*- coding: utf-8 -*-
"""
ETABS — Procedimiento explicativo + origen de datos + parser de export
======================================================================
Modulo ADITIVO. NO toca calculo_sismico_choc08.py ni calculo_estructural.py.

Contiene:
  1) ORIGEN_INPUTS  : de donde viene cada dato de entrada de la hoja sismica
                      (origen + tabla/norma/documento). Fuente textual.
  2) PROCEDIMIENTO  : pasos reproducidos de los MD fuente de verdad
                      (manual_procedimiento_etabs_estimastruct.md,
                       datos_ingreso_etabs_choc08_estimastruct.md,
                       flujo_etabs_modelado_choc08.md). Cada paso:
                       objetivo, navegacion ETABS exacta, datos+origen,
                       verificacion, errores comunes.
  3) EXPORT_ETABS_DOC : que contiene el archivo de export de ETABS,
                        como producirlo y como cargarlo.
  4) parse_export_etabs(): parser TOLERANTE (CSV/TSV/texto pegado de tablas
                        ETABS) que autollena W, T, V_din y deriva.

Fuentes de verdad (D:\\...\\03 Automation Projects\\etabs-flow\\):
  - manual_procedimiento_etabs_estimastruct.md  (15 capitulos)
  - datos_ingreso_etabs_choc08_estimastruct.md  (18 pasos)
  - flujo_etabs_modelado_choc08.md
"""

import csv
import io
import re


# ─────────────────────────────────────────────────────────────────────────────
# 1) ORIGEN DE CADA INPUT — textual (tooltip + procedimiento)
#    clave = key del input en la hoja (igual que SismoChoc08 / etabsState)
# ─────────────────────────────────────────────────────────────────────────────
ORIGEN_INPUTS = {
    "zona": {
        "titulo": "Zona sismica / Z",
        "origen": "Mapa de zonificacion sismica CHOC-08 Fig. 1.3.4-1 segun la "
                  "ubicacion del proyecto; el factor de zona Z (APS) sale de la "
                  "Tabla 1.3.4-1.",
        "referencia": "CHOC-08 Fig. 1.3.4-1 / Tabla 1.3.4-1",
        "ejemplo_cc135": "CC-135 Comayagua = Zona 3b -> Z = 0.25.",
    },
    "suelo": {
        "titulo": "Perfil de suelo (S, Ta, Tb, c)",
        "origen": "Estudio geotecnico del sitio (SPT / clasificacion del perfil); "
                  "el perfil (S1..S4) y sus parametros S, Ta, Tb, c salen de la "
                  "Tabla 1.3.4-2.",
        "referencia": "Estudio de suelo del proyecto + CHOC-08 Tabla 1.3.4-2",
        "ejemplo_cc135": "CC-135 = S1 (estudio de suelo CC-135: SPT 0-4.20 m, sin "
                         "arcilla blanda) -> S=1.0, Ta=0.155 s, Tb=0.364 s, c=2.0.",
    },
    "I": {
        "titulo": "Factor de importancia I",
        "origen": "Categoria de ocupacion / uso del edificio segun la Tabla 1.3.4-3.",
        "referencia": "CHOC-08 Tabla 1.3.4-3",
        "ejemplo_cc135": "CC-135 (aulas) = ocupacion especial -> I = 1.00.",
    },
    "Rw": {
        "titulo": "Factor de reduccion Rw",
        "origen": "Sistema estructural sismorresistente segun la Tabla 1.3.4-6.",
        "referencia": "CHOC-08 Tabla 1.3.4-6",
        "ejemplo_cc135": "CC-135 = marco intermedio de concreto -> Rw = 8. "
                         "(Marco comun Rw=5 esta PROHIBIDO en Zona 3.)",
    },
    "hn_m": {
        "titulo": "Altura total hn",
        "origen": "Altura total del edificio sobre la base. Sale de la arquitectura "
                  "o de ETABS Story Data (suma de las alturas de entrepiso).",
        "referencia": "Arquitectura / ETABS Story Data",
        "ejemplo_cc135": "CC-135 ~ 3 m (nivel de aulas).",
    },
    "W_t": {
        "titulo": "Peso sismico W",
        "origen": "Peso sismico (CHOC 1.3.5.1) = Dead + Super Dead + fraccion de Live. "
                  "Viene de ETABS: Mass Source / tabla Mass Summary, o de Base "
                  "Reactions FZ del caso Dead (sumando Super Dead + fraccion Live).",
        "referencia": "CHOC-08 1.3.5.1 / ETABS Base Reactions o Mass Summary",
        "ejemplo_cc135": "CC-135 ~ 1206 kgf (valor piloto).",
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# 2) PROCEDIMIENTO — pasos reproducidos de los MD (no inventado)
#    Cada paso: {n, titulo, capitulo, objetivo, navegacion[], datos[],
#                verificacion, errores[]}
# ─────────────────────────────────────────────────────────────────────────────
PROCEDIMIENTO = {
    "meta": {
        "titulo": "Procedimiento estructural en ETABS — CHOC-08",
        "proyecto": "CC-135, Aulas de Capacitacion Fittacori, Comayagua",
        "normas": "CHOC-08 (sismo) · ACI 318-19/14 (concreto) · AISC 360-16 (acero)",
        "etabs": "v22 · unidades kgf, m, C",
        "fuentes": [
            "manual_procedimiento_etabs_estimastruct.md (15 capitulos)",
            "datos_ingreso_etabs_choc08_estimastruct.md (18 pasos)",
            "flujo_etabs_modelado_choc08.md",
        ],
        "datos_resueltos": "Z=0.25 | Suelo S1 (S=1.0, Ta=0.155 s, Tb=0.364 s, c=2.0) | "
                           "I=1.00 | Rw=8 | Deriva limite=1/200 | f'c=210 kgf/cm2 (3000 psi) | "
                           "fy=4200 (Gr60)",
    },
    "pasos": [
        {
            "n": 1,
            "titulo": "Preparacion del modelo: unidades, grilla y niveles",
            "capitulo": "Cap. 1 · Pasos 1 (datos_ingreso)",
            "objetivo": "Crear el modelo con unidades correctas y la cuadricula/niveles "
                        "reales del proyecto, base de todo lo que sigue.",
            "navegacion": [
                "File > New Model > Model Initialization",
                "Desplegable de unidades (esquina inf. der.) > kgf, m, C",
                "Plantilla = Grid Only (control total, no plantillas automaticas)",
                "Edit > Grid Data > Modify/Show System (afinar ordenadas reales)",
                "Custom Story Data (alturas reales por nivel)",
            ],
            "datos": [
                "Unidades: kgf, m, C (todo el flujo asume estas unidades).",
                "CC-135: ejes X de A a C (luces 1.95-11.7 m), ejes Y 1 a 4.",
                "Niveles CC-135: L1 Nivel de Piso 0.8 m, L2 Nivel Solera 2.8 m, techo.",
            ],
            "verificacion": "La planta muestra los ejes con burbujas (A,B,C / 1,2,3) y la "
                            "elevacion los niveles con sus alturas.",
            "errores": [
                "Dejar unidades en kN o pulgadas.",
                "Olvidar el nivel de cimentacion/base.",
                "Medir alturas de entrepiso a ejes en vez de a caras.",
            ],
        },
        {
            "n": 2,
            "titulo": "Materiales (concreto + acero de refuerzo)",
            "capitulo": "Cap. 2 · Pasos 2-3 (datos_ingreso)",
            "objetivo": "Definir los materiales reales del proyecto para que ETABS calcule "
                        "rigidez, peso propio y diseno de refuerzo.",
            "navegacion": [
                "Define > Material Properties > Add New Material",
                "Region = User · Material Type = Concrete",
                "Add New Material > Material Type = Rebar (acero de refuerzo)",
                "Define > Reinforcing Bar Sizes (tamanos de varilla #2..#8)",
            ],
            "datos": [
                "Concreto 3000psi: f'c=210 kgf/cm2, E=218820 (15100·sqrt(f'c)), "
                "peso 2400 kgf/m3, Poisson 0.2.",
                "Concreto 4000psi: f'c=280, E=252671 (columnas/zapatas exigidas).",
                "Rebar A615 Gr60: fy=4200, fu=6300 kgf/cm2.",
            ],
            "verificacion": "La lista de materiales muestra Concreto 3000psi, "
                            "Concreto 4000psi y A615 Gr60.",
            "errores": [
                "Dejar el E por defecto de ETABS en otras unidades.",
                "No crear el material Rebar (ETABS no podra disenar el acero).",
            ],
        },
        {
            "n": 3,
            "titulo": "Secciones (columnas, vigas, soleras, losas) + refuerzo",
            "capitulo": "Cap. 3 · Pasos 4-8 (datos_ingreso)",
            "objetivo": "Reproducir en ETABS las secciones reales (las mismas con ficha en "
                        "EstimaStruct Div 03) con su material y armado.",
            "navegacion": [
                "Define > Section Properties > Frame Sections > Add New Property",
                "Shape = Concrete Rectangular (Depth t3, Width t2, Material)",
                "[Modify/Show Rebar]: Design Type Column (P-M-M) o Beam, Rebar Material A615 Gr60",
                "Columna: Cover 0.04 m · Viga/solera: Cover top/bottom 0.06 m",
                "Define > Section Properties > Slab Sections (losa maciza Shell-Thin 0.08 m)",
            ],
            "datos": [
                "Columnas: R-2 20x20 4#3 · R-3 20x30 4#4+2#3 · R-5 35x35 · R-6 60x60 8#6 · "
                "R-7 25x25 4#4 · R-9 40x40 6#6.",
                "Vigas: V-1 20x25 2#5+2#3 · V-3 20x30 4#5+2#3 · V-4 25x35 6#5 · V-5 30x40.",
                "Soleras: S-1 35x35 8#6 · S-3 20x20 4#3 (vigas de amarre).",
                "Pedestales P-1..P-7 y losa maciza 8 cm 3000psi.",
            ],
            "verificacion": "Define > Section Properties lista cada seccion con su material "
                            "y dimension.",
            "errores": [
                "Olvidar asignar el material correcto a la seccion.",
                "No entrar a [Modify/Show Rebar] (sin refuerzo el diseno falla).",
                "Confundir Shell-Thin con Membrane.",
            ],
        },
        {
            "n": 4,
            "titulo": "Dibujar y asignar el modelo",
            "capitulo": "Cap. 4",
            "objetivo": "Trazar columnas, vigas, soleras y losas en cada nivel con la "
                        "seccion correcta.",
            "navegacion": [
                "Barra lateral > Draw Beam/Column/Brace · Draw Floor/Wall",
                "Panel Properties of Object > elegir la seccion antes de dibujar",
                "Edit > Replicate (Story) para copiar elementos entre niveles",
                "Set Display Options > Section Names (verificar)",
            ],
            "datos": [
                "Columnas en cruces de ejes; vigas/soleras conectando columnas; "
                "losas sobre los panos.",
            ],
            "verificacion": "Section Names muestra la seccion de cada elemento; la vista 3D "
                            "confirma la geometria.",
            "errores": [
                "Elementos no conectados en los nudos (revisar Analyze > Check Model).",
                "Losas con sentido de carga equivocado.",
                "Vigas a media altura por nivel mal definido.",
            ],
        },
        {
            "n": 5,
            "titulo": "Apoyos, brazos rigidos y diafragma",
            "capitulo": "Cap. 5 · Paso 12 (datos_ingreso)",
            "objetivo": "Empotrar la base, modelar la zona rigida del nudo y definir el "
                        "diafragma de piso.",
            "navegacion": [
                "Select nudos base > Assign > Joint > Restraints > Fixed (empotrado)",
                "Select vigas+columnas > Assign > Frame > End Length Offsets > "
                "Automatic from Connectivity, Rigid-zone factor = 0.75",
                "Define > Diaphragms > Add New Diaphragm > Semi-Rigid (D1)",
                "Select puntos del nivel > Assign > Joint > Diaphragms > D1",
            ],
            "datos": [
                "CC-135: columnas empotradas en la base sobre los pedestales P-x.",
                "Rigid Zone Factor = 0.75 · Diafragma Semi-Rigid.",
            ],
            "verificacion": "Los apoyos muestran el simbolo de empotramiento; el diafragma "
                            "se ve sombreado en Display Options.",
            "errores": [
                "Dejar apoyos articulados cuando el diseno asume empotramiento.",
                "Diafragma rigido en planta irregular (CHOC pide semi-rigido + 3D).",
            ],
        },
        {
            "n": 6,
            "titulo": "Cargas (patrones, valores y asignacion)",
            "capitulo": "Cap. 6 · Pasos 9-11 (datos_ingreso)",
            "objetivo": "Definir Dead/Super Dead/Live, sus valores reales y asignarlos a "
                        "losas y vigas.",
            "navegacion": [
                "Define > Load Patterns: Dead (self weight 1), Super Dead (0), Live (0)",
                "Define > Shell Uniform Load Sets (agrupar Live y Super Dead por nivel)",
                "Assign > Shell Loads > Uniform Load Sets (a losas)",
                "Assign > Frame Loads > Distributed (mamposteria como carga lineal Super Dead)",
            ],
            "datos": [
                "Live por uso: aulas 200 kgf/m2, conferencias 250, asamblea 480. "
                "(CC-135 traia 132 -> insuficiente, corregir.)",
                "Super Dead mamposteria Div 04: bloque 4\" ~150, 6\" ~210, 8\" ~280 kgf/m2.",
                "Mamposteria sobre viga: peso (kgf/m2) x altura = carga lineal kgf/m.",
            ],
            "verificacion": "Display > Load Assignments muestra las flechas de carga con "
                            "sus valores.",
            "errores": [
                "Olvidar el Super Dead de mamposteria (el error de CC-135).",
                "Cargar Live sobre losas que no la reciben.",
                "Mezclar carga de area con lineal.",
            ],
        },
        {
            "n": 7,
            "titulo": "Fuente de masa y P-Delta",
            "capitulo": "Cap. 7 · Paso 13 (datos_ingreso)",
            "objetivo": "Definir la masa que ETABS sacude en el sismo (peso sismico W) y "
                        "activar P-Delta.",
            "navegacion": [
                "Define > Mass Source > Modify/Show > Specified Load Patterns",
                "Agregar Dead (1.0) + Super Dead (1.0) + porcion de Live segun uso",
                "Define > P-Delta Options > Non-iterative - Based on Mass",
            ],
            "datos": [
                "Peso sismico W (CHOC 1.3.5.1) = Dead + Super Dead + porcion de Live.",
                "En bodegas/almacenes incluir >= 25% de la carga viva.",
            ],
            "verificacion": "Tras correr, Display > Tables > Mass Summary muestra la masa "
                            "por nivel. Ese W alimenta la hoja sismica.",
            "errores": [
                "Dejar la masa por defecto (solo peso propio) sin Super Dead/Live.",
                "No activar P-Delta donde gobierna.",
            ],
        },
        {
            "n": 8,
            "titulo": "Accion sismica CHOC-08: funcion de espectro + caso SH",
            "capitulo": "Cap. 8 · Pasos 14-15 (datos_ingreso)",
            "objetivo": "Definir el espectro CHOC-08 (User Defined) y el caso de carga "
                        "espectral SH en ambas direcciones. Corazon del flujo.",
            "navegacion": [
                "Define > Functions > Response Spectrum > Add New Function > User Defined "
                "(NO 'Peru', NO codigo predefinido)",
                "Pegar pares T(s)-a/g de la hoja (boton Exportar espectro ETABS) · Damping 0.05",
                "Nombrar la funcion CHOC08_Z025_S1",
                "Define > Load Cases > Add New Case > Response Spectrum (SH)",
                "Acceleration U1 y U2, funcion CHOC08, Scale Factor g/Rw = 9.81/8 = 1.226",
                "Modal Combination = CQC · Directional = SRSS o 100/30 (0.30) · "
                "Eccentricity Ratio = 0.05",
            ],
            "datos": [
                "Espectro CHOC-08 Z=0.25 S1 (pares T,a/g exportados por esta hoja).",
                "a = 2.5·Z·g·(0.4+0.7·T/Ta) si T<Ta; a = 2.75·Z·g si Ta<=T<=Tb; "
                "a = 2.75·Z·g·(Tb/T)^c si T>Tb.",
            ],
            "verificacion": "El caso SH aparece como Response Spectrum; la grafica muestra la "
                            "meseta plana entre 0.155 y 0.364 s y la rama descendente.",
            "errores": [
                "Usar un espectro de otro pais (Peru).",
                "Olvidar el factor de escala con Rw.",
                "No aplicar en ambas direcciones U1/U2 · dejar excentricidad en 0.",
            ],
        },
        {
            "n": 9,
            "titulo": "Analisis modal (masa participante >= 90%)",
            "capitulo": "Cap. 9 · Paso 16 (datos_ingreso)",
            "objetivo": "Asegurar suficientes modos para capturar >= 90% de masa "
                        "participante en X e Y.",
            "navegacion": [
                "Define > Load Cases > Modal > Modify/Show > Type = Eigen",
                "Maximum Number of Modes: empezar en 12",
                "Post-run: Display > Tables > Modal > Participating Mass Ratios",
            ],
            "datos": [
                "Regla CHOC 1.3.6.5.1: masa participante >= 90% en cada direccion (no "
                "'12 modos' fijo).",
            ],
            "verificacion": "La suma de UX y de UY en el ultimo modo es >= 0.90. Si no, "
                            "subir el numero de modos y recorrer.",
            "errores": [
                "Quedarse en 12 modos sin verificar masa participante.",
            ],
        },
        {
            "n": 10,
            "titulo": "Combinaciones de carga (ACI 318 + sismo)",
            "capitulo": "Cap. 10 · Paso 17 (datos_ingreso)",
            "objetivo": "Crear las combinaciones ACI con el sismo E=SH en ambos sentidos.",
            "navegacion": [
                "Define > Load Combinations > Add New Combo",
                "Agregar las filas Case con su Scale Factor",
            ],
            "datos": [
                "1.4 D",
                "1.2 D + 1.6 L",
                "1.2 D + 1.0 L +/- 1.0 SH",
                "0.9 D +/- 1.0 SH",
                "(D incluye Dead + Super Dead.)",
            ],
            "verificacion": "Las combos sismicas referencian el caso SH existente.",
            "errores": [
                "Combos que referencian un caso sismico inexistente (lo que paso en CC-135).",
                "Olvidar el caso 0.9D +/- E que gobierna el volcamiento.",
            ],
        },
        {
            "n": 11,
            "titulo": "Correr y leer resultados (Base Reactions)",
            "capitulo": "Cap. 11 · Paso 18.1 (datos_ingreso)",
            "objetivo": "Ejecutar el analisis y confirmar que el sismo realmente actua.",
            "navegacion": [
                "Analyze > Check Model (nudos sueltos / duplicados)",
                "Analyze > Run Analysis (F5)",
                "Display > Tables > Analysis > Base Reactions",
                "Display > Force/Stress Diagrams > Frame/Pier Forces (por elemento)",
            ],
            "datos": [
                "Base Reactions: FX y FY != 0 en las combos con SH (el W tambien sale de "
                "FZ del caso Dead).",
            ],
            "verificacion": "Si FX = FY = 0 en las combos sismicas, el espectro no actua: "
                            "regresar al paso 8.",
            "errores": [
                "Combos sismicas con cortante 0 (sismo no aplicado, caso CC-135).",
            ],
        },
        {
            "n": 12,
            "titulo": "Escalar el cortante dinamico (1.3.6.5.3)",
            "capitulo": "Cap. 12 · Paso 18.2 (datos_ingreso)",
            "objetivo": "Garantizar que el cortante dinamico no quede por debajo del minimo "
                        "respecto al estatico.",
            "navegacion": [
                "Leer V_dinamico de Base Reactions (caso SH)",
                "Comparar con V estatico de esta hoja (V = Z·I·C/Rw·W)",
                "Si bajo: Define > Load Cases > SH > Modify > Scale Factor x (0.9·V_est/V_din)",
                "Volver a correr",
            ],
            "datos": [
                "C = 1.25·S/T^(2/3) <= 2.75 · T Metodo A = 0.0731·hn^(3/4).",
                "Regular: dinamico >= 90% del estatico. Irregular: >= 100%.",
            ],
            "verificacion": "Repetir la lectura de Base Reactions hasta cumplir el "
                            "porcentaje.",
            "errores": [
                "No escalar el dinamico cuando queda por debajo del minimo.",
            ],
        },
        {
            "n": 13,
            "titulo": "Verificar derivas (1.3.5.8.2)",
            "capitulo": "Cap. 13 · Paso 18.3 (datos_ingreso)",
            "objetivo": "Comprobar que la deriva de entrepiso no supera el limite.",
            "navegacion": [
                "Display > Story Response > Story Drifts (o Display > Tables > Story Drifts)",
                "Revisar cada piso en X e Y bajo combinaciones sismicas",
            ],
            "datos": [
                "Limite con Rw=8 y T<0.7 s: deriva <= min(0.04/Rw, 0.005) = 0.005 = 1/200.",
                "T>=0.7 s: min(0.03/Rw, 0.004).",
            ],
            "verificacion": "Toda deriva sismica por piso queda por debajo del limite; si "
                            "no, rigidizar (secciones, muros, soleras) y recorrer.",
            "errores": [
                "Leer derivas de combos gravitacionales (no aplican).",
                "Confundir desplazamiento total con deriva de entrepiso.",
            ],
        },
        {
            "n": 14,
            "titulo": "Diseno de concreto y enlace con EstimaStruct",
            "capitulo": "Cap. 14 · Paso 18.4-18.5 (datos_ingreso)",
            "objetivo": "Disenar el concreto en ETABS y llevar fuerzas/secciones a "
                        "EstimaStruct para generar partidas Div 03.",
            "navegacion": [
                "Design > Concrete Frame Design > View/Revise Preferences > "
                "Design Code = ACI 318-19",
                "Design > Concrete Frame Design > Start Design/Check",
                "Clic derecho elemento > Summary (acero requerido + combo que gobierna)",
                "Display > Tables: exportar secciones por miembro y fuerzas por combinacion",
            ],
            "datos": [
                "Exportar Mu, Vu (vigas) y Pu, Mu2, Mu3 (columnas) por combinacion.",
                "En EstimaStruct: mapear seccion -> type_mark (R-5, V-3, S-1), crear "
                "CasoDiseno por combinacion, correr motor ACI, generar partidas Div 03.",
            ],
            "verificacion": "El presupuesto refleja el acero y concreto que el analisis "
                            "realmente exige.",
            "errores": [
                "Disenar con un Design Code distinto al del proyecto.",
            ],
        },
        {
            "n": 15,
            "titulo": "Acero estructural (AISC 360-16, resumen)",
            "capitulo": "Cap. 15",
            "objetivo": "Para elementos de acero (Div 05): mismo modelado hasta correr, con "
                        "diferencias de codigo y metodo.",
            "navegacion": [
                "Design > Steel Frame Design > Preferences > AISC 360-16 (LRFD)",
                "Analysis Options > Metodo de Analisis Directo (reduccion rigidez 0.8, "
                "cargas nocionales 0.002·Yi)",
            ],
            "datos": [
                "Rw del sismo cambia: acero especial 12, comun 6, arriostrado 8.",
                "Estados limite LRFD: compresion §E3, flexion §F2, corte §G2, "
                "interaccion §H1.1 -> D/C ratio.",
            ],
            "verificacion": "EstimaStruct toma perfiles, longitudes, conexiones y peso de "
                            "acero para partidas Div 05.",
            "errores": [
                "Usar Rw de concreto en estructura de acero.",
            ],
        },
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# 3) DOCUMENTACION DEL ARCHIVO DE EXPORT DE ETABS
# ─────────────────────────────────────────────────────────────────────────────
EXPORT_ETABS_DOC = {
    "que_es": "Un archivo (CSV o Excel) que ETABS genera con sus tablas de resultados. "
              "De ahi se leen los valores que cierran la hoja sismica: peso sismico W, "
              "periodo T, cortante dinamico V_din y deriva maxima.",
    "como_producirlo": [
        "En ETABS, tras correr el analisis, abre las tablas: Display > Show Tables "
        "(atajo Ctrl+T).",
        "Marca las tablas relevantes (Base Reactions, Modal Periods And Frequencies, "
        "Story Drifts, opcional Mass Summary by Story) y pulsa OK.",
        "File > Export Current Table > To Excel / To CSV. (O File > Export > "
        "ETABS Tables to Excel.)",
        "Tambien puedes copiar la tabla mostrada y pegarla directamente en el loader.",
    ],
    "tablas": [
        {
            "tabla": "Base Reactions",
            "columnas": "OutputCase, FX, FY, FZ",
            "lee": "W = FZ del caso Dead (sumar Super Dead + fraccion Live). "
                   "V_dinamico = FX/FY del caso SH / Response Spectrum.",
        },
        {
            "tabla": "Modal Periods And Frequencies",
            "columnas": "Mode, Period",
            "lee": "T = periodo del modo gobernante (mayor periodo / mayor masa).",
        },
        {
            "tabla": "Story Drifts",
            "columnas": "Story, OutputCase, Direction, Drift",
            "lee": "Deriva maxima por piso (combos sismicas).",
        },
        {
            "tabla": "Mass Summary by Story (opcional)",
            "columnas": "Story, Mass / UX",
            "lee": "Masa por nivel (verificacion de W).",
        },
    ],
    "como_cargarlo": "Boton 'Cargar export ETABS' > selecciona el .csv/.xlsx (o pega el "
                     "texto de la tabla). El parser es tolerante con los nombres de columna "
                     "(varian poco entre versiones), autollena W, T, V_din y deriva en la "
                     "hoja, recalcula y muestra que leyo y de que tabla. Si una tabla "
                     "falta, avisa y sigue. Nota: el .xlsx se exporta mejor como CSV "
                     "(File > Export Current Table > To CSV).",
}

# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENTACION: Export XLSX para Import Acero (Steel Frame Summary)
# ─────────────────────────────────────────────────────────────────────────────
EXPORT_ETABS_ACERO_DOC = {
    "que_es": "Archivo XLSX con el Steel Frame Design Summary de ETABS. "
              "Contiene perfil, longitud y razon D/C por elemento. "
              "El importador genera partidas Div 05 en EstimaStruct automaticamente.",
    "paso_a_paso": [
        "1. Corre el diseno de acero: Design > Steel Frame Design > Start Design/Check.",
        "2. Abre tablas: Display > Show Tables (Ctrl+T).",
        "3. En el arbol de tablas, marca SOLO estas dos hojas:",
        "   - Design Data > Steel Frame Design > Steel Frame Design Summary",
        "   - Assignments > Frame Assigns - Summary",
        "4. Unidades recomendadas: kN, m (File > Change Tables Units antes de exportar).",
        "5. Exporta: File > Export Tables to Excel (o boton Export en la barra de tablas).",
        "6. Guarda el .xlsx — no renombrar hojas.",
    ],
    "hojas_requeridas": {
        "Steel Frame Design Summary": [
            "Frame (o Member / UniqueName)",
            "DesignSect (o Section / Profile / AnalSect)",
            "PRatio (o DCRatio / Ratio / TotalRatio)",
            "ComboName (o Combo)",
        ],
        "Frame Assigns - Summary": [
            "Frame (o UniqueName)",
            "Length (longitud en metros si unidades = m)",
        ],
    },
    "notas": [
        "El parser hace JOIN entre ambas hojas por ID de frame para combinar D/C + longitud.",
        "Perfil W6X16 (imperial) y W150X24 (metrico) son equivalentes — ambos mapeados.",
        "Elementos sin perfil mapeado en FICHAS_ACERO aparecen como 'no mapeados' (aviso).",
        "El endpoint POST /diseno/{pid}/import-etabs-acero?generar=true importa Y genera partidas en un solo llamado.",
    ],
}


# ─────────────────────────────────────────────────────────────────────────────
# 4) PARSER TOLERANTE DEL EXPORT DE ETABS
# ─────────────────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    """Normaliza un encabezado: minusculas, sin espacios/guiones/parentesis."""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _to_float(s):
    """Convierte texto a float toleando comas de miles y signos. None si no aplica."""
    if s is None:
        return None
    t = str(s).strip().replace(",", "")
    if t == "" or t in ("-", "—"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", t)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _split_rows(texto: str):
    """Divide el texto en filas de celdas. Detecta delimitador (coma o tab)."""
    texto = texto.replace("\r\n", "\n").replace("\r", "\n")
    lineas = [ln for ln in texto.split("\n") if ln.strip() != ""]
    if not lineas:
        return []
    # Detectar delimitador por la linea con mas columnas
    delim = ","
    if "\t" in texto and texto.count("\t") >= texto.count(","):
        delim = "\t"
    rows = []
    reader = csv.reader(io.StringIO("\n".join(lineas)), delimiter=delim)
    for r in reader:
        rows.append([c.strip() for c in r])
    return rows


def _find_header(rows, claves):
    """Busca la fila de encabezado que contenga al menos una de las claves
    normalizadas. Devuelve (idx_fila, {clave_norm: idx_col}) o (None, None)."""
    for i, row in enumerate(rows):
        norm = [_norm(c) for c in row]
        hit = {}
        for j, cell in enumerate(norm):
            for k in claves:
                if cell == k or (k in cell and len(cell) <= len(k) + 6):
                    hit.setdefault(k, j)
        if len(hit) >= max(1, len(claves) // 2):
            return i, hit
    return None, None


def _es_caso_dead(nombre: str) -> bool:
    n = _norm(nombre)
    return ("dead" in n) or n in ("d", "cm", "deadload")


def _es_caso_sismo(nombre: str) -> bool:
    n = _norm(nombre)
    return ("sh" in n) or ("spec" in n) or ("rs" in n) or ("sis" in n) or ("eq" in n) \
        or ("seism" in n) or ("quake" in n)


def parse_export_etabs(texto: str) -> dict:
    """
    Parser tolerante del export de ETABS (CSV/TSV/texto pegado).
    Reconoce, por nombre de columna, las tablas:
      - Base Reactions       -> W (FZ caso Dead), V_din (FX/FY caso sismo)
      - Modal Periods...     -> T (mayor periodo)
      - Story Drifts         -> deriva (max Drift)
    Devuelve {W, T, V_din, deriva, leido[], avisos[]}.
    Valores no encontrados quedan en None.
    """
    out = {"W": None, "T": None, "V_din": None, "deriva": None,
           "derivas_por_piso": [], "leido": [], "avisos": []}
    if not texto or not texto.strip():
        out["avisos"].append("Archivo vacio.")
        return out

    rows = _split_rows(texto)
    if not rows:
        out["avisos"].append("No se pudieron leer filas del archivo.")
        return out

    # ── Base Reactions: columnas OutputCase + FX/FY/FZ ──────────────────────
    hi, hit = _find_header(rows, ["outputcase", "fx", "fy", "fz"])
    if hi is not None and ("fz" in hit or "fx" in hit):
        col_case = hit.get("outputcase")
        col_fx, col_fy, col_fz = hit.get("fx"), hit.get("fy"), hit.get("fz")
        w_val, v_val = None, None
        for row in rows[hi + 1:]:
            if not any(row):
                continue
            case = row[col_case] if (col_case is not None and col_case < len(row)) else ""
            fz = _to_float(row[col_fz]) if (col_fz is not None and col_fz < len(row)) else None
            fx = _to_float(row[col_fx]) if (col_fx is not None and col_fx < len(row)) else None
            fy = _to_float(row[col_fy]) if (col_fy is not None and col_fy < len(row)) else None
            if _es_caso_dead(case) and fz is not None and w_val is None:
                w_val = abs(fz)
            if _es_caso_sismo(case):
                comp = max(abs(fx or 0), abs(fy or 0))
                if comp and (v_val is None or comp > v_val):
                    v_val = comp
        if w_val is not None:
            out["W"] = round(w_val, 1)
            out["leido"].append(f"W = {out['W']} kgf (Base Reactions, FZ caso Dead)")
        if v_val is not None:
            out["V_din"] = round(v_val, 1)
            out["leido"].append(f"V_din = {out['V_din']} kgf (Base Reactions, max FX/FY caso sismo)")
    else:
        out["avisos"].append("No se encontro la tabla 'Base Reactions' (W / V_din).")

    # ── Modal Periods And Frequencies: Mode + Period ────────────────────────
    hi, hit = _find_header(rows, ["mode", "period"])
    if hi is not None and "period" in hit:
        col_per = hit["period"]
        per_max = None
        for row in rows[hi + 1:]:
            if not any(row):
                continue
            per = _to_float(row[col_per]) if col_per < len(row) else None
            if per is not None and (per_max is None or per > per_max):
                per_max = per
        if per_max is not None:
            out["T"] = round(per_max, 4)
            out["leido"].append(f"T = {out['T']} s (Modal, mayor periodo)")
    else:
        out["avisos"].append("No se encontro la tabla 'Modal Periods And Frequencies' (T).")

    # ── Story Drifts: Drift ─────────────────────────────────────────────────
    hi, hit = _find_header(rows, ["story", "drift", "direction"])
    if hi is not None and "drift" in hit:
        col_dr = hit["drift"]
        col_st = hit.get("story")
        col_di = hit.get("direction")
        dr_max = None
        por_piso = []
        for row in rows[hi + 1:]:
            if not any(row):
                continue
            dr = _to_float(row[col_dr]) if col_dr < len(row) else None
            if dr is None:
                continue
            story = (row[col_st].strip() if (col_st is not None and col_st < len(row)) else "") or None
            direction = (row[col_di].strip() if (col_di is not None and col_di < len(row)) else "") or None
            por_piso.append({"story": story, "direction": direction, "drift": round(dr, 6)})
            if dr_max is None or dr > dr_max:
                dr_max = dr
        if por_piso:
            out["derivas_por_piso"] = por_piso
        if dr_max is not None:
            out["deriva"] = round(dr_max, 5)
            out["leido"].append(
                f"deriva max = {out['deriva']} (Story Drifts, {len(por_piso)} filas por piso)")
    else:
        out["avisos"].append("No se encontro la tabla 'Story Drifts' (deriva).")

    if not out["leido"]:
        out["avisos"].append("No se reconocio ninguna tabla. Verifica que el archivo "
                             "tenga encabezados de ETABS (OutputCase/FZ, Period, Drift).")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 5) PARSER DE EXPORT .xlsx (ETABS multi-hoja, via openpyxl)
#    Reusa parse_export_etabs() por hoja: convierte cada hoja a texto CSV y
#    aplica el mismo extractor tolerante. Cada tabla ETABS suele venir en su
#    propia hoja (Base Reactions, Modal..., Story Drifts, Mass Summary).
# ─────────────────────────────────────────────────────────────────────────────
def _hoja_a_texto(ws) -> str:
    """Serializa una hoja openpyxl a texto CSV (filas de celdas separadas por coma)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    for fila in ws.iter_rows(values_only=True):
        if fila is None:
            continue
        w.writerow(["" if c is None else c for c in fila])
    return buf.getvalue()


def parse_export_etabs_bytes(raw: bytes) -> dict:
    """
    Parser del export .xlsx de ETABS (multi-hoja, una tabla por hoja).
    Lee cada hoja con openpyxl, la pasa por parse_export_etabs() y combina los
    valores encontrados (W, T, V_din, deriva). Tolerante con nombres de hoja.
    Devuelve {W, T, V_din, deriva, leido[], avisos[], formato:"xlsx", hojas[]}.
    """
    out = {"W": None, "T": None, "V_din": None, "deriva": None,
           "derivas_por_piso": [], "leido": [], "avisos": [],
           "formato": "xlsx", "hojas": []}
    try:
        from openpyxl import load_workbook
    except ImportError:
        out["avisos"].append("openpyxl no esta instalado en el backend; "
                             "exporta la tabla como CSV.")
        return out

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        out["avisos"].append(f"No se pudo abrir el .xlsx: {e}")
        return out

    for ws in wb.worksheets:
        texto = _hoja_a_texto(ws)
        if not texto.strip():
            continue
        parcial = parse_export_etabs(texto)
        leido_hoja = []
        # Tomar el primer valor no nulo de cada cantidad; anotar de que hoja salio
        for clave in ("W", "T", "V_din", "deriva"):
            if out[clave] is None and parcial.get(clave) is not None:
                out[clave] = parcial[clave]
                leido_hoja.append(clave)
        # La lista de derivas por piso se toma de la primera hoja que la traiga
        if not out["derivas_por_piso"] and parcial.get("derivas_por_piso"):
            out["derivas_por_piso"] = parcial["derivas_por_piso"]
            leido_hoja.append("derivas_por_piso")
        if leido_hoja:
            etiquetas = [x for x in parcial.get("leido", [])]
            out["leido"].extend(f"[{ws.title}] {x}" for x in etiquetas)
            out["hojas"].append({"hoja": ws.title, "leyo": leido_hoja})
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass

    if not out["leido"]:
        out["avisos"].append("No se reconocio ninguna tabla en el .xlsx. Verifica "
                             "que las hojas tengan encabezados de ETABS "
                             "(Base Reactions, Modal Periods, Story Drifts).")
    return out
