from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from collections import Counter, defaultdict
import io, sys, os, re, math, unicodedata
from backend.db import get_db
from backend.models import Presupuesto, Capitulo, Partida, InsumoPartida
from backend.services.pricing import calc_base, precio_unitario as calc_pu

router = APIRouter(tags=["export"])


def safe_fname(nombre: str) -> str:
    """Nombre de archivo seguro para el header Content-Disposition (latin-1/ASCII).
    Las cabeceras HTTP no aceptan caracteres fuera de latin-1 (em-dash, etc.)."""
    base = (nombre or "obra").replace(" ", "_").replace("/", "-")
    base = unicodedata.normalize("NFKD", base).encode("ascii", "ignore").decode("ascii")
    base = "".join(ch for ch in base if ch.isalnum() or ch in "_-.")
    return (base[:40] or "obra")

# Layout de presupuesto — Opción 1 (BaseDatos-style + Type Mark)
MAIN_HEADERS = [
    "CSI", "Type Mark", "Descripción",
    "Cantidad", "Fórmula", "Cantidad Calc.",
    "Unidad", "Mano de Obra", "INSUMOS",
    "PRECIO UNITARIO", "Total"
]
RESUMEN_HEADERS = ["CSI", "Descripción", "Unidad", "Cantidad", "PRECIO UNITARIO", "Total"]
INSUMOS_DETAIL_HEADERS = ["Clave", "Descripción", "Unidad", "Rendimiento", "Costo Unit.", "Total"]
INSUMOS_GLOBAL_HEADERS = ["Tipo", "Clave", "Descripción", "Unidad", "Rendimiento", "Total Agrupado"]
INSUMOS_GLOBAL_FINAL_HEADERS = ["Tipo", "Clave", "Descripción", "Unidad", "Rendimiento", "Total Agrupado", "Rendimiento x Total Agrupado"]

COLOR_TIPO_FILL = {
    "amarillo": "FFF2C94C",
    "verde":    "FF92D050",
    "azul":     "FF00B0F0",
    "rosa":     "FFE91E8C",
    "blanco":   None,
}

# Paleta ConsuConstruct
HDR_FILL = "FFF5C518"   # amarillo casco
HDR_TEXT = "FF1A1A1A"   # negro
DIV_FILL = "FF3A3A3A"   # gris concreto oscuro
DIV_TEXT = "FFF5C518"
SUBT_TEXT = "FFD4A815"
TOT_TEXT = "FFE94560"


def _partida_export_values(pa, sobrecosto: float) -> tuple[float, float, float, float]:
    revit_q = float(pa.revit_q or 0)
    factor_e = float(pa.factor_e or 1) or 1.0
    factor_f = float(pa.factor_f or 1) or 1.0
    cantidad = float(pa.cantidad or 0)
    if cantidad == 0 and revit_q > 0:
        cantidad = revit_q * factor_e * factor_f

    base = calc_base(pa.costo_mo, pa.costo_ma, pa.unitario_matriz)
    precio_unitario = float(pa.precio_unitario or 0)
    if precio_unitario == 0 and base:
        precio_unitario = calc_pu(base, sobrecosto * 100)

    total = float(pa.total or 0)
    if total == 0 and cantidad and precio_unitario:
        total = cantidad * precio_unitario

    return revit_q, cantidad, precio_unitario, total


@router.get("/presupuestos/{pid}/export")
def exportar(pid: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.config),
        joinedload(Presupuesto.capitulos).joinedload(Capitulo.partidas)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    sobrecosto = float(p.config.sobrecosto) / 100 if (p.config and p.config.sobrecosto is not None) else 0.20

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "presupuesto"
    ws_resumen = wb.create_sheet("resumen")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===================== MAIN =====================
    # Título
    ws_main.merge_cells("A1:K1")
    ws_main["A1"] = p.nombre
    ws_main["A1"].font = Font(bold=True, size=14, color=HDR_TEXT)
    ws_main["A1"].fill = PatternFill("solid", fgColor=HDR_FILL)
    ws_main["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_main.row_dimensions[1].height = 24

    ws_main.merge_cells("A2:K2")
    ws_main["A2"] = f"Cliente: {p.cliente or '—'}   |   Moneda: {p.moneda}   |   Sobrecosto: {sobrecosto*100:.1f}%"
    ws_main["A2"].font = Font(size=10, italic=True, color="666666")
    ws_main["A2"].alignment = Alignment(horizontal="center")

    # Headers (fila 4)
    hdr_row = 4
    for ci, txt in enumerate(MAIN_HEADERS, 1):
        cell = ws_main.cell(row=hdr_row, column=ci, value=txt)
        cell.font = Font(bold=True, color=HDR_TEXT, size=10)
        cell.fill = PatternFill("solid", fgColor=HDR_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws_main.row_dimensions[hdr_row].height = 30

    row = hdr_row + 1

    for cap in sorted(p.capitulos, key=lambda c: c.orden):
        # División header
        ws_main.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        hcell = ws_main.cell(row=row, column=1, value=f"{cap.clave} — {cap.nombre}")
        hcell.font = Font(bold=True, color=DIV_TEXT, size=11)
        hcell.fill = PatternFill("solid", fgColor=DIV_FILL)
        hcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_main.row_dimensions[row].height = 20
        row += 1

        cap_first = row

        for pa in sorted(cap.partidas, key=lambda x: x.orden):
            fill_color = COLOR_TIPO_FILL.get(pa.color_tipo or "blanco")
            fill = PatternFill("solid", fgColor=fill_color) if fill_color else None

            r = row  # fila absoluta de esta partida
            revit_q, cantidad, precio_unitario, total = _partida_export_values(pa, sobrecosto)

            # A=CSI  B=Type Mark  C=Descripción
            ws_main.cell(row=r, column=1, value=pa.clave_csi)
            ws_main.cell(row=r, column=2, value=pa.type_mark or "")
            ws_main.cell(row=r, column=3, value=pa.descripcion)

            # D = Cantidad bruta (revit_q)
            ws_main.cell(row=r, column=4, value=revit_q)

            # E: cantidad real de la obra actual.
            ws_main.cell(row=r, column=5, value=cantidad)

            # F = redondeo
            ws_main.cell(row=r, column=6,
                value=f'=IF(E{r}=0,"",IF(E{r}>5,ROUNDUP(E{r}+1,0),ROUNDUP(E{r},0)))')

            # G = Unidad
            ws_main.cell(row=r, column=7, value=pa.unidad or "")

            # H = MO unit, I = INSUMOS (MA + matriz OPUS: subcontrato/flete/equipo) -> H+I == base (bug P1)
            ws_main.cell(row=r, column=8, value=float(pa.costo_mo or 0))
            ws_main.cell(row=r, column=9, value=float(pa.costo_ma or 0) + float(pa.unitario_matriz or 0))

            # J = P.U. = (H + I) * (1 + sobrecosto)
            ws_main.cell(row=r, column=10, value=precio_unitario)

            # K = Total real de la partida
            ws_main.cell(row=r, column=11, value=total)

            # Estilos
            for ci in range(1, 12):
                c = ws_main.cell(row=r, column=ci)
                c.border = border
                if fill:
                    c.fill = fill
                if ci in (4, 5, 6, 8, 9, 10, 11):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '#,##0.00' if ci >= 8 else '#,##0.0000'
                if ci == 1:
                    c.font = Font(size=10, color="666666")
                if ci == 2:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            row += 1

        cap_last = row - 1
        # Subtotal del capítulo
        ws_main.cell(row=row, column=10, value="Subtotal").font = Font(bold=True, color=SUBT_TEXT)
        ws_main.cell(row=row, column=10).alignment = Alignment(horizontal="right")
        if cap_last >= cap_first:
            ws_main.cell(row=row, column=11, value=f"=SUM(K{cap_first}:K{cap_last})")
        else:
            ws_main.cell(row=row, column=11, value=0)
        ws_main.cell(row=row, column=11).font = Font(bold=True, color=SUBT_TEXT)
        ws_main.cell(row=row, column=11).number_format = '#,##0.00'
        row += 2  # blank line

    # Total de obra
    total_row = row + 1
    # Sumar todos los subtotales — más simple: SUM toda la columna K hasta total_row-2 dividido por 2 (porque cada partida y cada subtotal se sumarían). Usar SUMIF a partidas mejor.
    # Práctico: sumar todos los subtotales (que tienen "Subtotal" en col J).
    ws_main.cell(row=total_row, column=10, value="TOTAL OBRA").font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_main.cell(row=total_row, column=10).alignment = Alignment(horizontal="right")
    ws_main.cell(row=total_row, column=11, value=f'=SUMIF(J5:J{total_row-1},"Subtotal",K5:K{total_row-1})')
    ws_main.cell(row=total_row, column=11).font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_main.cell(row=total_row, column=11).number_format = '#,##0.00'

    # Anchos
    widths_main = [13, 11, 50, 11, 11, 13, 9, 12, 12, 13, 14]
    for ci, w in enumerate(widths_main, 1):
        ws_main.column_dimensions[get_column_letter(ci)].width = w

    ws_main.freeze_panes = "A5"

    # ===================== RESUMEN =====================
    ws_resumen.sheet_view.showGridLines = False

    header_rows = [
        ("Consultorías de Construcción S. de R.L.", True, 16, HDR_FILL, 24),
        ("Nosotros lo Construimos", False, 11, DIV_FILL, 20),
        ("Ing. David A. Chinchilla  |  CICH 8222", False, 10, None, 18),
        ("Tegucigalpa, MDC, Honduras", False, 10, None, 18),
        ("+504 9662-8408   |   davidchinchilla@consuconstruct.com", False, 10, None, 18),
        ("https://consuconstruct.com", False, 10, None, 18),
    ]
    for i, (txt, bold, size, fill_color, height) in enumerate(header_rows, 1):
        ws_resumen.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
        c = ws_resumen.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=size, color=HDR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=fill_color) if fill_color else PatternFill(fill_type=None)
        ws_resumen.row_dimensions[i].height = height

    title_row = 8
    ws_resumen.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=7)
    tc = ws_resumen.cell(row=title_row, column=1, value=f"ESTIMACION DE OBRA — {p.nombre}")
    tc.font = Font(bold=True, size=13, color=HDR_TEXT)
    tc.fill = PatternFill("solid", fgColor=HDR_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[title_row].height = 24

    meta_row = title_row + 1
    ws_resumen.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=7)
    meta_cliente = ws_resumen.cell(row=meta_row, column=1, value=f"CLIENTE: {p.cliente or '—'}   |   MONEDA: {p.moneda}")
    meta_cliente.font = Font(bold=True, size=10, color=HDR_TEXT)
    meta_cliente.fill = PatternFill("solid", fgColor="FFF8E58A")
    meta_cliente.alignment = Alignment(horizontal="center", vertical="center")
    meta_cliente.border = border

    meta_row += 1
    ws_resumen.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=7)
    meta_fecha = ws_resumen.cell(row=meta_row, column=1, value=f"OBRA ACTUAL: {p.nombre}  |  FECHA: {p.fecha.strftime('%d/%m/%Y') if p.fecha else '—'}")
    meta_fecha.font = Font(size=10, italic=True, color="666666")
    meta_fecha.alignment = Alignment(horizontal="center", vertical="center")

    # Headers de tabla
    tbl_hdr_row = meta_row + 2
    for ci, txt in enumerate(RESUMEN_HEADERS, 1):
        c = ws_resumen.cell(row=tbl_hdr_row, column=ci, value=txt)
        c.font = Font(bold=True, color=HDR_TEXT, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws_resumen.row_dimensions[tbl_hdr_row].height = 24

    # Filas: una por capítulo + sus partidas
    pr = tbl_hdr_row + 1
    for cap in sorted(p.capitulos, key=lambda c: c.orden):
        cap_rows = []
        for pa in sorted(cap.partidas, key=lambda x: x.orden):
            _, cantidad, precio_unitario, total = _partida_export_values(pa, sobrecosto)
            if cantidad > 0:
                cap_rows.append((pa, cantidad, precio_unitario, total))
        if not cap_rows:
            continue

        ws_resumen.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=6)
        hc = ws_resumen.cell(row=pr, column=1, value=f"{cap.clave} — {cap.nombre}")
        hc.font = Font(bold=True, color=DIV_TEXT, size=11)
        hc.fill = PatternFill("solid", fgColor=DIV_FILL)
        hc.alignment = Alignment(horizontal="left", indent=1)
        pr += 1

        for pa, cantidad, precio_unitario, total in cap_rows:
            ws_resumen.cell(row=pr, column=1, value=pa.clave_csi).font = Font(size=10, color="666666")
            ws_resumen.cell(row=pr, column=2, value=pa.descripcion).font = Font(size=10)
            ws_resumen.cell(row=pr, column=3, value=pa.unidad or "")
            ws_resumen.cell(row=pr, column=4, value=cantidad)
            ws_resumen.cell(row=pr, column=5, value=precio_unitario)
            ws_resumen.cell(row=pr, column=6, value=total)
            for ci in (4, 5, 6):
                cc = ws_resumen.cell(row=pr, column=ci)
                cc.alignment = Alignment(horizontal="right")
                cc.number_format = '#,##0.00'
            for ci in range(1, 7):
                ws_resumen.cell(row=pr, column=ci).border = border
            pr += 1

    # Total final en resumen
    pr += 1
    ws_resumen.cell(row=pr, column=5, value="TOTAL GLOBAL DE LA OBRA").font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_resumen.cell(row=pr, column=5).alignment = Alignment(horizontal="right")
    ws_resumen.cell(row=pr, column=6, value=sum(float(ws_resumen.cell(row=r, column=6).value or 0) for r in range(tbl_hdr_row + 1, pr)))
    ws_resumen.cell(row=pr, column=6).font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_resumen.cell(row=pr, column=6).number_format = '#,##0.00'

    widths_prop = [13, 50, 9, 12, 14, 16]
    for ci, w in enumerate(widths_prop, 1):
        ws_resumen.column_dimensions[get_column_letter(ci)].width = w
    ws_resumen.column_dimensions["F"].width = 2
    ws_resumen.column_dimensions["G"].width = 2

    ws_resumen.freeze_panes = f"A{tbl_hdr_row+1}"

    # ===================== Output =====================
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = safe_fname(p.nombre)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_{nombre}.xlsx"'},
    )


# ───────────────────────────────────────────────────────────────────────────
#  EXPORT BASE DE DATOS COMPLETA — un renglón por (matriz, insumo)
# ───────────────────────────────────────────────────────────────────────────

DB_HEADERS = ["CSI / Recurso", "Type Mark / Clave", "Descripción", "Unidad",
              "Cantidad / Rendimiento", "MO / Costo Unit.", "Insumos / Total"]

MATRIZ_TEXT_COLOR = "FF1F6FB7"   # azul para fila matriz (estilo OPUS)


@router.get("/presupuestos/{pid}/export-db")
def exportar_db(pid: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.capitulos)
        .joinedload(Capitulo.partidas)
        .joinedload(Partida.insumos)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "BD"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DB_HEADERS))
    t = ws.cell(row=1, column=1, value=f"{p.nombre} — Base de Datos Completa")
    t.font = Font(bold=True, size=13, color=HDR_TEXT)
    t.fill = PatternFill("solid", fgColor=HDR_FILL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for ci, txt in enumerate(DB_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=txt)
        c.font = Font(bold=True, color=HDR_TEXT, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 26

    def csi_natural_key(pa):
        s = pa.clave_csi or ""
        parts = []
        for tok in re.findall(r"\d+|\D+", s):
            parts.append((0, int(tok)) if tok.isdigit() else (1, tok.lower()))
        return parts

    row = 3
    for cap in sorted(p.capitulos, key=lambda c: (c.orden if c.orden is not None else 999, c.clave or "")):
        # Header de capítulo
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DB_HEADERS))
        hc = ws.cell(row=row, column=1, value=f"{cap.clave} — {cap.nombre}")
        hc.font = Font(bold=True, color=DIV_TEXT, size=11)
        hc.fill = PatternFill("solid", fgColor=DIV_FILL)
        hc.alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        for pa in sorted(cap.partidas, key=csi_natural_key):
            color_fill = COLOR_TIPO_FILL.get(pa.color_tipo or "blanco")
            fill = PatternFill("solid", fgColor=color_fill) if color_fill else None

            # ── Fila MATRIZ (valores duros: cantidad, MO total, INSUMOS total) ──
            mo_total  = sum(float(i.total or 0) for i in pa.insumos if i.tipo == "MANO_OBRA")
            ins_total = sum(float(i.total or 0) for i in pa.insumos if i.tipo != "MANO_OBRA")
            matriz_vals = [
                pa.clave_csi or "",
                pa.type_mark or "",
                pa.descripcion or "",
                pa.unidad or "",
                float(pa.cantidad or 0),
                mo_total,
                ins_total,
            ]
            for ci, v in enumerate(matriz_vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = border
                if fill:
                    c.fill = fill
                c.font = Font(bold=True, color=MATRIZ_TEXT_COLOR, size=10)
                if ci in (5, 6, 7) and isinstance(v, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '#,##0.0000' if ci == 5 else '#,##0.00'
            row += 1

            # ── Filas RECURSO: insumos no-MO primero, MO después; alfabético por clave ──
            insumos = sorted(
                pa.insumos,
                key=lambda i: (1 if i.tipo == "MANO_OBRA" else 0, (i.clave or "").lower())
            )
            for ins in insumos:
                ins_vals = [
                    "Recurso",
                    ins.clave or "",
                    ins.descripcion or "",
                    ins.unidad or "",
                    float(ins.cantidad or 0),
                    float(ins.costo_unit or 0),
                    float(ins.total or 0),
                ]
                for ci, v in enumerate(ins_vals, 1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.border = border
                    c.font = Font(size=10)
                    if ci == 1:
                        c.font = Font(size=9, color="888888", italic=True)
                    if ci in (5, 6, 7) and isinstance(v, (int, float)):
                        c.alignment = Alignment(horizontal="right", vertical="center")
                        c.number_format = '#,##0.0000' if ci == 5 else '#,##0.00'
                row += 1

            # Fila vacía entre matrices
            row += 1

    # Anchos de columna
    widths = [13, 14, 50, 10, 14, 14, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = safe_fname(p.nombre)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_BD_{nombre}.xlsx"'},
    )


@router.get("/presupuestos/{pid}/export-insumos")
def exportar_insumos_necesarios(pid: str, db: Session = Depends(get_db)):
    """
    Exporta un reporte detallado de insumos necesarios por actividad.

    La salida agrupa materiales y mano de obra por partida con cantidad > 0
    y deja un resumen global consolidado al final del libro.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.config),
        joinedload(Presupuesto.capitulos).joinedload(Capitulo.partidas).joinedload(Partida.insumos)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    def _n(v):
        return float(v or 0)

    def _csi_natural_key(value: str):
        parts = []
        for tok in re.findall(r"\d+|\D+", value or ""):
            parts.append((0, int(tok)) if tok.isdigit() else (1, tok.lower()))
        return parts

    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "detalle"
    ws_global = wb.create_sheet("global")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_title(ws, end_col, title, subtitle=None):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(bold=True, size=14, color=HDR_TEXT)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
            s = ws.cell(row=2, column=1, value=subtitle)
            s.font = Font(size=10, italic=True, color="666666")
            s.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_block_header(ws, row, end_col, text, fill=DIV_FILL, text_color=DIV_TEXT):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color=text_color, size=11)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 19

    def style_table_headers(ws, row, headers, end_col):
        for ci, txt in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=txt)
            c.font = Font(bold=True, color=HDR_TEXT, size=10)
            c.fill = PatternFill("solid", fgColor=HDR_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[row].height = 22

    activity_count = 0
    detail_material_total = 0.0
    detail_mo_total = 0.0
    detail_all_total = 0.0
    global_agg = {}
    omitted_types = set()

    style_title(
        ws_detail,
        6,
        f"{p.nombre} — Insumos necesarios por actividad",
        "Solo actividades con cantidad > 0. Se separan materiales y mano de obra por partida. El total de cada insumo se calcula como rendimiento x costo unitario x cantidad de la matriz."
    )
    detail_row = 4

    for cap in sorted(p.capitulos, key=lambda c: (c.orden if c.orden is not None else 999, c.clave or "")):
        partidas = [pa for pa in sorted(cap.partidas, key=lambda x: (_csi_natural_key(x.clave_csi or ""), x.orden or 0)) if _n(pa.cantidad) > 0]
        if not partidas:
            continue

        style_block_header(ws_detail, detail_row, 6, f"{cap.clave} — {cap.nombre}")
        detail_row += 1

        for pa in partidas:
            activity_count += 1
            actividad_qty = _n(pa.cantidad)
            actividad_unidad = pa.unidad or ""
            activity_material_total = 0.0
            activity_mo_total = 0.0
            activity_total_all = _n(pa.total)
            detail_all_total += activity_total_all

            style_block_header(
                ws_detail,
                detail_row,
                6,
                f"{pa.clave_csi} — {pa.descripcion}",
                fill="FF4A4A4A",
                text_color="FFF5C518"
            )
            detail_row += 1

            # Cantidad de la matriz separada de los insumos.
            ws_detail.cell(row=detail_row, column=1, value="Unidad matriz")
            ws_detail.cell(row=detail_row, column=2, value=actividad_unidad)
            ws_detail.cell(row=detail_row, column=3, value="Cantidad matriz")
            ws_detail.cell(row=detail_row, column=4, value=actividad_qty)
            for ci in range(1, 5):
                cell = ws_detail.cell(row=detail_row, column=ci)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if ci in (1, 3) else "right", vertical="center")
                if ci in (1, 3):
                    cell.font = Font(bold=True, color=HDR_TEXT)
                    cell.fill = PatternFill("solid", fgColor=HDR_FILL)
                else:
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0.0000' if ci == 4 else '@'
            ws_detail.merge_cells(start_row=detail_row, start_column=5, end_row=detail_row, end_column=6)
            note = ws_detail.cell(row=detail_row, column=5, value="Base de cálculo de la ficha matriz")
            note.font = Font(size=9, italic=True, color="666666")
            note.alignment = Alignment(horizontal="center", vertical="center")
            note.border = border
            detail_row += 1

            for ins in pa.insumos:
                rendimiento = _n(ins.cantidad)
                unit = _n(ins.costo_unit)
                total = rendimiento * unit * actividad_qty
                key = (ins.tipo, ins.clave or "", ins.descripcion or "", ins.unidad or "")
                agg = global_agg.setdefault(key, {"tipo": ins.tipo, "rendimiento_sum": 0.0, "count": 0, "total": 0.0})
                agg["rendimiento_sum"] += rendimiento
                agg["count"] += 1
                agg["total"] += total
                if ins.tipo not in ("MATERIAL", "MANO_OBRA"):
                    omitted_types.add(ins.tipo)

            sections = [
                ("Materiales", "MATERIAL", [i for i in pa.insumos if i.tipo == "MATERIAL"]),
                ("Mano de obra", "MANO_OBRA", [i for i in pa.insumos if i.tipo == "MANO_OBRA"]),
            ]

            for section_label, tipo, insumos in sections:
                style_block_header(ws_detail, detail_row, 6, section_label)
                detail_row += 1
                style_table_headers(ws_detail, detail_row, INSUMOS_DETAIL_HEADERS, 6)
                detail_row += 1

                if not insumos:
                    ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=6)
                    empty = ws_detail.cell(row=detail_row, column=1, value=f"Sin {section_label.lower()}")
                    empty.font = Font(size=10, italic=True, color="888888")
                    empty.alignment = Alignment(horizontal="center", vertical="center")
                    empty.border = border
                    detail_row += 1
                    detail_row += 1
                    continue

                section_total = 0.0
                for ins in sorted(insumos, key=lambda i: ((i.clave or "").lower(), (i.descripcion or "").lower())):
                    rendimiento = _n(ins.cantidad)
                    unit = _n(ins.costo_unit)
                    total = rendimiento * unit * actividad_qty
                    row_values = [
                        ins.clave or "",
                        ins.descripcion or "",
                        ins.unidad or "",
                        rendimiento,
                        unit,
                        total,
                    ]
                    for ci, value in enumerate(row_values, 1):
                        cell = ws_detail.cell(row=detail_row, column=ci, value=value)
                        cell.border = border
                        if ci in (4, 5, 6):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0.0000' if ci == 4 else '#,##0.00'
                    section_total += total

                    if tipo == "MATERIAL":
                        detail_material_total += total
                        activity_material_total += total
                    elif tipo == "MANO_OBRA":
                        detail_mo_total += total
                        activity_mo_total += total
                    else:
                        omitted_types.add(tipo)

                    detail_row += 1

                ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=3)
                subtotal_label = ws_detail.cell(row=detail_row, column=1, value=f"Subtotal {section_label.lower()}")
                subtotal_label.font = Font(bold=True, color=SUBT_TEXT)
                subtotal_label.alignment = Alignment(horizontal="right", vertical="center")
                subtotal_label.fill = PatternFill("solid", fgColor="FFF4E7AA")
                subtotal_label.border = border

                subtotal_qty = ws_detail.cell(row=detail_row, column=4, value="")
                subtotal_qty.fill = PatternFill("solid", fgColor="FFF4E7AA")
                subtotal_qty.border = border

                subtotal_unit = ws_detail.cell(row=detail_row, column=5, value="")
                subtotal_unit.fill = PatternFill("solid", fgColor="FFF4E7AA")
                subtotal_unit.border = border

                subtotal_total = ws_detail.cell(row=detail_row, column=6, value=section_total)
                subtotal_total.font = Font(bold=True, color=SUBT_TEXT)
                subtotal_total.alignment = Alignment(horizontal="right", vertical="center")
                subtotal_total.number_format = '#,##0.00'
                subtotal_total.fill = PatternFill("solid", fgColor="FFF4E7AA")
                subtotal_total.border = border
                detail_row += 2

            ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=5)
            activity_total_label = ws_detail.cell(row=detail_row, column=1, value="Total = Total materiales + Total mano de obra")
            activity_total_label.font = Font(bold=True, color=TOT_TEXT)
            activity_total_label.fill = PatternFill("solid", fgColor="FFF4F0C3")
            activity_total_label.alignment = Alignment(horizontal="right", vertical="center")
            activity_total_label.border = border

            activity_total = ws_detail.cell(row=detail_row, column=6, value=activity_material_total + activity_mo_total)
            activity_total.font = Font(bold=True, color=TOT_TEXT)
            activity_total.fill = PatternFill("solid", fgColor="FFF4F0C3")
            activity_total.alignment = Alignment(horizontal="right", vertical="center")
            activity_total.number_format = '#,##0.00'
            activity_total.border = border
            detail_row += 2

            detail_row += 1

    if activity_count == 0:
        ws_detail.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)
        c = ws_detail.cell(row=4, column=1, value="No hay actividades con cantidad > 0 para exportar")
        c.font = Font(italic=True, color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws_detail.column_dimensions["A"].width = 16
    ws_detail.column_dimensions["B"].width = 48
    ws_detail.column_dimensions["C"].width = 12
    ws_detail.column_dimensions["D"].width = 14
    ws_detail.column_dimensions["E"].width = 14
    ws_detail.column_dimensions["F"].width = 14
    ws_detail.freeze_panes = "A4"

    global_items = sorted(
        global_agg.items(),
        key=lambda item: (
            0 if item[1]["tipo"] == "MATERIAL" else 1 if item[1]["tipo"] == "MANO_OBRA" else 2,
            item[0][1].lower(),
            item[0][2].lower(),
            item[0][3].lower(),
        )
    )

    style_title(
        ws_global,
        7,
        f"{p.nombre} — Resumen global de insumos",
        "Lista consolidada de materiales, mano de obra, equipo y subcontrato acumulada desde todas las actividades con cantidad > 0."
    )

    ws_global.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    t1 = ws_global.cell(row=4, column=1, value=f"Materiales: {detail_material_total:,.2f}")
    t1.font = Font(bold=True, color=HDR_TEXT)
    t1.fill = PatternFill("solid", fgColor=HDR_FILL)
    t1.alignment = Alignment(horizontal="center", vertical="center")

    ws_global.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)
    t2 = ws_global.cell(row=4, column=4, value=f"Mano de obra: {detail_mo_total:,.2f}")
    t2.font = Font(bold=True, color=HDR_TEXT)
    t2.fill = PatternFill("solid", fgColor=HDR_FILL)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    ws_global.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
    # Esta fila debe coincidir con el Costo Directo del front end.
    # Es la suma exacta de materiales + mano de obra + equipo + subcontrato.
    t3 = ws_global.cell(row=5, column=1, value=f"Total general: {detail_all_total:,.2f}")
    t3.font = Font(bold=True, color=TOT_TEXT)
    t3.alignment = Alignment(horizontal="center", vertical="center")

    hdr_row = 7
    style_block_header(ws_global, hdr_row, 6, "Consolidado global de insumos")
    hdr_row += 1
    style_table_headers(ws_global, hdr_row, INSUMOS_GLOBAL_HEADERS, 6)

    row = hdr_row + 1
    for key, agg in global_items:
        tipo = agg["tipo"]
        clave = key[1]
        descripcion = key[2]
        unidad = key[3]
        rendimiento = (_n(agg["rendimiento_sum"]) / agg["count"]) if agg["count"] else 0
        total = _n(agg["total"])
        values = [
            "Material" if tipo == "MATERIAL" else "Mano de obra" if tipo == "MANO_OBRA" else tipo,
            clave,
            descripcion,
            unidad,
            rendimiento,
            total,
        ]
        for ci, value in enumerate(values, 1):
            c = ws_global.cell(row=row, column=ci, value=value)
            c.border = border
            if ci in (5, 6):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.0000' if ci == 5 else '#,##0.00'
        row += 1

    if not global_items:
        ws_global.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws_global.cell(row=row, column=1, value="No hay insumos materiales o mano de obra para consolidar.")
        c.font = Font(italic=True, color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")

    row += 2
    style_block_header(ws_global, row, 7, "Cantidad de insumos")
    row += 1
    style_table_headers(ws_global, row, INSUMOS_GLOBAL_FINAL_HEADERS, 7)

    row += 1
    material_items = [item for item in global_items if item[1]["tipo"] == "MATERIAL"]
    for key, agg in material_items:
        tipo = agg["tipo"]
        clave = key[1]
        descripcion = key[2]
        unidad = key[3]
        rendimiento = (_n(agg["rendimiento_sum"]) / agg["count"]) if agg["count"] else 0
        total_agrupado = _n(agg["total"])
        base_resultado = rendimiento * total_agrupado
        resultado = math.ceil(base_resultado) if tipo == "MATERIAL" else base_resultado
        values = [
            "Material" if tipo == "MATERIAL" else "Mano de obra" if tipo == "MANO_OBRA" else tipo,
            clave,
            descripcion,
            unidad,
            rendimiento,
            total_agrupado,
            resultado,
        ]
        for ci, value in enumerate(values, 1):
            c = ws_global.cell(row=row, column=ci, value=value)
            c.border = border
            if ci in (5, 6, 7):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.0000' if ci == 5 else '#,##0'
        row += 1

    if not material_items:
        ws_global.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        c = ws_global.cell(row=row, column=1, value="No hay insumos materiales para consolidar.")
        c.font = Font(italic=True, color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws_global.column_dimensions["A"].width = 16
    ws_global.column_dimensions["B"].width = 18
    ws_global.column_dimensions["C"].width = 48
    ws_global.column_dimensions["D"].width = 12
    ws_global.column_dimensions["E"].width = 14
    ws_global.column_dimensions["F"].width = 14
    ws_global.column_dimensions["G"].width = 14
    ws_global.freeze_panes = "A9"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = safe_fname(p.nombre)
    extra_note = ""
    if omitted_types:
        extra_note = "_solo_ma_mo"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_Insumos_{nombre}{extra_note}.xlsx"'},
    )


@router.get("/presupuestos/{pid}/audit-report")
def exportar_reporte_auditoria(pid: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.config),
        joinedload(Presupuesto.capitulos).joinedload(Capitulo.partidas).joinedload(Partida.insumos)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    def _n(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    def _clean_text(v) -> str:
        return (str(v or "").replace("_x000D_", "").strip())

    def _is_audit_insumo(code: str) -> bool:
        c = (code or "").strip().upper()
        return c.startswith(("MA-", "MO-", "SC-", "EQ-", "HER-", "DIS-", "FL-"))

    version = (p.config.template_version if p.config and p.config.template_version else "v1.2").strip()

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "resumen"
    ws_insumos = wb.create_sheet("insumos")
    ws_codigos = wb.create_sheet("codigos")
    ws_zero = wb.create_sheet("precio_cero")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
    div_fill = PatternFill("solid", fgColor=DIV_FILL)

    def block_title(sheet, row, text, end_col):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        c = sheet.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color=DIV_TEXT, size=11)
        c.fill = div_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def table_header(sheet, row, headers):
        for i, h in enumerate(headers, 1):
            c = sheet.cell(row=row, column=i, value=h)
            c.font = Font(bold=True, color=HDR_TEXT, size=10)
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

    rows = []
    code_groups: dict[str, list[dict]] = defaultdict(list)
    desc_groups: dict[str, Counter] = defaultdict(Counter)
    tipo_groups: dict[str, Counter] = defaultdict(Counter)
    price_groups: dict[str, Counter] = defaultdict(Counter)
    partidas_con_qty = 0

    for cap in sorted(p.capitulos, key=lambda c: c.orden):
        for pa in sorted(cap.partidas, key=lambda x: x.orden):
            qty = _n(pa.cantidad)
            if qty <= 0:
                continue
            partidas_con_qty += 1
            for ins in pa.insumos:
                code = str(ins.clave or "").strip().upper()
                if not _is_audit_insumo(code):
                    continue
                row = {
                    "capitulo": cap.clave,
                    "partida_csi": pa.clave_csi,
                    "partida_desc": pa.descripcion or "",
                    "tipo": ins.tipo or "",
                    "codigo": code,
                    "descripcion": _clean_text(ins.descripcion),
                    "unidad": _clean_text(ins.unidad),
                    "rendimiento": _n(ins.cantidad),
                    "precio": _n(ins.costo_unit),
                    "total": _n(ins.total),
                }
                rows.append(row)
                code_groups[code].append(row)
                if row["descripcion"]:
                    desc_groups[code][row["descripcion"]] += 1
                if row["tipo"]:
                    tipo_groups[code][row["tipo"]] += 1
                if row["precio"] > 0:
                    price_groups[code][round(row["precio"], 4)] += 1

    total_visible = sum(r["total"] for r in rows)
    totales_tipo = defaultdict(float)
    for r in rows:
        totales_tipo[r["tipo"]] += r["total"]

    ws_resumen.merge_cells("A1:B1")
    ws_resumen["A1"] = f"{p.nombre} — Auditoría de Insumos"
    ws_resumen["A1"].font = Font(bold=True, size=14, color=HDR_TEXT)
    ws_resumen["A1"].fill = hdr_fill
    ws_resumen["A1"].alignment = Alignment(horizontal="center", vertical="center")
    resumen_rows = [
        ("Obra", p.nombre),
        ("Versión", version),
        ("Alcance", "Solo insumos visibles en Materiales y Mano de Obra"),
        ("HER-00", "Variable por ficha: 5% de la mano de obra total"),
        ("Partidas con cantidad > 0", partidas_con_qty),
        ("Insumos auditados", len(rows)),
        ("Códigos únicos", len(code_groups)),
        ("Total visible", total_visible),
        ("Estado", "OK" if rows else "Sin datos"),
    ]
    for i, (k, v) in enumerate(resumen_rows, start=3):
        ws_resumen[f"A{i}"] = k
        ws_resumen[f"B{i}"] = v
        ws_resumen[f"A{i}"].border = border
        ws_resumen[f"B{i}"].border = border
        if isinstance(v, (int, float)):
            ws_resumen[f"B{i}"].alignment = Alignment(horizontal="right")
            ws_resumen[f"B{i}"].number_format = '#,##0.00'

    row = 3
    block_title(ws_insumos, row, f"{p.nombre} — Insumos visibles", 10)
    row += 1
    table_header(ws_insumos, row, ["Tipo", "Código", "Descripción", "Unidad", "Rendimiento", "Precio unit.", "Total", "Capítulo", "CSI", "Partida"])
    row += 1
    for r in rows:
        values = [
            r["tipo"],
            r["codigo"],
            r["descripcion"],
            r["unidad"],
            r["rendimiento"],
            r["precio"],
            r["total"],
            r["capitulo"],
            r["partida_csi"],
            r["partida_desc"],
        ]
        for cidx, val in enumerate(values, 1):
            cell = ws_insumos.cell(row=row, column=cidx, value=val)
            cell.border = border
            if cidx in (5, 6, 7):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
        row += 1

    row = 3
    block_title(ws_codigos, row, f"{p.nombre} — Resumen por código", 8)
    row += 1
    table_header(ws_codigos, row, ["Veces", "Código", "Descripción canónica", "Variantes de nombre", "Tipo canónico", "Precio canónico", "Regla", "Total acumulado"])
    row += 1
    for code, items in sorted(code_groups.items(), key=lambda x: (-len(x[1]), x[0])):
        desc_counter = desc_groups.get(code, Counter())
        tipo_counter = tipo_groups.get(code, Counter())
        canon_desc = ""
        if desc_counter:
            max_count = max(desc_counter.values())
            candidates = [desc for desc, count in desc_counter.items() if count == max_count]
            canon_desc = sorted(candidates, key=lambda s: (len(s), s))[0]
        canon_tipo = ""
        if tipo_counter:
            max_count = max(tipo_counter.values())
            candidates = [tipo for tipo, count in tipo_counter.items() if count == max_count]
            canon_tipo = sorted(candidates, key=lambda s: (len(s), s))[0]
        variaciones = " | ".join([f"{desc} ({count})" for desc, count in desc_counter.most_common()])
        canon_price = None
        if code != "HER-00":
            price_counter = price_groups.get(code, Counter())
            if price_counter:
                max_count = max(price_counter.values())
                candidates = [precio for precio, count in price_counter.items() if count == max_count]
                canon_price = sorted(candidates)[0]
            else:
                canon_price = 0.0
        regla = "Variable: 5% de la MO total de la ficha" if code == "HER-00" else "Precio fijo por código"
        total_sum = sum(i["total"] for i in items)
        values = [
            len(items),
            code,
            canon_desc,
            variaciones,
            canon_tipo,
            canon_price,
            regla,
            total_sum,
        ]
        for cidx, val in enumerate(values, 1):
            cell = ws_codigos.cell(row=row, column=cidx, value=val)
            cell.border = border
            if cidx in (1, 6, 8):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
        row += 1

    row = 3
    block_title(ws_zero, row, f"{p.nombre} — Insumos con precio 0", 10)
    row += 1
    table_header(ws_zero, row, ["Tipo", "Código", "Descripción", "Unidad", "Rendimiento", "Precio unit.", "Total", "Capítulo", "CSI", "Partida"])
    row += 1
    for r in rows:
        if r["precio"] > 0:
            continue
        values = [
            r["tipo"],
            r["codigo"],
            r["descripcion"],
            r["unidad"],
            r["rendimiento"],
            r["precio"],
            r["total"],
            r["capitulo"],
            r["partida_csi"],
            r["partida_desc"],
        ]
        for cidx, val in enumerate(values, 1):
            cell = ws_zero.cell(row=row, column=cidx, value=val)
            cell.border = border
            if cidx in (5, 6, 7):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
        row += 1

    for sh, widths in [
        (ws_resumen, [28, 38]),
        (ws_insumos, [16, 16, 46, 12, 14, 14, 14, 12, 16, 42]),
        (ws_codigos, [10, 16, 38, 44, 18, 14, 26, 14]),
        (ws_zero, [16, 16, 46, 12, 14, 14, 14, 12, 16, 42]),
    ]:
        for idx, width in enumerate(widths, 1):
            sh.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = safe_fname(p.nombre)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_Auditoria_{nombre}.xlsx"'},
    )
