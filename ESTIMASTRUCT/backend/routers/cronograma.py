"""
Cronograma / Gantt de una obra.
  GET /presupuestos/{pid}/cronograma          -> JSON para el front (Gantt)
  GET /presupuestos/{pid}/export-cronograma   -> XLSX (tabla + grilla de semanas)

Duraciones por tiempo unitario del catalogo V1.2 (cronograma.py).
"""
import io
import os
import sys
from datetime import date, timedelta
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session, joinedload

from backend.db import get_db
from backend.models import Presupuesto, Capitulo, CronogramaOverride
from backend import cronograma as engine
from backend.routers.export import safe_fname

router = APIRouter(tags=["cronograma"])

CUADRILLAS_MAX = 12   # tope de cuadrillas en paralelo por actividad

# Paleta ConsuConstruct (espejo de export.py)
HDR_FILL = "FFF5C518"
HDR_TEXT = "FF1A1A1A"
DIV_FILL = "FF3A3A3A"
DIV_TEXT = "FFF5C518"
TOT_TEXT = "FFE94560"

# Color por fase (para el grid del Gantt en Excel y referencia del front)
FASE_FILL = {
    "1.": "FF9AA0A6", "2.": "FF9AA0A6", "3.": "FFB0B0B0",
    "4.": "FF378ADD", "5.": "FF1D9E75", "6.": "FFEF6C5A",
    "7.": "FF9C6ADE", "8.": "FF5DCAA5", "9.": "FF5DA8E8", "10": "FFB0B0B0",
}


def _fase_color(fase: str) -> str:
    return FASE_FILL.get((fase or "")[:2], "FF378ADD")


def _overrides(db: Session, pid: str) -> dict:
    """{partida_id: (n_esp, n_ay)} para las partidas que el usuario ajusto."""
    rows = db.query(CronogramaOverride).filter(
        CronogramaOverride.presupuesto_id == pid).all()
    return {r.partida_id: (max(1, int(r.n_esp or engine.DEF_ESP)),
                           max(1, int(r.n_ay or engine.DEF_AY))) for r in rows}


def _partidas_obra(p: Presupuesto, overrides: dict):
    """crono_input para el motor: partidas con cantidad > 0 (+ n_esp/n_ay)."""
    crono_input = []
    for cap in sorted(p.capitulos, key=lambda c: (c.orden if c.orden is not None else 999)):
        for pa in sorted(cap.partidas, key=lambda x: x.orden or 0):
            if float(pa.cantidad or 0) <= 0:
                continue
            csi = (pa.clave_csi or "00").strip()
            ne, na = overrides.get(pa.id, (engine.DEF_ESP, engine.DEF_AY))
            crono_input.append({
                "clave_csi": csi,
                "descripcion": pa.descripcion or "",
                "unidad": pa.unidad or "",
                "cantidad": float(pa.cantidad or 0),
                "capitulo_clave": (cap.clave or csi[:2]).strip(),
                "partida_id": pa.id,
                "n_esp": ne,
                "n_ay": na,
            })
    return crono_input


def _calcular(p: Presupuesto, overrides: dict):
    crono_input = _partidas_obra(p, overrides)
    if not crono_input:
        raise HTTPException(400, "La obra no tiene partidas con cantidad > 0")
    filas = engine.construir_cronograma(crono_input)
    # fin (dia laboral) por actividad
    for f in filas:
        ini = date.fromisoformat(f["fecha_inicio"])
        f["fecha_fin"] = engine._suma_dias_laborales(ini, max(0, f["duracion_dias"] - 1)).isoformat()
    return filas


@router.get("/presupuestos/{pid}/cronograma")
def get_cronograma(pid: str, db: Session = Depends(get_db)):
    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.capitulos).joinedload(Capitulo.partidas)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    filas = _calcular(p, _overrides(db, pid))
    inicio = min(f["fecha_inicio"] for f in filas)
    fin = max(f["fecha_fin"] for f in filas)
    d0 = date.fromisoformat(inicio)
    dfin = date.fromisoformat(fin)
    dias_cal = (dfin - d0).days + 1

    # Días laborables: lunes-viernes completos, sábado medio día, domingo libre.
    dias_lab = 0.0
    d = d0
    while d <= dfin:
        wd = d.weekday()
        if wd <= 4:
            dias_lab += 1.0
        elif wd == 5:
            dias_lab += 0.5
        d += timedelta(days=1)

    fases_orden = []
    for f in filas:
        if f["fase"] not in fases_orden:
            fases_orden.append(f["fase"])

    acts = []
    for f in filas:
        ini = date.fromisoformat(f["fecha_inicio"])
        fn = date.fromisoformat(f["fecha_fin"])
        acts.append({
            "orden": f["orden"],
            "csi": f["clave_csi"],
            "descripcion": f["descripcion"],
            "unidad": f["unidad"],
            "cantidad": f["cantidad"],
            "fase": f["fase"],
            "fase_color": "#" + _fase_color(f["fase"])[2:],
            "fecha_inicio": f["fecha_inicio"],
            "fecha_fin": f["fecha_fin"],
            "duracion_dias": f["duracion_dias"],
            "offset_dias": (ini - d0).days,           # para posicionar la barra (calendario)
            "span_dias": (fn - ini).days + 1,         # ancho de la barra (calendario)
            "fuente": f["fuente"],
            "jh_esp": f.get("jh_esp", 0),
            "jh_ay": f.get("jh_ay", 0),
            "n_esp": f.get("n_esp", engine.DEF_ESP),
            "n_ay": f.get("n_ay", engine.DEF_AY),
            "partida_id": f.get("partida_id"),
        })

    return {
        "obra_id": p.id,
        "nombre": p.nombre,
        "fecha_inicio": inicio,
        "fecha_fin": fin,
        "dias_calendario": dias_cal,
        "dias_laborables": dias_lab,   # L-V completos + sábado ½
        "semanas": (dias_cal + 6) // 7,
        "meses": round(dias_cal / 30.44, 1),
        "fases": fases_orden,
        "actividades": acts,
    }


class PersonalIn(BaseModel):
    partida_id: str
    n_esp: int
    n_ay: int


@router.post("/presupuestos/{pid}/cronograma/personal")
def set_personal(pid: str, body: PersonalIn, db: Session = Depends(get_db)):
    """Ajusta especialistas + ayudantes de una actividad. Default (3,3) elimina el override."""
    p = db.query(Presupuesto).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")
    ne = max(1, min(CUADRILLAS_MAX, int(body.n_esp or engine.DEF_ESP)))
    na = max(1, min(CUADRILLAS_MAX, int(body.n_ay or engine.DEF_AY)))
    ov = db.query(CronogramaOverride).filter(
        CronogramaOverride.partida_id == body.partida_id).first()
    if ne == engine.DEF_ESP and na == engine.DEF_AY:
        if ov:
            db.delete(ov)
    elif ov:
        ov.n_esp = ne
        ov.n_ay = na
    else:
        db.add(CronogramaOverride(presupuesto_id=pid, partida_id=body.partida_id,
                                  n_esp=ne, n_ay=na))
    db.commit()
    return {"ok": True, "partida_id": body.partida_id, "n_esp": ne, "n_ay": na}


@router.get("/presupuestos/{pid}/export-cronograma")
def export_cronograma(pid: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.capitulos).joinedload(Capitulo.partidas)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    filas = _calcular(p, _overrides(db, pid))
    inicio = min(f["fecha_inicio"] for f in filas)
    fin = max(f["fecha_fin"] for f in filas)
    d0 = date.fromisoformat(inicio)
    dfin = date.fromisoformat(fin)
    n_sem = ((dfin - d0).days // 7) + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "cronograma"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    TBL = ["#", "CSI", "Descripción", "Unidad", "Cantidad", "Fase",
           "Inicio", "Días", "Esp", "Ay", "Fin", "Fuente"]
    ncol_tbl = len(TBL)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol_tbl + n_sem)
    t = ws.cell(row=1, column=1, value=f"{p.nombre} — Cronograma (Gantt)")
    t.font = Font(bold=True, size=13, color=HDR_TEXT)
    t.fill = PatternFill("solid", fgColor=HDR_FILL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol_tbl + n_sem)
    sub = ws.cell(row=2, column=1, value=(
        f"Inicio {inicio}  |  Fin {fin}  |  {n_sem} semanas  |  "
        f"{len(filas)} actividades  |  Duraciones por tiempo unitario catálogo V1.2"))
    sub.font = Font(size=10, italic=True, color="666666")
    sub.alignment = Alignment(horizontal="center")

    # Headers tabla + semanas
    hr = 4
    for ci, txt in enumerate(TBL, 1):
        c = ws.cell(row=hr, column=ci, value=txt)
        c.font = Font(bold=True, color=HDR_TEXT, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    for w in range(n_sem):
        c = ws.cell(row=hr, column=ncol_tbl + 1 + w, value=f"S{w + 1}")
        c.font = Font(bold=True, color=HDR_TEXT, size=9)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[hr].height = 24

    row = hr + 1
    fase_actual = None
    for f in filas:
        if f["fase"] != fase_actual:
            fase_actual = f["fase"]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol_tbl + n_sem)
            hc = ws.cell(row=row, column=1, value=fase_actual)
            hc.font = Font(bold=True, color=DIV_TEXT, size=11)
            hc.fill = PatternFill("solid", fgColor=DIV_FILL)
            hc.alignment = Alignment(horizontal="left", indent=1)
            row += 1

        ini = date.fromisoformat(f["fecha_inicio"])
        fn = date.fromisoformat(f["fecha_fin"])
        vals = [f["orden"] + 1, f["clave_csi"], f["descripcion"], f["unidad"],
                f["cantidad"], f["fase"], f["fecha_inicio"], f["duracion_dias"],
                f.get("n_esp", engine.DEF_ESP), f.get("n_ay", engine.DEF_AY),
                f["fecha_fin"], f["fuente"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = border
            c.font = Font(size=9)
            if ci in (5, 8, 9, 10):
                c.alignment = Alignment(horizontal="right")
                c.number_format = '#,##0.00' if ci == 5 else '0'
        # barra de la fase en la grilla de semanas
        wfill = PatternFill("solid", fgColor=_fase_color(f["fase"]))
        sem_ini = (ini - d0).days // 7
        sem_fin = (fn - d0).days // 7
        for w in range(n_sem):
            cell = ws.cell(row=row, column=ncol_tbl + 1 + w)
            cell.border = border
            if sem_ini <= w <= sem_fin:
                cell.fill = wfill
        row += 1

    widths = [5, 14, 46, 8, 10, 30, 11, 6, 5, 5, 11, 9]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for w in range(n_sem):
        ws.column_dimensions[get_column_letter(ncol_tbl + 1 + w)].width = 3.2
    ws.freeze_panes = f"{get_column_letter(ncol_tbl + 1)}{hr + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = safe_fname(p.nombre)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_Cronograma_{nombre}.xlsx"'},
    )
