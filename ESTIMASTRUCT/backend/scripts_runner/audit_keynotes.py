"""Audit CSI keynotes: EVERYTHING in the Revit model that carries a keynote
(placed instances, element TYPES — placed or not —, and MATERIALS) vs the
EstimaStruct V1.2 catalog.

Reads the model dump written by execute_revit_code (model_audit_raw.json), which
now has the shape:
    {
      "keynote_table": { "<CSI key>": "<keynote text>", ... },   # project keynote .txt
      "objects":       [ {kind, id, category, family, type, keynote, type_mark, placed}, ... ]
    }
kind is one of: "type" | "material" | "instance". A keynote almost always lives on
the TYPE (or material), not the placed instance, so the previous instance-only audit
missed loaded-but-unplaced types (e.g. windows) and every material. This version keys
off the object's own keynote regardless of kind or placement.

Classification per keyed object (Director instruction 2026-07-07, option B):
  GREEN = CSI exists in catalog AND the Revit keynote TEXT matches the EstimaStruct
          ficha description (normalized similarity >= TEXT_MATCH_RATIO).
  RED   = CSI not in catalog, or keynote text diverges from the catalog description.

The keynote text comes from the project keynote table (key -> text). Accents in that
.txt are often mojibake (�); normalize_text() strips non-alphanumerics and accents
and compares by difflib ratio so mojibake does not cause false mismatches.

Keynote must be read via BuiltInParameter.KEYNOTE_PARAM (localized display name is
"Nota de clave" on Spanish Revit) — see revit_dump_snippet.py.

Output: audit_keynotes_report.csv (same canonical CSV, now also including
multicapa rows for complex assemblies: compound type + every compound layer).
"""
import json
import os
import re
import csv
import unicodedata
import asyncio
import datetime
from difflib import SequenceMatcher
from collections import defaultdict

MODEL_DUMP = r"D:\OneDrive\Bots\Estimbot\EXPORTS\model_audit_raw.json"
FICHAS_DIR = r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\development\Template2_Updated\v1.2\fichas"
OUT_CSV  = r"D:\OneDrive\Bots\Estimbot\auditorias Revit MCP\audit_keynotes_report.csv"
OUT_XLSX = r"D:\OneDrive\Bots\Estimbot\auditorias Revit MCP\audit_keynotes_report.xlsx"

# Similarity threshold (0..1) above which keynote text == catalog description.
TEXT_MATCH_RATIO = 0.82

# Postgres — primary catalog source (v1.3+)
_PG_SECRET = r"D:\Secrets\postgres_credentials.txt"
_PG_HOST   = "127.0.0.1"
_PG_PORT   = 5432
_PG_DB     = "estimastruct"
_PG_USER   = "postgres"

_SQL_CATALOG = """
    SELECT DISTINCT ON (clave_csi)
        clave_csi, descripcion, unidad, precio_unitario
    FROM partida
    WHERE clave_csi IS NOT NULL AND clave_csi != ''
    ORDER BY clave_csi, precio_unitario DESC NULLS LAST
"""


def normalize_csi_key(key):
    """Same normalization as scripts_runner/import_quantities.py::_normalize_csi_key."""
    if not key:
        return ""
    raw = str(key).replace("_x000D_", "").strip()
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    if lines and all(ln == lines[0] for ln in lines):
        raw = lines[0]
    else:
        raw = " ".join(lines)
    raw = re.sub(r"\s*\.\s*", ".", raw)
    raw = re.sub(r"\s+", " ", raw)
    parts = raw.split(" ")
    if len(parts) <= 3:
        return raw
    return " ".join(parts[:3]) + "." + ".".join(parts[3:])


def normalize_text(s):
    """Lowercase, strip accents + mojibake, keep alphanumerics only. Used for the
    fuzzy keynote-text vs catalog-description comparison."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", u"{}".format(s))
    s = u"".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def text_ratio(a, b):
    na, nb = normalize_text(a), normalize_text(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


def load_codigo_map() -> dict:
    """Returns {normalized_csi: codigo} from fichas JSON. Separate from catalog (PG doesn't have codigo)."""
    fichas_path = os.path.join(os.path.dirname(__file__), "..", "..",
                               "development", "Template2_Updated", "v1.2",
                               "fichas", "fichas_v1.2.live.json")
    try:
        with open(fichas_path, encoding="utf-8") as f:
            fichas = json.load(f)
        return {normalize_csi_key(fi.get("csi") or fi.get("clave_csi") or ""): fi.get("codigo", "")
                for fi in fichas if (fi.get("csi") or fi.get("clave_csi")) and fi.get("codigo")}
    except Exception:
        return {}


def _pg_password():
    try:
        for line in open(_PG_SECRET, encoding="utf-8"):
            line = line.strip()
            if line.startswith("password="):
                return line[9:]
    except Exception:
        pass
    return None


async def _pg_fetch_catalog(pw):
    import asyncpg
    conn = await asyncpg.connect(
        host=_PG_HOST, port=_PG_PORT,
        database=_PG_DB, user=_PG_USER, password=pw
    )
    rows = await conn.fetch(_SQL_CATALOG)
    await conn.close()
    return rows


def load_catalog_pg():
    """Load CSI catalog from Postgres (v1.3 primary source).

    Returns (by_csi, count) where each entry has keys matching ficha dicts
    plus 'unidad', 'precio_unitario', '_source'='pg'.
    Returns (None, 0) if Postgres is unavailable.
    """
    pw = _pg_password()
    if not pw:
        return None, 0
    try:
        rows = asyncio.run(_pg_fetch_catalog(pw))
        by_csi = {}
        for r in rows:
            key = normalize_csi_key(r["clave_csi"])
            if key and key not in by_csi:
                by_csi[key] = {
                    "csi":            r["clave_csi"],
                    "descripcion":    r["descripcion"] or "",
                    "unidad":         r["unidad"] or "",
                    "precio_unitario": float(r["precio_unitario"] or 0),
                    "_source":        "pg",
                }
        return by_csi, len(by_csi)
    except Exception as e:
        print("WARN: PG catalog unavailable — {}".format(e))
        return None, 0


def load_catalog_json():
    """Load CSI catalog from fichas JSON (v1.2 fallback source)."""
    live = os.path.join(FICHAS_DIR, "fichas_v1.2.live.json")
    canon = os.path.join(FICHAS_DIR, "fichas_v1.2.json")
    candidates = [p for p in (live, canon) if os.path.exists(p)]
    path = max(candidates, key=os.path.getmtime)
    with open(path, encoding="utf-8") as f:
        fichas = json.load(f)
    by_csi = {}
    for ficha in fichas:
        if not isinstance(ficha, dict):
            continue
        key = normalize_csi_key(ficha.get("csi", ""))
        if key:
            entry = dict(ficha)
            entry["_source"] = "json"
            entry.setdefault("precio_unitario", 0.0)
            entry.setdefault("unidad", "")
            by_csi[key] = entry
    return by_csi, path, len(fichas)


def load_catalog():
    """Load catalog: Postgres primary, fichas JSON fallback.

    Returns (by_csi, catalog_path, total_count).
    Each entry in by_csi has at least: descripcion, _source ('pg'|'json'),
    precio_unitario, unidad.
    """
    pg_catalog, pg_count = load_catalog_pg()
    json_catalog, json_path, json_count = load_catalog_json()

    if pg_catalog is None:
        # Postgres unavailable — use JSON only
        print("INFO: Catalog source = fichas JSON (PG unavailable)")
        return json_catalog, json_path, json_count

    # Merge: PG primary, JSON fills in CSI codes not yet in PG
    merged = dict(json_catalog)   # start with JSON base
    merged.update(pg_catalog)     # PG overwrites JSON where CSI matches
    # Track which CSI are only in JSON (not promoted to PG yet)
    json_only = set(json_catalog) - set(pg_catalog)
    if json_only:
        print("INFO: {} CSI codes in fichas JSON but NOT in PG (not yet promoted): {}".format(
            len(json_only), ", ".join(sorted(json_only)[:10])
        ))

    print("INFO: Catalog source = PG ({} CSI) + JSON fallback ({} JSON-only)".format(
        pg_count, len(json_only)
    ))
    return merged, "{} + {}".format("postgres:estimastruct", json_path), pg_count + len(json_only)


def classify_keynote(norm_key, keynote_raw, keytext, ficha_desc, ficha):
    if not ficha:
        return (
            "RED",
            u"CSI '{}' no existe en catálogo V1.2".format(keynote_raw or ""),
            0.0,
        )

    ratio = text_ratio(keytext, ficha_desc)
    if not keytext:
        return (
            "RED",
            u"CSI en catálogo pero keynote sin texto en tabla Revit",
            ratio,
        )
    if ratio >= TEXT_MATCH_RATIO:
        return (
            "GREEN",
            u"OK (texto coincide, sim={:.2f})".format(ratio),
            ratio,
        )
    return (
        "RED",
        u"texto keynote != descripción catálogo (sim={:.2f})".format(ratio),
        ratio,
    )


def _write_xlsx(rows_out, fieldnames, stats, catalog_path, catalog_count, kt_count, sched_qty):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARN: openpyxl not installed — skipping XLSX")
        return

    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    BOLD        = Font(bold=True)
    CENTER      = Alignment(horizontal="center", vertical="center")
    WRAP        = Alignment(wrap_text=True, vertical="top")
    THIN        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Resumen ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16

    date_str = datetime.date.today().isoformat()
    ws.append(["Auditoría Keynotes EstimaStruct", date_str])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].alignment = CENTER

    ws.append([])
    ws.append(["Catálogo", str(catalog_path)[:80]])
    ws.append(["Total fichas catálogo", catalog_count])
    ws.append(["Entradas keynote table (Revit)", kt_count])
    ws.append([])
    ws.append(["Filas auditadas", len(rows_out)])
    ws.append(["  Tipos de familia", stats.get("kind_type", 0)])
    ws.append(["  Materiales", stats.get("kind_material", 0)])
    ws.append(["  Instancias", stats.get("kind_instance", 0)])
    ws.append(["  Compound types", stats.get("kind_compound_type", 0)])
    ws.append(["  Compound layers", stats.get("kind_compound_layer", 0)])
    ws.append([])

    green_count = stats.get("GREEN", 0)
    red_count   = stats.get("RED", 0)
    total       = green_count + red_count or 1

    placed_green = sum(1 for r in rows_out if r.get("status") == "GREEN" and r.get("placed") not in (False, "False", "false"))
    placed_red   = sum(1 for r in rows_out if r.get("status") == "RED"   and r.get("placed") not in (False, "False", "false") and r.get("keynote","").strip())
    unplaced_red = sum(1 for r in rows_out if r.get("status") == "RED"   and (r.get("placed") in (False,"False","false") or not r.get("keynote","").strip()))

    ws.append(["GREEN  ✓ (texto coincide)", green_count])
    ws["A{}".format(ws.max_row)].fill = GREEN_FILL
    ws["B{}".format(ws.max_row)].fill = GREEN_FILL
    ws.append(["RED    ✗ (falta/mismatch) — TOTAL", red_count])
    ws["A{}".format(ws.max_row)].fill = RED_FILL
    ws["B{}".format(ws.max_row)].fill = RED_FILL
    ws.append(["  RED placed/con CSI (accionables)", placed_red])
    ws["A{}".format(ws.max_row)].fill = RED_FILL
    ws.append(["  RED unplaced/sin CSI (plantilla Revit, no accionable)", unplaced_red])
    ws.append(["% coincidencia (sobre total)", "{:.1f}%".format(100.0 * green_count / total)])
    ws.append(["% coincidencia (sobre placed)", "{:.1f}%".format(100.0 * placed_green / max(placed_green + placed_red, 1))])
    ws.append([])

    # RED breakdown by reason
    from collections import Counter
    red_reasons = Counter(r.get("reason", "") for r in rows_out if r.get("status") == "RED")
    ws.append(["Motivo RED", "Cantidad"])
    ws["A{}".format(ws.max_row)].font = BOLD
    ws["B{}".format(ws.max_row)].font = BOLD
    for reason, cnt in red_reasons.most_common():
        ws.append([reason[:80], cnt])
        ws["A{}".format(ws.max_row)].fill = RED_FILL

    ws.append([])
    pg_breakdown = Counter(r.get("pg_exists", "") for r in rows_out)
    ws.append(["pg_exists", "Cantidad"])
    ws["A{}".format(ws.max_row)].font = BOLD
    ws["B{}".format(ws.max_row)].font = BOLD
    for k in ("YES", "JSON", "NO"):
        ws.append([k, pg_breakdown.get(k, 0)])

    ws.append([])
    marca_match_count = sum(
        1 for r in rows_out
        if r.get("marca") and r.get("expected_codigo")
        and r["marca"] == r["expected_codigo"]
    )
    ws.append(["Marca match (marca == expected_codigo, non-empty):", marca_match_count])
    ws["A{}".format(ws.max_row)].font = BOLD

    # ── Sheet 2: Auditoría completa ─────────────────────────────────────────
    ws2 = wb.create_sheet("Auditoría")
    DISPLAY_COLS = [
        ("audit_scope", 12), ("kind", 14), ("category", 22), ("family", 28),
        ("type", 36), ("placed", 7), ("keynote", 14), ("type_mark", 10),
        ("marca", 8), ("expected_codigo", 12),
        ("keynote_text", 48), ("ficha_desc", 48), ("status", 8), ("reason", 44),
        ("pg_exists", 8), ("precio_u", 10), ("qty_m2", 10), ("row_role", 8),
        ("layer_function", 16), ("layer_width_mm", 10),
        ("material_name", 28), ("material_keynote", 14),
    ]
    col_keys = [c[0] for c in DISPLAY_COLS]

    for ci, (col, width) in enumerate(DISPLAY_COLS, start=1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws2.column_dimensions[get_column_letter(ci)].width = width

    ws2.row_dimensions[1].height = 18
    ws2.freeze_panes = "A2"

    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

    for ri, row in enumerate(rows_out, start=2):
        fill = GREEN_FILL if row.get("status") == "GREEN" else RED_FILL
        for ci, key in enumerate(col_keys, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            cell = ws2.cell(row=ri, column=ci, value=val)
            # Special fill for marca column: green=match, yellow=mismatch with expected
            if key == "marca":
                marca_val = row.get("marca", "") or ""
                expected_val = row.get("expected_codigo", "") or ""
                if marca_val and expected_val and marca_val == expected_val:
                    cell.fill = GREEN_FILL
                elif expected_val and marca_val != expected_val:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = fill
            else:
                cell.fill = fill
            cell.alignment = WRAP
            if key in ("precio_u", "qty_m2") and val:
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                except (TypeError, ValueError):
                    pass

    ws2.auto_filter.ref = "A1:{}1".format(get_column_letter(len(DISPLAY_COLS)))

    # ── Sheet 3: Solo REDs ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("REDs")
    for ci, (col, width) in enumerate(DISPLAY_COLS, start=1):
        cell = ws3.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws3.column_dimensions[get_column_letter(ci)].width = width
    ws3.row_dimensions[1].height = 18
    ws3.freeze_panes = "A2"

    ri = 2
    for row in rows_out:
        if row.get("status") != "RED":
            continue
        for ci, key in enumerate(col_keys, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.fill = RED_FILL
            cell.alignment = WRAP
            if key in ("precio_u", "qty_m2") and val:
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                except (TypeError, ValueError):
                    pass
        ri += 1
    ws3.auto_filter.ref = "A1:{}1".format(get_column_letter(len(DISPLAY_COLS)))

    # ── Sheet 4: Cantidades (schedule_quantities) ───────────────────────────
    ws4 = wb.create_sheet("Cantidades")
    sq_cols = [
        ("schedule", 28), ("csi", 16), ("mat_name", 36),
        ("area_m2", 12), ("type_mark", 14),
    ]
    for ci, (col, width) in enumerate(sq_cols, start=1):
        cell = ws4.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws4.column_dimensions[get_column_letter(ci)].width = width
    ws4.row_dimensions[1].height = 18
    ws4.freeze_panes = "A2"

    real_rows = [
        r for r in sched_qty
        if r.get("area_m2", 0) > 0 and "Keynote" not in r.get("csi", "")
    ]
    for ri, row in enumerate(real_rows, start=2):
        for ci, (key, _) in enumerate(sq_cols, start=1):
            val = row.get(key, "")
            cell = ws4.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP
            if key == "area_m2":
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                except (TypeError, ValueError):
                    pass

    ws4.auto_filter.ref = "A1:{}1".format(get_column_letter(len(sq_cols)))

    # ── Sheet 5: Sugerencias para EstimaStruct ─────────────────────────────
    # Raw linking data for every Revit element not mapped in EstimaStruct:
    #   - elements with no keynote/CSI assigned (need CSI in Revit)
    #   - elements with a CSI that is absent from the EstimaStruct catalog
    # All identification fields included so each row can be traced back to
    # its Revit element and later written back once a partida exists.
    ws5 = wb.create_sheet("Sugerencias→EstimaStruct")
    ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
    sug_cols = [
        ("element_id",       12),
        ("kind",             14),
        ("category",         22),
        ("family",           28),
        ("type",             36),
        ("marca",             8),
        ("type_mark",        10),
        ("placed",            7),
        ("keynote",          16),
        ("keynote_text",     52),
        ("qty_m2",           12),
        ("pg_exists",         8),
        ("reason",           36),
    ]
    for ci, (col, width) in enumerate(sug_cols, start=1):
        cell = ws5.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws5.column_dimensions[get_column_letter(ci)].width = width
    ws5.row_dimensions[1].height = 18
    ws5.freeze_panes = "A2"

    sug_keys = [c[0] for c in sug_cols]
    sug_rows = [r for r in rows_out if r.get("pg_exists") == "NO"]
    # sort: rows with a keynote first (CSI missing from catalog), then blank keynotes
    sug_rows.sort(key=lambda r: (not bool(r.get("keynote", "").strip()),
                                  r.get("category", ""), r.get("family", ""), r.get("type", "")))
    for ri, row in enumerate(sug_rows, start=2):
        for ci, key in enumerate(sug_keys, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            cell = ws5.cell(row=ri, column=ci, value=val)
            cell.fill = ORANGE_FILL
            cell.alignment = WRAP
            if key == "qty_m2" and val not in ("", None):
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                except (TypeError, ValueError):
                    pass

    ws5.auto_filter.ref = "A1:{}1".format(get_column_letter(len(sug_cols)))

    try:
        wb.save(OUT_XLSX)
    except PermissionError:
        alt = OUT_XLSX.replace(".xlsx", "_new.xlsx")
        try:
            wb.save(alt)
            print("NOTE: XLSX locked — saved to {}".format(alt))
        except PermissionError:
            import datetime as _dt
            ts = _dt.datetime.now().strftime("%H%M%S")
            alt2 = OUT_XLSX.replace(".xlsx", "_{}.xlsx".format(ts))
            wb.save(alt2)
            print("NOTE: both paths locked — saved to {}".format(alt2))
    return True


def audit():
    with open(MODEL_DUMP, encoding="utf-8") as f:
        model = json.load(f)

    keynote_table = model.get("keynote_table", {})
    # normalized keynote table: CSI key -> text
    ktab = {}
    for k, v in keynote_table.items():
        if k == "__error__":
            continue
        ktab[normalize_csi_key(k)] = v

    catalog, catalog_path, catalog_count = load_catalog()
    codigo_map = load_codigo_map()

    # aggregate schedule quantities: normalized_csi -> total area_m2
    qty_by_csi = defaultdict(float)
    for sq in model.get("schedule_quantities", []):
        csi_raw = sq.get("csi", "")
        if "Keynote" in csi_raw:
            continue
        area = sq.get("area_m2", 0) or 0
        if area > 0:
            qty_by_csi[normalize_csi_key(csi_raw)] += area

    rows_out = []
    stats = defaultdict(int)

    for obj in model.get("objects", []):
        keynote_raw = obj.get("keynote")
        norm_key = normalize_csi_key(keynote_raw)
        ficha = catalog.get(norm_key)
        keytext = ktab.get(norm_key, "")
        ficha_desc = ficha.get("descripcion", "") if ficha else ""
        status, reason, ratio = classify_keynote(
            norm_key, keynote_raw, keytext, ficha_desc, ficha
        )

        stats[status] += 1
        stats["kind_" + obj.get("kind", "?")] += 1
        rows_out.append({
            "audit_scope": "base_object",
            "kind": obj.get("kind"),
            "element_id": obj.get("id"),
            "category": obj.get("category"),
            "family": obj.get("family"),
            "type": obj.get("type"),
            "placed": obj.get("placed"),
            "keynote": keynote_raw or "",
            "type_mark": obj.get("type_mark") or "",
            "marca": obj.get("marca", "") or "",
            "expected_codigo": codigo_map.get(normalize_csi_key(keynote_raw), "") if keynote_raw else "",
            "keynote_text": keytext,
            "ficha_desc": ficha_desc,
            "status": status,
            "reason": reason,
            "pg_exists": "YES" if (ficha and ficha.get("_source") == "pg") else ("JSON" if ficha else "NO"),
            "precio_u": ficha.get("precio_unitario", 0) if ficha else 0,
            "qty_m2": qty_by_csi.get(norm_key, 0) or "",
            "parent_type_id": "",
            "parent_type_name": "",
            "row_role": "",
            "layer_index": "",
            "layer_function": "",
            "layer_width_mm": "",
            "material_id": "",
            "material_name": "",
            "material_keynote": "",
        })

    for comp in model.get("compound_elements", []):
        comp_type_id = comp.get("type_id")
        comp_type_name = comp.get("type_name")
        comp_category = comp.get("category_label")
        comp_placed = comp.get("placed")
        comp_type_mark = comp.get("type_type_mark") or ""
        comp_type_keynote = comp.get("type_keynote") or ""
        comp_type_norm = normalize_csi_key(comp_type_keynote)
        comp_type_ficha = catalog.get(comp_type_norm)
        comp_type_text = ktab.get(comp_type_norm, "")
        comp_type_desc = comp_type_ficha.get("descripcion", "") if comp_type_ficha else ""
        comp_status, comp_reason, _ = classify_keynote(
            comp_type_norm,
            comp_type_keynote,
            comp_type_text,
            comp_type_desc,
            comp_type_ficha,
        )

        rows_out.append({
            "audit_scope": "compound",
            "kind": "compound_type",
            "element_id": comp_type_id,
            "category": comp_category,
            "family": "",
            "type": comp_type_name,
            "placed": comp_placed,
            "keynote": comp_type_keynote,
            "type_mark": comp_type_mark,
            "marca": comp.get("type_marca", "") or "",
            "expected_codigo": codigo_map.get(normalize_csi_key(comp_type_keynote), "") if comp_type_keynote else "",
            "keynote_text": comp_type_text,
            "ficha_desc": comp_type_desc,
            "status": comp_status,
            "reason": comp_reason,
            "pg_exists": "YES" if (comp_type_ficha and comp_type_ficha.get("_source") == "pg") else ("JSON" if comp_type_ficha else "NO"),
            "precio_u": comp_type_ficha.get("precio_unitario", 0) if comp_type_ficha else 0,
            "qty_m2": qty_by_csi.get(comp_type_norm, 0) or "",
            "parent_type_id": comp_type_id,
            "parent_type_name": comp_type_name,
            "row_role": "type",
            "layer_index": "",
            "layer_function": "",
            "layer_width_mm": "",
            "material_id": "",
            "material_name": "",
            "material_keynote": "",
        })
        stats[comp_status] += 1
        stats["kind_compound_type"] += 1

        for li, layer in enumerate(comp.get("layers", [])):
            layer_keynote = layer.get("material_keynote") or ""
            layer_norm = normalize_csi_key(layer_keynote)
            layer_ficha = catalog.get(layer_norm)
            layer_text = ktab.get(layer_norm, "")
            layer_desc = layer_ficha.get("descripcion", "") if layer_ficha else ""
            layer_status, layer_reason, _ = classify_keynote(
                layer_norm,
                layer_keynote,
                layer_text,
                layer_desc,
                layer_ficha,
            )
            rows_out.append({
                "audit_scope": "compound",
                "kind": "compound_layer",
                "element_id": "",
                "category": comp_category,
                "family": "",
                "type": comp_type_name,
                "placed": comp_placed,
                "keynote": layer_keynote,
                "type_mark": comp_type_mark,
                "marca": layer.get("marca", "") or "",
                "expected_codigo": codigo_map.get(normalize_csi_key(layer_keynote), "") if layer_keynote else "",
                "keynote_text": layer_text,
                "ficha_desc": layer_desc,
                "status": layer_status,
                "reason": layer_reason,
                "pg_exists": "YES" if (layer_ficha and layer_ficha.get("_source") == "pg") else ("JSON" if layer_ficha else "NO"),
                "precio_u": layer_ficha.get("precio_unitario", 0) if layer_ficha else 0,
                "qty_m2": qty_by_csi.get(layer_norm, 0) or "",
                "parent_type_id": comp_type_id,
                "parent_type_name": comp_type_name,
                "row_role": "layer",
                "layer_index": li,
                "layer_function": layer.get("function", ""),
                "layer_width_mm": layer.get("width_mm", ""),
                "material_id": layer.get("material_id", ""),
                "material_name": layer.get("material_name", ""),
                "material_keynote": layer_keynote,
            })
            stats[layer_status] += 1
            stats["kind_compound_layer"] += 1

    fieldnames = [
        "audit_scope", "kind", "element_id", "category", "family", "type", "placed",
        "keynote", "type_mark", "marca", "expected_codigo",
        "keynote_text", "ficha_desc", "status", "reason",
        "pg_exists", "precio_u", "qty_m2",
        "parent_type_id", "parent_type_name", "row_role", "layer_index",
        "layer_function", "layer_width_mm", "material_id", "material_name",
        "material_keynote",
    ]

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_out)

    schedule_quantities = model.get("schedule_quantities", [])
    mirror_xlsx_written = _write_xlsx(
        rows_out, fieldnames, stats, catalog_path, catalog_count, len(ktab), schedule_quantities
    )

    print("Catálogo: {} ({} fichas)".format(catalog_path, catalog_count))
    print("Tabla keynotes Revit: {} entradas".format(len(ktab)))
    print("Filas auditadas: {}  (tipos {}, materiales {}, instancias {}, compound types {}, compound layers {})".format(
        len(rows_out),
        stats.get("kind_type", 0),
        stats.get("kind_material", 0),
        stats.get("kind_instance", 0),
        stats.get("kind_compound_type", 0),
        stats.get("kind_compound_layer", 0),
    ))
    print("  Compound types: {} | Compound layers: {}".format(
        stats.get("kind_compound_type", 0), stats.get("kind_compound_layer", 0)
    ))
    print("")
    print("GREEN (CSI + texto coincide):  {}".format(stats["GREEN"]))
    print("RED  (falta/mismatch):         {}".format(stats["RED"]))
    print("")
    print("Reporte CSV:  {}".format(OUT_CSV))
    if mirror_xlsx_written:
        print("Reporte XLSX: {}".format(OUT_XLSX))
    else:
        print("Reporte XLSX: omitido por lock; CSV canónico generado y el pipeline puede seguir.")


if __name__ == "__main__":
    audit()
