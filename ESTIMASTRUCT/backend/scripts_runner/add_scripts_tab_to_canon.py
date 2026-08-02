"""Add/refresh 'Scripts & Skills' sheet in Auditoria_MCP_Master_CANON.xlsx.

Run:  python -m backend.scripts_runner.add_scripts_tab_to_canon
      (from D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\)

Adds a sheet with every IronPython MCP snippet, offline Python script,
and Claude skill in the audit pipeline so they can be found from the XLSX.
"""
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CANON_XLSX = r"D:\OneDrive\Bots\Estimbot\auditorias Revit MCP\Auditoria_MCP_Master_CANON.xlsx"
SHEET_NAME = "Scripts & Skills"

# ─── Registry ───────────────────────────────────────────────────────────────
# Columns: nombre, tipo, ruta, que_hace, cuando_usar, inputs, outputs
ENTRIES = [
    # ── IronPython snippets (inyectados via execute_revit_code MCP) ──────────
    (
        "revit_dump_snippet.py (DUMP_AUDIT_CODE)",
        "IronPython MCP",
        r"D:\GitHub\revit-mcp-stdio\revit_mcp\pipe\estimastruct_tools.py",
        "Vuelca modelo Revit completo: ElementTypes, Materials, compound layers, keynote_table, "
        "schedule_quantities (T04/T02/T01). Escribe model_audit_raw.json.",
        "FASE 1: siempre al inicio del pipeline. Antes de auditoría o set marks.",
        "Modelo Revit abierto + MCP execute_revit_code disponible",
        r"D:\OneDrive\Bots\Estimbot\EXPORTS\model_audit_raw.json",
    ),
    (
        "revit_marks_master.py (SET_MARKS_CODE)",
        "IronPython MCP",
        r"D:\GitHub\revit-mcp-stdio\revit_mcp\pipe\estimastruct_tools.py",
        "Asigna TypeMark/Marca desde csi_to_codigo.json. Cubre: Materials (ALL_MODEL_MARK), "
        "ElementTypes no-compuestos (ALL_MODEL_TYPE_MARK), FloorType (TypeMark + rename Name), "
        "Door/Window instancias (inst mark = type TypeMark), Floor instancias (inst mark sync). "
        "NO usa DB.Transaction (execute_revit_code ya envuelve).",
        "FASE 4: después de auditoría, cuando REDs tienen CSI válido pero mark vacío.",
        r"D:\OneDrive\Bots\Estimbot\EXPORTS\csi_to_codigo.json (debe existir antes)",
        "Prints por elemento modificado + stats al final",
    ),
    (
        "revit_set_marks_snippet.py",
        "IronPython MCP (LEGACY)",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\revit_set_marks_snippet.py",
        "Versión anterior de set marks. USA DB.Transaction internamente — "
        "NO inyectar via execute_revit_code (causa 'Starting a new transaction is not permitted'). "
        "Usar revit_marks_master.py en su lugar.",
        "DEPRECADO para inyección MCP. Referencia histórica únicamente.",
        r"D:\OneDrive\Bots\Estimbot\EXPORTS\csi_to_codigo.json",
        "Prints stats",
    ),
    # ── Python offline (corren con D:\LLM\python\python.exe) ────────────────
    (
        "audit_keynotes.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\audit_keynotes.py",
        "Cross-referencia model_audit_raw.json contra catálogo (Postgres primario, fichas JSON fallback). "
        "Clasifica cada objeto: GREEN (CSI existe + texto keynote sim≥0.82) o RED. "
        "Incluye compound types + layers. Genera CSV + XLSX con 5 hojas.",
        "FASE 2: después de dump (Fase 1). También como verificación final tras set marks.",
        r"model_audit_raw.json + Postgres:5432 (o fichas_v1.2.live.json fallback)",
        r"audit_keynotes_report.csv + audit_keynotes_report.xlsx (D:\OneDrive\Bots\Estimbot\auditorias Revit MCP\)",
    ),
    (
        "generate_audit_xlsx.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\generate_audit_xlsx.py",
        "Genera XLSX detallado adicional con 5 sheets: Resumen, Elemento x Elemento, "
        "Materiales-Elem.Compuestos, Familias sin usar, Keynotes Corruptos-Duplicados. "
        "Lee audit_keynotes_report.csv + model_audit_raw.json.",
        "FASE 3 (opcional): para entregable detallado por proyecto.",
        r"audit_keynotes_report.csv + model_audit_raw.json + fichas_v1.2.live.json",
        r"Auditoria_CSI_<proyecto>_<ts>.xlsx (D:\OneDrive\Bots\Estimbot\auditorias Revit MCP\)",
    ),
    (
        "generate_csi_to_codigo.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\generate_csi_to_codigo.py",
        "Genera csi_to_codigo.json mapeando clave_csi → codigo EstimaStruct "
        "(ej. '09 29 00.6' → 'WS-04'). Fuente: fichas_v1.2.live.json.",
        "PREFLIGHT: correr antes de revit_marks_master.py si csi_to_codigo.json no existe o está desactualizado.",
        r"fichas_v1.2.live.json",
        r"D:\OneDrive\Bots\Estimbot\EXPORTS\csi_to_codigo.json",
    ),
    (
        "generate_keynotes.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\generate_keynotes.py",
        "Genera RevitKeynotes*.txt desde catálogo EstimaStruct. "
        "El .txt se carga en Revit (Manage > Keynotes) para poblar la keynote table del proyecto.",
        "Antes de dump/auditoría si la keynote table de Revit está vacía o desactualizada.",
        r"fichas_v1.2.live.json (o Postgres)",
        r"D:\OneDrive\Bots\Estimbot\EXPORTS\RevitKeynotes_*.txt",
    ),
    (
        "suggest_keynotes.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\suggest_keynotes.py",
        "Para elementos RED, sugiere el keynote correcto buscando en catálogo por nombre/familia/categoría.",
        "Después de auditoría cuando hay REDs con CSI sin asignar — para acelerar corrección manual.",
        r"audit_keynotes_report.csv + catálogo",
        "Sugerencias a stdout / CSV",
    ),
    (
        "sync_audit_colors.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\sync_audit_colors.py",
        "Sincroniza colores de auditoría desde XLSX hacia EstimaStruct BD.",
        "Post-auditoría para reflejar estado GREEN/RED en la app.",
        r"audit_keynotes_report.csv / xlsx",
        "EstimaStruct DB actualizada",
    ),
    (
        "run_audit_pipeline.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\run_audit_pipeline.py",
        "Orquestador: corre generate_csi_to_codigo → audit_keynotes → generate_audit_xlsx en secuencia.",
        "Para correr el pipeline offline completo de una vez (asume que el dump ya fue hecho).",
        r"model_audit_raw.json debe existir (dump via MCP primero)",
        "CSV + XLSX de auditoría",
    ),
    (
        "import_quantities.py",
        "Python offline",
        r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\backend\scripts_runner\import_quantities.py",
        "Importa cantidades desde schedules Revit exportados hacia EstimaStruct BD.",
        "Después de exportar schedules desde el botón pyRevit 'Exportar Schedules'.",
        "Excel/CSV de schedules exportados",
        "EstimaStruct BD — partidas con cantidades",
    ),
    # ── Claude Skills ─────────────────────────────────────────────────────────
    (
        "revit-estimastruct-audit (SKILL)",
        "Claude Skill",
        r"C:\Users\consu\.claude\skills\revit-estimastruct-audit\SKILL.md",
        "Skill unificada del pipeline completo: preflight → dump → audit → xlsx → set marks → verify. "
        "Contiene GOTCHAs API, reglas canónicas marks 2026-07-19, tabla reparación archivos faltantes. "
        "Reemplaza: auditoria-csi-revit-mcp, renombrar-keynotes-revit-mcp, revit-project-audit.",
        "Invocar al inicio de cualquier sesión de auditoría o marks Revit vía MCP.",
        "N/A (es documentación/instrucciones para Claude)",
        "N/A",
    ),
    (
        "auditoria-csi-revit-mcp (SKILL DEPRECADA)",
        "Claude Skill (DEPRECADA)",
        r"C:\Users\consu\.claude\skills\auditoria-csi-revit-mcp\SKILL.md",
        "DEPRECADA 2026-07-20. Contenido migrado a revit-estimastruct-audit.",
        "NO usar. Ver revit-estimastruct-audit.",
        "N/A",
        "N/A",
    ),
    (
        "renombrar-keynotes-revit-mcp (SKILL DEPRECADA)",
        "Claude Skill (DEPRECADA)",
        r"C:\Users\consu\.claude\skills\renombrar-keynotes-revit-mcp\SKILL.md",
        "DEPRECADA 2026-07-20. Contenido migrado a revit-estimastruct-audit.",
        "NO usar. Ver revit-estimastruct-audit.",
        "N/A",
        "N/A",
    ),
    (
        "revit-project-audit (SKILL DEPRECADA)",
        "Claude Skill (DEPRECADA)",
        r"C:\Users\consu\.claude\skills\revit-project-audit\SKILL.md",
        "DEPRECADA 2026-07-20. Contenido migrado a revit-estimastruct-audit.",
        "NO usar. Ver revit-estimastruct-audit.",
        "N/A",
        "N/A",
    ),
]

COLUMNS = [
    ("Nombre", 32),
    ("Tipo", 18),
    ("Ruta", 60),
    ("Qué hace", 55),
    ("Cuándo usar", 38),
    ("Inputs", 44),
    ("Outputs", 44),
]


def run():
    if not os.path.exists(CANON_XLSX):
        print("ERROR: No se encuentra", CANON_XLSX)
        return

    wb = openpyxl.load_workbook(CANON_XLSX)

    # Remove existing sheet if present
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME)

    # Styles
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP = Alignment(wrap_text=True, vertical="top")
    THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    TYPE_FILLS = {
        "IronPython MCP":          PatternFill("solid", fgColor="FFE699"),   # gold
        "IronPython MCP (LEGACY)": PatternFill("solid", fgColor="F2DCDB"),   # light red
        "Python offline":          PatternFill("solid", fgColor="DDEBF7"),   # light blue
        "Claude Skill":            PatternFill("solid", fgColor="E2EFDA"),   # light green
        "Claude Skill (DEPRECADA)":PatternFill("solid", fgColor="F2F2F2"),   # grey
    }

    # Header row
    for ci, (col, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    for ri, entry in enumerate(ENTRIES, start=2):
        tipo = entry[1]
        fill = TYPE_FILLS.get(tipo, PatternFill("solid", fgColor="FFFFFF"))
        for ci, val in enumerate(entry, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = THIN
            cell.alignment = WRAP
        ws.row_dimensions[ri].height = 48

    # Legend below data
    legend_row = len(ENTRIES) + 4
    ws.cell(row=legend_row, column=1, value="Leyenda de tipos:").font = Font(bold=True)
    for i, (tipo, fill) in enumerate(TYPE_FILLS.items(), start=1):
        cell = ws.cell(row=legend_row + i, column=1, value=tipo)
        cell.fill = fill
        cell.border = THIN

    ws.auto_filter.ref = "A1:{}1".format(get_column_letter(len(COLUMNS)))

    wb.save(CANON_XLSX)
    print("OK — hoja '{}' actualizada en {}".format(SHEET_NAME, CANON_XLSX))
    print("  {} scripts/skills registrados.".format(len(ENTRIES)))


if __name__ == "__main__":
    run()
