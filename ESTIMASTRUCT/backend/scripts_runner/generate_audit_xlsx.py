"""Generate the auditable XLSX deliverable from the NEW audit_keynotes.py design
(fuzzy keynote-text-vs-catalog-description match, not exact CSI+TypeMark).

Reads:
  - audit_keynotes_report.csv (from audit_keynotes.py — kind/element_id/category/
    family/type/placed/keynote/type_mark/keynote_text/ficha_desc/status/reason)
  - model_audit_raw.json (compound_elements, unplaced_types — see revit_dump_snippet.py)
  - fichas_v1.2.live.json / .json (catalog, mtime-newest)

Writes:
  auditorias Revit MCP/Auditoria_CSI_ValleDeAngeles_<timestamp>.xlsx, 4 sheets:
  1. Resumen                        — dashboard counts
  2. Elemento x Elemento             — one row per keyed object (type/material/instance)
  3. Materiales - Elementos Compuestos — Muros/Suelos/Cubiertas/Techos placed in the
     model, each with its own Type keynote + every layer's material + that material's
     own keynote. This is where wall/floor/roof/ceiling GAPS live (Director instruction
     2026-07-07: only walls belong in a "GAP" view — this sheet covers exactly that,
     material by material, instead of a generic category-level GAP list).
  4. Familias sin usar               — unplaced ElementTypes (physical categories only,
     noise filtered), with their current keynote (blank if none) so it's obvious what
     still needs assignment.

Explicitly DROPPED per Director instruction 2026-07-07: "Ambiguos" sheet (those rows
were just base elements with no keynote yet — not real ambiguity) and the old
category-level "GAP Priorizados" sheet (superseded by the compound-materials sheet,
scoped to walls only).

Color convention (matches EstimaStruct's Partida.color_tipo / bases-drawer.js):
  verde = confirmado correcto (CSI existe + texto keynote coincide con catálogo)
  rosa  = falta keynote / CSI no existe / sin texto en tabla Revit
  azul  = informativo (familia sin usar sin keynote todavía, no es un error)
"""
import json
import os
import re
import csv
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MODEL_DUMP = r"D:\OneDrive\Bots\Estimbot\EXPORTS\model_audit_raw.json"
FICHAS_DIR = r"D:\GitHub\EstimBot\ConsuConstructEstimBot\ESTIMASTRUCT\development\Template2_Updated\v1.2\fichas"
AUDIT_CSV = r"D:\OneDrive\Bots\Estimbot\auditorias Revit MCP\audit_keynotes_report.csv"
OUT_DIR = r"D:\OneDrive\Bots\Estimbot\auditorias Revit MCP"

NOISE_CATEGORY_SUBSTRINGS = [
    "etiqueta", "símbolo", "simbolo", "marca", "leyenda", "material",
    "eje", "boceto", "plano", "cámara", "camara", "anotaci", "circuito",
    "cad de", "curva de nivel", "punto ", "origen interno",
    "información de proyecto", "informacion de proyecto", "referencia a vista",
    "entorno", "emplazamiento", "vínculo", "vinculo", "camino de sol",
    "camino del recorrido", "base de camino",
    "tabla de planificacion", "tabla de planificación", "patron", "patrón",
    "fluido", "vista", "relleno",
]

F_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
F_PINK = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
F_BLUE = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
F_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def is_noise_category(category):
    if not category:
        return True
    low = category.lower()
    return any(sub in low for sub in NOISE_CATEGORY_SUBSTRINGS)


def normalize_csi_key(key):
    if not key:
        return ""
    raw = str(key).replace("_x000D_", "").strip()
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    if lines and all(ln == lines[0] for ln in lines):
        raw = lines[0]
    else:
        raw = " ".join(lines)
    raw = re.sub(r"\s*\.\s*", ".", raw)
    raw = re.sub(r"\s+", " ", raw)
    parts = raw.split(" ")
    if len(parts) <= 3:
        return raw
    return " ".join(parts[:3]) + "." + ".".join(parts[3:])


def load_catalog():
    live = os.path.join(FICHAS_DIR, "fichas_v1.2.live.json")
    canon = os.path.join(FICHAS_DIR, "fichas_v1.2.json")
    candidates = [p for p in (live, canon) if os.path.exists(p)]
    path = max(candidates, key=os.path.getmtime)
    with open(path, encoding="utf-8") as f:
        fichas = json.load(f)
    by_csi = {}
    for ficha in fichas:
        if not isinstance(ficha, dict):
            continue
        key = normalize_csi_key(ficha.get("csi", ""))
        if key:
            by_csi[key] = ficha
    return by_csi, path


def header_row(ws, headers, widths):
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = F_HEADER
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def write_row(ws, ri, vals, fill, wrap_cols=()):
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(ci in wrap_cols))


def build():
    with open(MODEL_DUMP, encoding="utf-8") as f:
        model = json.load(f)
    catalog, catalog_path = load_catalog()

    with open(AUDIT_CSV, encoding="utf-8-sig") as f:
        report_rows = list(csv.DictReader(f))

    green_ct = sum(1 for r in report_rows if r["status"] == "GREEN")
    red_ct = sum(1 for r in report_rows if r["status"] == "RED")
    no_existe_ct = sum(1 for r in report_rows if "no existe" in r["reason"])
    sin_texto_ct = sum(1 for r in report_rows if "sin texto en tabla" in r["reason"])
    diverge_ct = sum(1 for r in report_rows if "texto keynote" in r["reason"])

    # ---- workbook ----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Resumen
    ws = wb.create_sheet("Resumen")
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 15
    rows_summary = [
        ("Auditoría CSI — Apartamento Valle de Angeles", ""),
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Catálogo fuente", catalog_path),
        ("", ""),
        ("Objetos con keynote auditados", len(report_rows)),
        ("VERDE — CSI + texto coincide con catálogo", green_ct),
        ("ROJO total", red_ct),
        ("  CSI no existe en catálogo", no_existe_ct),
        ("  CSI existe pero sin texto en tabla Revit (falta reload keynote.txt)", sin_texto_ct),
        ("  Texto keynote diverge de descripción catálogo", diverge_ct),
        ("", ""),
        ("Elementos compuestos auditados (Muros/Suelos/Cubiertas/Techos)", len(model.get("compound_elements", []))),
        ("Familias sin usar (físicas, excluye ruido)", None),  # filled below
    ]
    for ri, (label, val) in enumerate(rows_summary, 1):
        c1 = ws.cell(row=ri, column=1, value=label)
        c2 = ws.cell(row=ri, column=2, value=val)
        if ri == 1:
            c1.font = Font(bold=True, size=14)
        elif label.strip() and not label.startswith("  "):
            c1.font = Font(bold=True)
    legend_row = len(rows_summary) + 3
    ws.cell(row=legend_row, column=1, value="Leyenda de colores").font = Font(bold=True)
    legend = [
        ("VERDE — confirmado correcto, no tocar", F_GREEN),
        ("ROSA — falta keynote / CSI inválido / sin texto en tabla Revit", F_PINK),
        ("AZUL — informativo (familia sin usar, aún no se asignó keynote)", F_BLUE),
    ]
    for i, (text, fill) in enumerate(legend, 1):
        cell = ws.cell(row=legend_row + i, column=1, value=text)
        cell.fill = fill
        cell.border = BORDER

    # Sheet 2: Elemento x Elemento
    ws2 = wb.create_sheet("Elemento x Elemento")
    headers2 = [
        "Kind", "Element ID", "Categoría", "Familia", "Tipo", "Colocado",
        "Keynote (CSI)", "Type Mark", "Texto en tabla Revit", "Descripción Catálogo",
        "Estado", "Motivo",
    ]
    widths2 = [10, 11, 20, 28, 28, 10, 14, 12, 45, 45, 10, 45]
    header_row(ws2, headers2, widths2)
    report_rows_sorted = sorted(report_rows, key=lambda r: (r["status"] != "RED", r["category"] or ""))
    for ri, r in enumerate(report_rows_sorted, 2):
        fill = F_GREEN if r["status"] == "GREEN" else F_PINK
        vals = [
            r["kind"], r["element_id"], r["category"], r["family"], r["type"],
            r["placed"], r["keynote"], r["type_mark"], r["keynote_text"],
            r["ficha_desc"], r["status"], r["reason"],
        ]
        write_row(ws2, ri, vals, fill, wrap_cols=(9, 10, 12))
    ws2.auto_filter.ref = "A1:L{}".format(len(report_rows_sorted) + 1)

    # Sheet 3: Materiales - Elementos Compuestos (Muros/Suelos/Cubiertas/Techos)
    ws3 = wb.create_sheet("Materiales - Elem. Compuestos")
    headers3 = [
        "Categoría", "Tipo", "Colocado", "Type Mark", "Keynote del Tipo", "Estado Tipo",
        "Capa (función)", "Material", "Ancho (mm)", "Keynote del Material", "Estado Material",
    ]
    widths3 = [14, 32, 10, 12, 16, 10, 18, 26, 12, 18, 30]
    header_row(ws3, headers3, widths3)
    ri = 2
    # Global audit: TODOS los tipos compuestos (colocados o no), ordenados con
    # los colocados primero (más relevantes) y luego por categoría/tipo.
    compound = sorted(
        model.get("compound_elements", []),
        key=lambda c: (not c.get("placed", True), c["category_label"], c["type_name"]),
    )
    for comp in compound:
        placed = comp.get("placed", True)
        type_kn = comp.get("type_keynote")
        type_norm = normalize_csi_key(type_kn)
        type_ficha = catalog.get(type_norm)
        type_ok = bool(type_kn) and type_ficha is not None
        type_fill = F_GREEN if type_ok else (F_PINK if placed else F_BLUE)
        type_estado = "OK" if type_ok else ("SIN KEYNOTE" if not type_kn else "CSI NO EXISTE")

        layers = comp.get("layers", [])
        if not layers:
            vals = [comp["category_label"], comp["type_name"], placed, comp.get("type_type_mark") or "",
                     type_kn or "", type_estado, "", "", "", "", ""]
            write_row(ws3, ri, vals, type_fill, wrap_cols=())
            ri += 1
            continue

        for li, layer in enumerate(layers):
            mat_kn = layer.get("material_keynote")
            mat_norm = normalize_csi_key(mat_kn)
            mat_ficha = catalog.get(mat_norm)
            mat_ok = bool(mat_kn) and mat_ficha is not None
            mat_estado = "OK" if mat_ok else ("SIN KEYNOTE" if not mat_kn else "CSI NO EXISTE")
            if mat_ok or (li == 0 and type_ok):
                fill = F_GREEN
            elif placed:
                fill = F_PINK
            else:
                fill = F_BLUE
            vals = [
                comp["category_label"] if li == 0 else "",
                comp["type_name"] if li == 0 else "",
                placed if li == 0 else "",
                (comp.get("type_type_mark") or "") if li == 0 else "",
                (type_kn or "") if li == 0 else "",
                type_estado if li == 0 else "",
                layer.get("function", ""),
                layer.get("material_name") or "",
                layer.get("width_mm"),
                mat_kn or "",
                mat_estado,
            ]
            write_row(ws3, ri, vals, fill, wrap_cols=())
            ri += 1
    ws3.auto_filter.ref = "A1:K{}".format(ri - 1)

    # Sheet 4: Familias sin usar (físicas, excluye ruido)
    unplaced = [t for t in model.get("unplaced_types", []) if not is_noise_category(t.get("category"))]
    unplaced_sorted = sorted(unplaced, key=lambda t: (t.get("category") or "", t.get("family") or ""))
    ws4 = wb.create_sheet("Familias sin usar")
    headers4 = ["Type ID", "Categoría", "Familia", "Tipo", "Keynote", "Type Mark", "Estado"]
    widths4 = [11, 20, 32, 30, 16, 12, 16]
    header_row(ws4, headers4, widths4)
    for ri, t in enumerate(unplaced_sorted, 2):
        kn = t.get("keynote") or ""
        fill = F_GREEN if kn else F_BLUE
        estado = "con keynote" if kn else "sin keynote (informativo)"
        vals = [t.get("type_id"), t.get("category") or "", t.get("family") or "",
                 t.get("type") or "", kn, t.get("type_mark") or "", estado]
        write_row(ws4, ri, vals, fill, wrap_cols=())
    ws4.auto_filter.ref = "A1:G{}".format(len(unplaced_sorted) + 1)

    # Sheet 5: Keynotes Corruptos/Duplicados — same "Lavatrasta" failure mode
    # detector as before (multi-line repeated keynote in one Revit cell).
    corrupt_rows = []
    for r in report_rows:
        raw = r.get("keynote") or ""
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        if len(lines) > 1:
            if all(ln == lines[0] for ln in lines):
                reason = "keynote repetido {}x en la misma celda".format(len(lines))
            else:
                reason = "keynote con {} líneas distintas en la misma celda".format(len(lines))
            corrupt_rows.append((r["element_id"], r["category"], r["family"], r["type"], raw, reason))
    ws5 = wb.create_sheet("Keynotes Corruptos-Duplicados")
    headers5 = ["Element ID", "Categoría", "Familia", "Tipo", "Keynote crudo (raw)", "Problema detectado"]
    widths5 = [11, 20, 30, 30, 30, 45]
    header_row(ws5, headers5, widths5)
    for ri, row in enumerate(corrupt_rows, 2):
        write_row(ws5, ri, list(row), F_PINK, wrap_cols=(5, 6))
    ws5.auto_filter.ref = "A1:F{}".format(len(corrupt_rows) + 1)

    # fill in the Resumen "familias sin usar" count now that we know it
    ws["B13"] = len(unplaced_sorted)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_DIR, "Auditoria_CSI_ValleDeAngeles_{}.xlsx".format(ts))
    wb.save(out_path)

    print("Objetos auditados:", len(report_rows), "| VERDE:", green_ct, "| ROJO:", red_ct)
    print("  CSI no existe:", no_existe_ct, "| sin texto en tabla:", sin_texto_ct, "| texto diverge:", diverge_ct)
    print("Elementos compuestos:", len(compound))
    print("Familias sin usar (físicas):", len(unplaced_sorted))
    print("Keynotes corruptos/duplicados:", len(corrupt_rows))
    print("")
    print("Archivo:", out_path)
    return out_path


if __name__ == "__main__":
    build()
