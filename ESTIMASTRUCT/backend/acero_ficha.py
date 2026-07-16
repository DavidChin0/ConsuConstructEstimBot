# -*- coding: utf-8 -*-
"""
acero_ficha.py — Puente ETABS -> EstimaStruct (ACERO, Div 05)
=============================================================
Modulo ADITIVO y PURO (sin ORM). MVP: importar ratios D/C y cantidades de la
salida de Steel Frame Design de ETABS. NO disena (el diseno LRFD lo hace
ETABS). Dos responsabilidades:

  1) mapear_perfil_a_ficha(perfil, rol) -> type_mark de ficha Div 05
     (VA-x vigas, C-x columnas). Match por perfil normalizado; perfiles
     duales (sirven como columna o viga) se desambiguan con rol.

  2) Parser tolerante de la tabla de miembros / Steel Frame Design de ETABS:
       - parse_acero_texto(texto)  : CSV/TSV/texto pegado
       - parse_acero_bytes(raw)    : .xlsx multi-hoja (openpyxl)
     Columnas reconocidas (tolerante a la version de ETABS):
       Frame/Member, Section/Profile/AnalSect, Length (m -> mL),
       DesignType/Type (Column/Beam), ratio D/C (PMMRatio/DCRatio/Ratio/Total),
       combo gobernante (Combo/Governing/OutputCase).

Fuente de verdad (etabs-flow/):
  - flujo_etabs_acero_aisc360_div05.md (§6.2 diccionario perfil->ficha,
    §6.1 datos ETABS->EstimaStruct, §6.3 conexiones [fuera de alcance]).
"""

import csv
import io
import json
import os
import re


# ─────────────────────────────────────────────────────────────────────────────
# 1) DICCIONARIO PERFIL AISC -> FICHA Div 05  (§6.2)
#    Clave = perfil normalizado (ver _norm_perfil). Valor:
#      - str           -> ficha unica (no depende del rol)
#      - {"col","viga"} -> perfil dual; se desambigua con rol COLUMNA/VIGA
# ─────────────────────────────────────────────────────────────────────────────

FICHAS_ACERO = {
    # Vigas W (VA-x)
    "W150X18":  "VA-1",
    "W180X22":  "VA-2",
    "W200X27":  "VA-3",
    "W250X33":  "VA-4",
    # Columnas HSS (C-1..C-3)
    "HSS6X6X3/16":  "C-1",
    "HSS8X8X1/4":   "C-2",
    "HSS10X10X1/2": "C-3",
    # Columnas W (C-4, C-5)
    "W200X46":  "C-4",
    "W250X73":  "C-5",
    # Perfiles duales (col / viga) — desambiguar con rol
    "W150X24":  {"col": "C-6",  "viga": "VA-7"},
    "W200X36":  {"col": "C-7",  "viga": "VA-8"},
    "W200X71":  {"col": "C-8",  "viga": "VA-9"},
    "W250X49":  {"col": "C-9",  "viga": "VA-10"},
    "W310X73":  {"col": "C-10", "viga": "VA-11"},
}


def _norm_perfil(perfil: str) -> str:
    """Normaliza un nombre de perfil AISC para hacer match contra el dicc.
    - mayusculas, sin espacios
    - 'x'/'×' -> 'X'
    - elimina sufijos imperiales entre parentesis: 'W310x73 (W12x53)' -> 'W310X73'
    - HSS conserva fracciones (3/16, 1/4, 1/2)
    Ejemplos: 'w 200 x 46' -> 'W200X46' ; 'HSS 6x6x3/16"' -> 'HSS6X6X3/16'
    """
    if not perfil:
        return ""
    s = str(perfil).strip().upper()
    # Quitar lo que venga entre parentesis (equivalencia imperial)
    s = re.sub(r"\([^)]*\)", "", s)
    # Comillas de pulgadas y simbolos varios
    s = s.replace('"', "").replace("”", "").replace("''", "")
    s = s.replace("×", "X")
    # Quitar todos los espacios
    s = re.sub(r"\s+", "", s)
    return s


def _norm_rol(rol: str) -> str:
    """Normaliza el rol a 'COLUMNA' o 'VIGA'. Tolerante a Column/Beam/etc."""
    r = (rol or "").strip().upper()
    if r.startswith("COL") or "COLUM" in r:
        return "COLUMNA"
    if r.startswith("BEAM") or "VIGA" in r or "BRACE" in r:
        return "VIGA"
    return ""


def mapear_perfil_a_ficha(perfil: str, rol: str = "") -> dict:
    """
    Mapea un perfil AISC + rol (COLUMNA/VIGA) a un type_mark de ficha Div 05.

    rol desambigua perfiles duales (W150x24, W200x36, W200x71, W250x49, W310x73).
    Si el perfil es dual y no se da rol -> se asume COLUMNA (C-x) y se marca
    'rol_asumido'.

    Devuelve dict:
      {perfil_norm, type_mark, rol, dual(bool), mapeado(bool), rol_asumido(bool)}
    type_mark == "" y mapeado=False si el perfil no esta en el diccionario.
    """
    pn = _norm_perfil(perfil)
    rol_n = _norm_rol(rol)
    entry = FICHAS_ACERO.get(pn)

    if entry is None:
        return {
            "perfil_norm": pn, "type_mark": "", "rol": rol_n or "",
            "dual": False, "mapeado": False, "rol_asumido": False,
        }

    if isinstance(entry, dict):
        rol_asumido = False
        if rol_n == "VIGA":
            mark = entry["viga"]
        elif rol_n == "COLUMNA":
            mark = entry["col"]
        else:
            mark = entry["col"]      # default columna
            rol_n = "COLUMNA"
            rol_asumido = True
        return {
            "perfil_norm": pn, "type_mark": mark, "rol": rol_n,
            "dual": True, "mapeado": True, "rol_asumido": rol_asumido,
        }

    # Ficha unica: el rol viene del prefijo de la ficha si no se dio
    if not rol_n:
        rol_n = "VIGA" if entry.startswith("VA") else "COLUMNA"
    return {
        "perfil_norm": pn, "type_mark": entry, "rol": rol_n,
        "dual": False, "mapeado": True, "rol_asumido": False,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 2) PARSER TOLERANTE — mismo enfoque que seccion_ficha.py / etabs_procedimiento
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _to_float(s):
    if s is None:
        return None
    t = str(s).strip().replace(",", "")
    if t == "" or t in ("-", "—"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", t)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _split_rows(texto: str):
    texto = (texto or "").replace("\r\n", "\n").replace("\r", "\n")
    lineas = [ln for ln in texto.split("\n") if ln.strip() != ""]
    if not lineas:
        return []
    delim = ","
    if "\t" in texto and texto.count("\t") >= texto.count(","):
        delim = "\t"
    rows = []
    for r in csv.reader(io.StringIO("\n".join(lineas)), delimiter=delim):
        rows.append([c.strip() for c in r])
    return rows


def _find_header(rows, requeridas, opcionales=(), min_hits=2):
    """Busca la MEJOR fila de encabezado: debe contener TODAS las claves
    'requeridas' (normalizadas) y al menos 'min_hits' columnas reconocidas.
    Devuelve (idx_fila, {clave_norm: idx_col}) o (None, None)."""
    todas = list(requeridas) + list(opcionales)
    mejor_i, mejor_hit, mejor_n = None, None, -1
    umbral = max(len(requeridas), min_hits)
    for i, row in enumerate(rows):
        norm = [_norm(c) for c in row]
        hit = {}
        for j, cell in enumerate(norm):
            if not cell:
                continue
            for k in todas:
                if k in hit:
                    continue
                if cell == k or (cell.startswith(k) and len(cell) <= len(k) + 4):
                    hit[k] = j
        if not all(k in hit for k in requeridas):
            continue
        if len(hit) < umbral:
            continue
        if len(hit) > mejor_n:
            mejor_i, mejor_hit, mejor_n = i, hit, len(hit)
    if mejor_i is None:
        return None, None
    return mejor_i, mejor_hit


def _buscar_col(hit, *alias):
    for a in alias:
        if a in hit:
            return hit[a]
    return None


def _to_mL(v) -> float:
    """Longitud de miembro. ETABS exporta en m (flujo 'kgf, m, C'). Si el valor
    es grande (>50) se asume que vino en cm y se divide. Devuelve metros."""
    v = float(v)
    return round(v / 100.0, 3) if v > 50 else round(v, 3)


def _parse_miembros(rows) -> list:
    """Lee la tabla de miembros / Steel Frame Design. Devuelve lista de
    {frame, perfil, rol, longitud_mL, dc, combo}.

    Alias de columnas ETABS -> clave normalizada:
      frame   : frame/member/uniquename
      perfil  : section/profile/analsect/designsect/analysissection
      rol     : designtype/type/frametype
      length  : length/longitud/length1
      dc      : pmmratio/dcratio/ratio/total/totalratio/dc
      combo   : combo/governingcombo/outputcase/loadcase/govcombo
    """
    hi, hit = _find_header(
        rows,
        requeridas=[],
        opcionales=["frame", "member", "uniquename",
                    "section", "profile", "analsect", "designsect",
                    "analysissection",
                    "designtype", "type", "frametype",
                    "length", "longitud", "length1",
                    "pmmratio", "dcratio", "ratio", "total", "totalratio", "dc",
                    "combo", "governingcombo", "outputcase", "loadcase",
                    "govcombo"],
    )
    col_frame = _buscar_col(hit or {}, "frame", "member", "uniquename")
    col_perfil = _buscar_col(hit or {}, "section", "profile", "analsect",
                             "designsect", "analysissection")
    col_rol = _buscar_col(hit or {}, "designtype", "type", "frametype")
    col_len = _buscar_col(hit or {}, "length", "longitud", "length1")
    col_dc = _buscar_col(hit or {}, "pmmratio", "dcratio", "ratio",
                         "totalratio", "total", "dc")
    col_combo = _buscar_col(hit or {}, "combo", "governingcombo", "govcombo",
                            "outputcase", "loadcase")

    # Necesitamos al menos un perfil para mapear; frame es deseable
    if hi is None or col_perfil is None:
        return []

    out = []
    for row in rows[hi + 1:]:
        if not any(row):
            continue
        if any((c or "").upper().startswith("TABLE") for c in row):
            continue

        def cell(idx):
            return row[idx] if (idx is not None and idx < len(row)) else None

        perfil = (cell(col_perfil) or "").strip()
        if not perfil:
            continue
        frame = (cell(col_frame) or "").strip()
        rol = (cell(col_rol) or "").strip()
        longitud = _to_float(cell(col_len))
        dc = _to_float(cell(col_dc))
        combo = (cell(col_combo) or "").strip()
        out.append({
            "frame": frame,
            "perfil": perfil,
            "rol": rol,
            "longitud_mL": _to_mL(longitud) if longitud is not None else None,
            "dc": dc,
            "combo": combo,
        })
    return out


def parse_acero_texto(texto: str) -> dict:
    """Parser de un export ETABS de ACERO en CSV/TSV/texto pegado.
    Devuelve {miembros:[{frame,perfil,rol,longitud_mL,dc,combo}], avisos:[], formato}."""
    out = {"miembros": [], "avisos": [], "formato": "texto"}
    rows = _split_rows(texto)
    if not rows:
        out["avisos"].append("Texto vacio o ilegible.")
        return out
    out["miembros"] = _parse_miembros(rows)
    if not out["miembros"]:
        out["avisos"].append("No se reconocio la tabla de miembros "
                             "(Frame/Section/Length/DesignType/Ratio).")
    return out


def parse_acero_bytes(raw: bytes) -> dict:
    """Parser de export .xlsx ETABS (multi-hoja). Une los miembros de todas las
    hojas. Tolerante a nombres de hoja."""
    out = {"miembros": [], "avisos": [], "formato": "xlsx"}
    try:
        from openpyxl import load_workbook
    except ImportError:
        out["avisos"].append("openpyxl no instalado; exporta como CSV.")
        return out
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        out["avisos"].append(f"No se pudo abrir el .xlsx: {e}")
        return out

    for ws in wb.worksheets:
        buf = io.StringIO()
        w = csv.writer(buf)
        for fila in ws.iter_rows(values_only=True):
            if fila is None:
                continue
            w.writerow(["" if c is None else c for c in fila])
        texto = buf.getvalue()
        if not texto.strip():
            continue
        rows = _split_rows(texto)
        mb = _parse_miembros(rows)
        if mb:
            out["miembros"].extend(mb)
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass

    if not out["miembros"]:
        out["avisos"].append("No se reconocio la tabla de miembros en el .xlsx.")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 3) AGREGACION POR FICHA — acumula longitud, D/C max, combo gobernante
# ─────────────────────────────────────────────────────────────────────────────

DC_LIMITE = 1.0   # AISC 360-16 §H1.1 — D/C ratio admisible


def agregar_por_ficha(miembros: list) -> dict:
    """Agrega una lista de miembros (salida del parser) por ficha Div 05.

    Devuelve dict:
      {por_ficha:[{ficha,perfil,rol,longitud_total_mL,n_miembros,dc_max,
                   combo_gobernante}],
       sobre_esforzados:[{frame,perfil,dc,combo}]  (D/C > DC_LIMITE),
       perfiles_no_mapeados:[{frame,perfil,rol}],
       avisos:[]}
    El combo_gobernante de cada ficha es el del miembro con mayor D/C.
    """
    acum = {}   # ficha -> {...}
    sobre = []
    no_map = []
    avisos = []

    for m in miembros:
        perfil = m.get("perfil") or ""
        rol = m.get("rol") or ""
        mapa = mapear_perfil_a_ficha(perfil, rol)
        dc = m.get("dc")
        frame = m.get("frame") or ""
        combo = m.get("combo") or ""
        longitud = m.get("longitud_mL") or 0.0

        if not mapa["mapeado"]:
            no_map.append({"frame": frame, "perfil": perfil, "rol": rol})
            continue

        if dc is not None and dc > DC_LIMITE:
            sobre.append({"frame": frame, "perfil": perfil,
                          "dc": round(dc, 3), "combo": combo})

        ficha = mapa["type_mark"]
        # Agrupar por (ficha, perfil_norm, rol) para no mezclar perfiles distintos
        clave = (ficha, mapa["perfil_norm"], mapa["rol"])
        a = acum.get(clave)
        if a is None:
            a = {
                "ficha": ficha, "perfil": mapa["perfil_norm"], "rol": mapa["rol"],
                "longitud_total_mL": 0.0, "n_miembros": 0,
                "dc_max": None, "combo_gobernante": "",
            }
            acum[clave] = a
        a["longitud_total_mL"] += longitud
        a["n_miembros"] += 1
        if dc is not None and (a["dc_max"] is None or dc > a["dc_max"]):
            a["dc_max"] = dc
            a["combo_gobernante"] = combo
        if mapa["rol_asumido"]:
            avisos.append(f"Perfil dual '{perfil}' (frame {frame or '-'}) sin "
                          "rol explicito; se asumio COLUMNA.")

    por_ficha = []
    for a in acum.values():
        por_ficha.append({
            "ficha": a["ficha"],
            "perfil": a["perfil"],
            "rol": a["rol"],
            "longitud_total_mL": round(a["longitud_total_mL"], 3),
            "n_miembros": a["n_miembros"],
            "dc_max": round(a["dc_max"], 3) if a["dc_max"] is not None else None,
            "combo_gobernante": a["combo_gobernante"],
        })
    por_ficha.sort(key=lambda x: x["ficha"])

    return {
        "por_ficha": por_ficha,
        "sobre_esforzados": sobre,
        "perfiles_no_mapeados": no_map,
        "avisos": avisos,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 4) FICHAS DE CONEXION (Div 05) + RESOLVER perfil_viga × perfil_columna -> ficha
#    Fuente: plan_modulo_conexiones_acero_aisc_lrfd.md (Capa 4).
#    tipo:  VC_CORTANTE | VC_MOMENTO | VV | SOLDADA | PLACA_BASE
# ─────────────────────────────────────────────────────────────────────────────

FICHAS_CONEXION = {
    # ── Empernadas viga-columna (CV) por clase de peso ────────────────────────
    "CV-1": {"tipo": "VC_CORTANTE", "perfil_col": "",        "perfil_vig": "",        "clase": "ligera",  "descripcion": "Viga-columna a cortante (placa simple) — perfiles ligeros ≤W8x24"},
    "CV-2": {"tipo": "VC_CORTANTE", "perfil_col": "",        "perfil_vig": "",        "clase": "media",   "descripcion": "Viga-columna a cortante (placa simple) — perfiles medios W8x48/W10x33"},
    "CV-3": {"tipo": "VC_CORTANTE", "perfil_col": "",        "perfil_vig": "",        "clase": "pesada",  "descripcion": "Viga-columna a cortante (placa simple) — perfiles pesados W12x53"},
    # ── Empernadas viga-viga (VV) por clase de peso ───────────────────────────
    "VV-1": {"tipo": "VV", "perfil_col": "", "perfil_vig": "", "clase": "ligera", "descripcion": "Viga-viga a cortante (placa simple) — perfiles ligeros ≤W8x24"},
    "VV-2": {"tipo": "VV", "perfil_col": "", "perfil_vig": "", "clase": "media",  "descripcion": "Viga-viga a cortante (placa simple) — perfiles medios W8x48/W10x33"},
    "VV-3": {"tipo": "VV", "perfil_col": "", "perfil_vig": "", "clase": "pesada", "descripcion": "Viga-viga a cortante (placa simple) — perfiles pesados W12x53"},
    # ── Soldadas mismo perfil ambos lados (CX-15..18) ─────────────────────────
    "CX-15": {"tipo": "SOLDADA", "perfil_col": "W200X36", "perfil_vig": "W200X36", "clase": "", "descripcion": "Conexion soldada (filete E70) — W200x36 ambos lados"},
    "CX-16": {"tipo": "SOLDADA", "perfil_col": "W200X71", "perfil_vig": "W200X71", "clase": "", "descripcion": "Conexion soldada (filete E70) — W200x71 ambos lados"},
    "CX-17": {"tipo": "SOLDADA", "perfil_col": "W250X49", "perfil_vig": "W250X49", "clase": "", "descripcion": "Conexion soldada (filete E70) — W250x49 ambos lados"},
    "CX-18": {"tipo": "SOLDADA", "perfil_col": "W310X73", "perfil_vig": "W310X73", "clase": "", "descripcion": "Conexion soldada (filete E70) — W310x73 ambos lados"},
    # ── Soldadas mixtas col≠vig (CX-19/20) ────────────────────────────────────
    "CX-19": {"tipo": "SOLDADA", "perfil_col": "W200X71", "perfil_vig": "W200X36", "clase": "", "descripcion": "Conexion soldada mixta (filete E70) — col W200x71 × vig W200x36"},
    "CX-20": {"tipo": "SOLDADA", "perfil_col": "W250X49", "perfil_vig": "W200X71", "clase": "", "descripcion": "Conexion soldada mixta (filete E70) — col W250x49 × vig W200x71"},
    # ── Placa base §J8 (BP-x) por clase de peso de la COLUMNA ──────────────────
    #   Insumos en la base curada (placa A36 + J-bolts F1554 + soldadura + labor).
    "BP-1": {"tipo": "PLACA_BASE", "perfil_col": "", "perfil_vig": "", "clase": "ligera",  "descripcion": "Placa base §J8 (A36 250x250x5/8 + 4 J-bolt 3/4 F1554) — columnas ligeras ≤W8x24/HSS6"},
    "BP-2": {"tipo": "PLACA_BASE", "perfil_col": "", "perfil_vig": "", "clase": "media",   "descripcion": "Placa base §J8 (A36 300x300x3/4 + 4 J-bolt 1 F1554) — columnas medias W8x48/W10x33/W200x46-71/HSS8"},
    "BP-3": {"tipo": "PLACA_BASE", "perfil_col": "", "perfil_vig": "", "clase": "pesada",  "descripcion": "Placa base §J8 (A36 400x400x1 + 4 J-bolt 1 F1554) — columnas pesadas W12x53/W250x73/W310x73/HSS10"},
}


def _peso_perfil(perfil_norm: str):
    """Peso lineal (kg/m) de la designacion. W250X33 -> 33 ; HSS no aplica -> None.
    Para clasificar por peso en el resolver empernado."""
    m = re.search(r"^W\d+X(\d+(?:\.\d+)?)", perfil_norm or "")
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


# Clase de peso EXPLICITA por perfil (alinea con las fichas base CV/VV-1/2/3).
# Resuelve la ambiguedad US (W8x48=48) vs metrico (W200x71=71): ambos -> media.
#   ligera = CV-1/VV-1 (≤W8x24) · media = CV-2/VV-2 (W8x48,W10x33) · pesada = CV-3/VV-3 (W12x53)
_PERFIL_CLASE = {
    # ligera
    "W6X16": "ligera", "W8X24": "ligera",
    "W150X18": "ligera", "W150X24": "ligera", "W180X22": "ligera",
    "W200X27": "ligera", "W200X36": "ligera", "W250X33": "ligera",
    # media
    "W8X48": "media", "W10X33": "media",
    "W200X46": "media", "W200X71": "media", "W250X49": "media",
    # pesada
    "W12X53": "pesada", "W250X73": "pesada", "W310X73": "pesada",
}


def _clase_peso(perfil_norm: str) -> str:
    """Clase de peso de la viga para resolver CV/VV. Mapa explicito (alineado con
    las fichas base); fallback por peso si el perfil no esta listado.
    Fallback metrico: ligera ≤40 · media 40–72 · pesada >72  (HSS/otros -> media)."""
    if perfil_norm in _PERFIL_CLASE:
        return _PERFIL_CLASE[perfil_norm]
    w = _peso_perfil(perfil_norm)
    if w is None:
        return "media"
    if w <= 40:
        return "ligera"
    if w <= 72:
        return "media"
    return "pesada"


# Indices auxiliares para el resolver (perfil_norm -> ficha)
_CX_MISMO = {f["perfil_vig"]: k for k, f in FICHAS_CONEXION.items()
             if f["tipo"] == "SOLDADA" and f["perfil_col"] == f["perfil_vig"]}
_CX_MIXTO = {(f["perfil_col"], f["perfil_vig"]): k for k, f in FICHAS_CONEXION.items()
             if f["tipo"] == "SOLDADA" and f["perfil_col"] != f["perfil_vig"]}
_CLASE_A_CV = {"ligera": "CV-1", "media": "CV-2", "pesada": "CV-3"}
_CLASE_A_VV = {"ligera": "VV-1", "media": "VV-2", "pesada": "VV-3"}
_CLASE_A_BP = {"ligera": "BP-1", "media": "BP-2", "pesada": "BP-3"}


def _norm_tipo_conexion(tipo: str) -> str:
    """Normaliza el tipo de conexion a una clave canonica."""
    t = (tipo or "").strip().upper().replace("-", "_").replace(" ", "_")
    if "SOLD" in t:
        return "SOLDADA"
    if "PLACA" in t or "BASE" in t:
        return "PLACA_BASE"
    if t.startswith("VV") or "VIGA_VIGA" in t:
        return "VV"
    if "MOMENTO" in t:
        return "VC_MOMENTO"
    if "CORTANTE" in t or t.startswith("VC") or "VIGA_COLUMNA" in t:
        return "VC_CORTANTE"
    return "VC_CORTANTE"


# ─────────────────────────────────────────────────────────────────────────────
# CARGA DE INSUMOS DESDE LA BASE CURADA (fichas_{version}.json) — fuente unica.
#   REGLA: toda ficha de conexion DEBE quedar respaldada por estos insumos.
# ─────────────────────────────────────────────────────────────────────────────
from backend.config import CONFIG
_BASE_FICHAS_DIR = CONFIG.FICHAS_DIR
_FICHAS_CACHE = {}

_TIPO_POR_PREFIJO = {
    "MA": "MATERIAL", "MO": "MANO_OBRA", "HER": "HERRAMIENTA",
    "EQ": "EQUIPO", "FL": "FLETE", "SC": "SUBCONTRATO", "DIS": "DISEÑO",
}


def _tipo_insumo(codigo: str) -> str:
    c = (codigo or "").upper()
    for pref, tipo in _TIPO_POR_PREFIJO.items():
        if c.startswith(pref):
            return tipo
    return "MATERIAL"


def _base_fichas_idx(version: str = "v1.2") -> dict:
    """Indice {codigo: ficha} de la base curada. Cacheado. {} si no se puede leer."""
    if version in _FICHAS_CACHE:
        return _FICHAS_CACHE[version]
    path = os.path.join(_BASE_FICHAS_DIR, version, "fichas", f"fichas_{version}.json")
    idx = {}
    try:
        with open(path, encoding="utf-8") as fh:
            for f in json.load(fh):
                cod = f.get("codigo")
                if cod:
                    idx[cod] = f
    except (OSError, ValueError):
        idx = {}
    _FICHAS_CACHE[version] = idx
    return idx


def insumos_ficha(codigo: str, version: str = "v1.2") -> dict:
    """Insumos + costo de una ficha (por codigo) desde la base curada.

    Devuelve {insumos:[{codigo,descripcion,unidad,tipo,cantidad,precio_unit,total}],
              costo_directo, precio_unitario_base, n_insumos, csi, unidad,
              descripcion_base, fuente, ok(bool)}.
    ok=False e insumos=[] si la ficha no existe en la base -> la regla 'ficha
    basada en insumos' NO se cumple y el caller debe avisarlo."""
    f = _base_fichas_idx(version).get(codigo)
    if not f:
        return {"insumos": [], "costo_directo": 0.0, "precio_unitario_base": 0.0,
                "n_insumos": 0, "csi": "", "unidad": "pza", "descripcion_base": "",
                "fuente": f"base {version}", "ok": False}
    insumos, total = [], 0.0
    for ins in f.get("insumos") or []:
        q = float(ins.get("cantidad") or 0)
        pu = float(ins.get("precioUnitario") or ins.get("precio_unitario") or 0)
        t = q * pu
        total += t
        insumos.append({
            "codigo": ins.get("codigo", ""),
            "descripcion": ins.get("descripcion", ""),
            "unidad": ins.get("unidad", ""),
            "tipo": _tipo_insumo(ins.get("codigo", "")),
            "cantidad": round(q, 6),
            "precio_unit": round(pu, 4),
            "total": round(t, 4),
        })
    pu_ficha = f.get("precio_unitario")
    return {
        "insumos": insumos,
        "costo_directo": round(total, 2),
        "precio_unitario_base": round(float(pu_ficha), 2) if pu_ficha is not None else round(total, 2),
        "n_insumos": len(insumos),
        "csi": f.get("csi", ""),
        "unidad": f.get("unidad", "pza"),
        "descripcion_base": f.get("descripcion", ""),
        "fuente": f"base {version}",
        "ok": len(insumos) > 0,
    }


def conexion_ficha(perfil_vig: str, perfil_col: str, tipo: str = "VC_CORTANTE") -> dict:
    """
    Resuelve la ficha de conexion (Div 05) a partir de los perfiles de viga y
    columna y el tipo. Sigue el plan §Capa 4.

    Devuelve dict:
      {ficha, tipo, descripcion, perfil_vig, perfil_col, aproximado(bool), rotulo}
    'aproximado'=True cuando no hay ficha exacta y se devuelve la de la clase
    mas cercana. 'rotulo' es la etiqueta lista para mostrar arriba de la hoja.
    """
    pv = _norm_perfil(perfil_vig)
    pc = _norm_perfil(perfil_col)
    tn = _norm_tipo_conexion(tipo)
    aprox = False
    ficha = ""
    desc = ""

    if tn == "SOLDADA":
        if pv == pc and pv in _CX_MISMO:
            ficha = _CX_MISMO[pv]
        elif (pc, pv) in _CX_MIXTO:
            ficha = _CX_MIXTO[(pc, pv)]
        else:
            # nearest: misma-perfil mas cercano por peso de la viga; si no, mixta CX-20
            aprox = True
            wv = _peso_perfil(pv)
            if wv is not None and _CX_MISMO:
                ficha = min(_CX_MISMO.items(),
                            key=lambda kv: abs((_peso_perfil(kv[0]) or 0) - wv))[1]
            else:
                ficha = "CX-20"
    elif tn == "VV":
        ficha = _CLASE_A_VV.get(_clase_peso(pv), "VV-2")
        aprox = (perfil_vig != "" and _peso_perfil(pv) is None)
    elif tn == "PLACA_BASE":
        # La placa base depende de la COLUMNA (no de la viga): clase por peso del col.
        ficha = _CLASE_A_BP.get(_clase_peso(pc), "BP-2")
        aprox = (perfil_col != "" and _peso_perfil(pc) is None)  # HSS -> clase media por fallback
    else:  # VC_CORTANTE / VC_MOMENTO -> empernada CV por clase
        ficha = _CLASE_A_CV.get(_clase_peso(pv), "CV-2")
        aprox = (perfil_vig != "" and _peso_perfil(pv) is None)

    f = FICHAS_CONEXION.get(ficha, {})
    desc = f.get("descripcion", "")
    tipo_out = f.get("tipo", tn)

    # Rotulo: ficha exacta usa su descripcion; aproximada usa rotulo generico
    nombre_tipo = {
        "VC_CORTANTE": "cortante", "VC_MOMENTO": "momento",
        "VV": "viga-viga", "SOLDADA": "soldada", "PLACA_BASE": "placa base",
    }.get(tipo_out, tipo_out.lower())
    pv_lbl = perfil_vig or "—"
    pc_lbl = perfil_col or "—"
    if aprox:
        rotulo = f"{ficha} (aprox.) · Conexion {nombre_tipo} — viga {pv_lbl} a columna {pc_lbl}"
    else:
        rotulo = f"{ficha} · {nombre_tipo} — viga {pv_lbl} a columna {pc_lbl}"

    # REGLA: la ficha SIEMPRE se respalda con insumos de la base curada.
    bom = insumos_ficha(ficha)
    return {
        "ficha": ficha,
        "tipo": tipo_out,
        "descripcion": desc or bom.get("descripcion_base", ""),
        "csi": bom.get("csi", ""),
        "perfil_vig": pv,
        "perfil_col": pc,
        "aproximado": aprox,
        "rotulo": rotulo,
        "insumos": bom["insumos"],
        "n_insumos": bom["n_insumos"],
        "costo_directo": bom["costo_directo"],
        "precio_unitario": bom["precio_unitario_base"],
        "unidad": bom["unidad"],
        "insumos_ok": bom["ok"],
    }


def catalogo_acero() -> dict:
    """
    Catalogo para poblar los menus del frontend (Capa 4-bis del plan).
    Lee de FICHAS_ACERO (perfil->ficha) + FICHAS_CONEXION.

    Devuelve:
      {vigas:[{ficha,perfil}], columnas:[{ficha,perfil}],
       conexiones:[{ficha,tipo,perfil_col,perfil_vig,descripcion}]}
    Los perfiles duales aparecen tanto en vigas (VA-x) como en columnas (C-x).
    """
    vigas, columnas = [], []
    for perfil, entry in FICHAS_ACERO.items():
        if isinstance(entry, dict):
            if "viga" in entry:
                vigas.append({"ficha": entry["viga"], "perfil": perfil})
            if "col" in entry:
                columnas.append({"ficha": entry["col"], "perfil": perfil})
        elif isinstance(entry, str):
            if entry.startswith("VA"):
                vigas.append({"ficha": entry, "perfil": perfil})
            else:
                columnas.append({"ficha": entry, "perfil": perfil})

    def _ficha_key(item):
        m = re.search(r"(\d+)", item["ficha"])
        return (item["ficha"][:2], int(m.group(1)) if m else 0)

    vigas.sort(key=_ficha_key)
    columnas.sort(key=_ficha_key)

    conexiones = [
        {"ficha": k, "tipo": v["tipo"], "perfil_col": v["perfil_col"],
         "perfil_vig": v["perfil_vig"], "descripcion": v["descripcion"]}
        for k, v in FICHAS_CONEXION.items()
    ]
    conexiones.sort(key=lambda x: (x["ficha"][:2], int(re.search(r"(\d+)", x["ficha"]).group(1))))

    return {"vigas": vigas, "columnas": columnas, "conexiones": conexiones}
