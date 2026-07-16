# -*- coding: utf-8 -*-
"""
seccion_ficha.py — Puente ETABS -> EstimaStruct (CONCRETO, Div 03)
==================================================================
Modulo ADITIVO y PURO (sin ORM). Dos responsabilidades:

  1) mapear_seccion_a_ficha(b_cm, h_cm, tipo) -> type_mark de ficha Div 03
     (R-x columnas, V-x vigas, S-x soleras). Match por dimensiones; si no
     hay match exacto devuelve el mas cercano + flag 'aproximado'.

  2) Parser tolerante del export de ETABS para concreto:
       - parse_concreto_texto(texto)  : CSV/TSV/texto pegado
       - parse_concreto_bytes(raw)    : .xlsx multi-hoja (openpyxl)
     Lee dos tablas por nombre de columna (tolerante a la version de ETABS):
       SECCIONES por miembro : Frame/Member, Section, b/t2, h/t3, fc
       FUERZAS por combinacion: Frame, OutputCase/Combo, P, V2, V3, T, M2, M3

Unidades: ver UNIDAD_FACTORES. ETABS suele exportar en kgf/kgf-m (flujo
CC-135: "kgf, m, C") o en kN/kN-m. El motor ACI de EstimaStruct trabaja en
t / t-m. Convertimos a t / t-m antes de crear los CasoDiseno.

Fuentes de verdad (etabs-flow/):
  - datos_ingreso_etabs_choc08_estimastruct.md (Pasos 4,5,6 -> tabla R/V/S)
  - integracion_estimastruct_concreto_choc08.md (seccion 5: diccionario)
"""

import csv
import io
import re


# ─────────────────────────────────────────────────────────────────────────────
# 1) DICCIONARIO SECCION (b x h cm) -> FICHA Div 03
#    Fuente: datos_ingreso_etabs_choc08_estimastruct.md, Pasos 4/5/6.
#    Clave = (min(b,h), max(b,h)) en cm  (insensible al orden b/h).
# ─────────────────────────────────────────────────────────────────────────────

# COLUMNAS / castillos (R-x)
FICHAS_COLUMNA = {
    (15, 15): "R-1",   # castillo
    (20, 20): "R-2",
    (20, 30): "R-3",   # 20x30 4#4+2#3  (R-4 misma seccion -> se prefiere R-3)
    (25, 25): "R-7",
    (30, 30): "R-8",
    (35, 35): "R-5",
    (40, 40): "R-9",
    (60, 60): "R-6",
}

# VIGAS (V-x)
FICHAS_VIGA = {
    (20, 25): "V-1",
    (15, 30): "V-2",
    (20, 30): "V-3",
    (25, 35): "V-4",
    (30, 40): "V-5",
}

# SOLERAS / tie-beams (S-x)  — vigas de amarre
FICHAS_SOLERA = {
    (35, 35): "S-1",
    (15, 20): "S-2",
    (20, 20): "S-3",
    (25, 25): "S-4",
}

# Pedestales (P-x) — columnas de cimentacion (referencia, no obligatorio mapear)
FICHAS_PEDESTAL = {
    (40, 40): "P-1",
    (50, 50): "P-2",
    (60, 60): "P-3",
    (70, 70): "P-4",
    (30, 30): "P-5",
    (35, 35): "P-6",
}


def _clave_dim(b_cm: float, h_cm: float):
    """(menor, mayor) redondeado a cm entero — insensible al orden."""
    a, b = sorted((round(float(b_cm)), round(float(h_cm))))
    return (a, b)


def _es_solera(tipo: str) -> bool:
    return "SOLERA" in (tipo or "").upper()


def _es_columna(tipo: str) -> bool:
    t = (tipo or "").upper()
    return "COLUMNA" in t or "PEDESTAL" in t or "CASTILLO" in t


def mapear_seccion_a_ficha(b_cm: float, h_cm: float, tipo: str = "VIGA") -> dict:
    """
    Mapea una seccion (b x h cm) + tipo a un type_mark de ficha Div 03.

    tipo: "VIGA"/"VIGA_SIMPLE" -> V-x ; "SOLERA" -> S-x ;
          "COLUMNA"/"PEDESTAL" -> R-x (o P-x).
    Si tipo es ambiguo, se infiere: secciones cuadradas grandes => columna.

    Devuelve dict:
      {type_mark, tipo_estructura, aproximado(bool),
       dim_ficha(b,h), dim_pedida(b,h), dist}
    Match exacto (tolerancia +-1 cm) -> aproximado=False.
    Sin match exacto -> ficha mas cercana por distancia euclidea, aproximado=True.
    """
    b_cm = float(b_cm)
    h_cm = float(h_cm)
    clave = _clave_dim(b_cm, h_cm)

    if _es_solera(tipo):
        tabla, tipo_estructura = FICHAS_SOLERA, "SOLERA"
    elif _es_columna(tipo):
        tabla, tipo_estructura = FICHAS_COLUMNA, "COLUMNA"
    else:
        tabla, tipo_estructura = FICHAS_VIGA, "VIGA"

    # 1) Match exacto (tolerancia +-1 cm en cada dimension)
    for (fb, fh), mark in tabla.items():
        if abs(fb - clave[0]) <= 1 and abs(fh - clave[1]) <= 1:
            return {
                "type_mark": mark,
                "tipo_estructura": tipo_estructura,
                "aproximado": False,
                "dim_ficha": {"b": fb, "h": fh},
                "dim_pedida": {"b": clave[0], "h": clave[1]},
                "dist": 0.0,
            }

    # 2) Mas cercano por distancia euclidea
    mejor_mark, mejor_dim, mejor_dist = None, None, None
    for (fb, fh), mark in tabla.items():
        dist = ((fb - clave[0]) ** 2 + (fh - clave[1]) ** 2) ** 0.5
        if mejor_dist is None or dist < mejor_dist:
            mejor_dist, mejor_mark, mejor_dim = dist, mark, (fb, fh)

    if mejor_mark is None:
        return {
            "type_mark": "", "tipo_estructura": tipo_estructura,
            "aproximado": True, "dim_ficha": None,
            "dim_pedida": {"b": clave[0], "h": clave[1]}, "dist": None,
        }

    return {
        "type_mark": mejor_mark,
        "tipo_estructura": tipo_estructura,
        "aproximado": True,
        "dim_ficha": {"b": mejor_dim[0], "h": mejor_dim[1]},
        "dim_pedida": {"b": clave[0], "h": clave[1]},
        "dist": round(mejor_dist, 2),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 2) UNIDADES — factores de conversion a las del motor (t , t-m)
#    1 t = 1000 kgf  ;  1 t = 9.80665 kN (aprox) -> 1 kN = 0.101972 t
# ─────────────────────────────────────────────────────────────────────────────
KN_POR_T = 9.80665   # 1 t-fuerza = 9.80665 kN

UNIDAD_FACTORES = {
    # unidad_entrada : (factor_fuerza_a_t, factor_momento_a_tm)
    "kgf":  (1.0 / 1000.0, 1.0 / 1000.0),    # kgf -> t ; kgf-m -> t-m
    "kg":   (1.0 / 1000.0, 1.0 / 1000.0),    # alias de kgf
    "ton":  (1.0, 1.0),                       # ya en t / t-m
    "t":    (1.0, 1.0),
    "kn":   (1.0 / KN_POR_T, 1.0 / KN_POR_T),  # kN -> t ; kN-m -> t-m
}


def factores_unidad(unidad_entrada: str):
    """Devuelve (factor_fuerza, factor_momento) para llevar a t / t-m.
    Default = kgf (flujo CC-135 'kgf, m, C'). Tolerante a mayusculas/espacios."""
    u = re.sub(r"[^a-z]", "", (unidad_entrada or "kgf").lower())
    return UNIDAD_FACTORES.get(u, UNIDAD_FACTORES["kgf"])


# ─────────────────────────────────────────────────────────────────────────────
# 3) PARSER TOLERANTE — reutiliza el enfoque de etabs_procedimiento.py
# ─────────────────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _to_float(s):
    if s is None:
        return None
    t = str(s).strip().replace(",", "")
    if t == "" or t in ("-", "—"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", t)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _split_rows(texto: str):
    texto = (texto or "").replace("\r\n", "\n").replace("\r", "\n")
    lineas = [ln for ln in texto.split("\n") if ln.strip() != ""]
    if not lineas:
        return []
    delim = ","
    if "\t" in texto and texto.count("\t") >= texto.count(","):
        delim = "\t"
    rows = []
    for r in csv.reader(io.StringIO("\n".join(lineas)), delimiter=delim):
        rows.append([c.strip() for c in r])
    return rows


def _find_header(rows, requeridas, opcionales=(), min_hits=2):
    """Busca la MEJOR fila de encabezado: debe contener TODAS las claves
    'requeridas' (normalizadas) y al menos 'min_hits' columnas reconocidas en
    total. Mapea tambien las 'opcionales'. Prefiere la fila con mas matches.
    Devuelve (idx_fila, {clave_norm: idx_col}) o (None, None).
    Match exacto o prefijo corto, para no confundir 'P' con 'Period'."""
    todas = list(requeridas) + list(opcionales)
    mejor_i, mejor_hit, mejor_n = None, None, -1
    umbral = max(len(requeridas), min_hits)
    for i, row in enumerate(rows):
        norm = [_norm(c) for c in row]
        hit = {}
        for j, cell in enumerate(norm):
            if not cell:
                continue
            for k in todas:
                if k in hit:
                    continue
                # exacto, o el header empieza por la clave (max +4 chars extra)
                if cell == k or (cell.startswith(k) and len(cell) <= len(k) + 4):
                    hit[k] = j
        if not all(k in hit for k in requeridas):
            continue
        if len(hit) < umbral:
            continue
        if len(hit) > mejor_n:
            mejor_i, mejor_hit, mejor_n = i, hit, len(hit)
    if mejor_i is None:
        return None, None
    return mejor_i, mejor_hit


# Alias de columnas ETABS -> clave normalizada interna
#   seccion:  frame/member/uniquename ; section/analysissection ; t2/width/b ; t3/depth/h ; fc/fce
#   fuerzas:  frame/member ; outputcase/combo/loadcase ; p ; v2 ; v3 ; t/tor ; m2 ; m3
def _buscar_col(hit, *alias):
    for a in alias:
        if a in hit:
            return hit[a]
    return None


def _parse_secciones(rows) -> dict:
    """Lee la tabla de SECCIONES por miembro. Devuelve {member: {section,b,h,fc}}."""
    # requiere al menos: un id de frame y dos dimensiones (o el nombre de seccion)
    hi, hit = _find_header(
        rows,
        requeridas=[],
        opcionales=["frame", "member", "uniquename", "section", "analysissection",
                    "t2", "t3", "width", "depth", "b", "h", "fc", "fce", "fc1",
                    "length", "longitud", "framelength", "largo"],
    )
    # Heuristica: necesitamos al menos (frame|member) y (section o t2/t3 o b/h)
    col_frame = _buscar_col(hit or {}, "frame", "member", "uniquename")
    col_sec = _buscar_col(hit or {}, "section", "analysissection")
    col_b = _buscar_col(hit or {}, "t2", "width", "b")
    col_h = _buscar_col(hit or {}, "t3", "depth", "h")
    col_fc = _buscar_col(hit or {}, "fc", "fce", "fc1")
    col_len = _buscar_col(hit or {}, "length", "longitud", "framelength", "largo")
    if hi is None or col_frame is None or (col_sec is None and (col_b is None or col_h is None)):
        return {}

    out = {}
    for row in rows[hi + 1:]:
        if not any(row):
            continue
        # Cortar al toparse con otra cabecera/tabla (ej. tabla de fuerzas)
        norm_row = {_norm(c) for c in row}
        if norm_row & {"outputcase", "combo", "loadcase", "casecombo"}:
            break
        if any(c.upper().startswith("TABLE") for c in row if c):
            continue
        def cell(idx):
            return row[idx] if (idx is not None and idx < len(row)) else None
        member = (cell(col_frame) or "").strip()
        if not member:
            continue
        # Si choca con otra tabla (encabezados repetidos), saltar filas no numericas
        b = _to_float(cell(col_b))
        h = _to_float(cell(col_h))
        fc = _to_float(cell(col_fc))
        section = (cell(col_sec) or "").strip()
        # Si no hay dimensiones explicitas, intentar extraerlas del nombre de seccion
        if (b is None or h is None) and section:
            m = re.search(r"(\d+(?:\.\d+)?)\s*[xX×]\s*(\d+(?:\.\d+)?)", section)
            if m:
                b = b if b is not None else float(m.group(1))
                h = h if h is not None else float(m.group(2))
        if b is None and h is None and not section:
            continue
        out[member] = {
            "section": section,
            "b": _to_cm(b) if b is not None else None,
            "h": _to_cm(h) if h is not None else None,
            "fc": fc,
            "L_m": _to_m(_to_float(cell(col_len))) if col_len is not None else None,
        }
    return out


def _to_m(v):
    """Longitud de frame ETABS -> metros. Flujo CC-135 exporta en m (kgf,m,C),
    asi que el caso normal es v<=30. Heuristica de unidad para exports en otra
    unidad (un frame >30 m es irreal en estos edificios -> viene en cm o mm):
      v>3000 -> mm (3 m=3000) ; 30<v<=3000 -> cm (3 m=300, 12 m=1200) ; else m."""
    if v is None:
        return None
    v = float(v)
    if v <= 0:
        return None
    if v > 3000:      # mm
        return round(v / 1000.0, 3)
    if v > 30:        # cm
        return round(v / 100.0, 3)
    return round(v, 3)


def _to_cm(v) -> float:
    """ETABS en 'm' exporta dimensiones < 5 (ej 0.35). Convertir a cm."""
    v = float(v)
    return round(v * 100, 1) if v < 5 else round(v, 1)


def _parse_fuerzas(rows) -> list:
    """Lee la tabla de FUERZAS por combinacion. Devuelve lista de
    {member, combo, P, V2, V3, T, M2, M3} (crudos, en unidad de ETABS)."""
    hi, hit = _find_header(
        rows,
        requeridas=["outputcase"],
        opcionales=["frame", "member", "combo", "loadcase", "casecombo",
                    "p", "v2", "v3", "t", "tor", "m2", "m3", "station"],
    )
    if hi is None:
        # ETABS a veces nombra la columna 'Combo' o 'Load Case/Combo'
        hi, hit = _find_header(
            rows, requeridas=["combo"],
            opcionales=["frame", "member", "p", "v2", "v3", "t", "tor", "m2", "m3"],
        )
    if hi is None:
        return []

    col_frame = _buscar_col(hit, "frame", "member")
    col_combo = _buscar_col(hit, "outputcase", "combo", "loadcase", "casecombo")
    col_p = _buscar_col(hit, "p")
    col_v2 = _buscar_col(hit, "v2")
    col_v3 = _buscar_col(hit, "v3")
    col_t = _buscar_col(hit, "t", "tor")
    col_m2 = _buscar_col(hit, "m2")
    col_m3 = _buscar_col(hit, "m3")
    if col_combo is None:
        return []

    out = []
    for row in rows[hi + 1:]:
        if not any(row):
            continue
        def cell(idx):
            return row[idx] if (idx is not None and idx < len(row)) else None
        combo = (cell(col_combo) or "").strip()
        member = (cell(col_frame) or "").strip()
        if not combo:
            continue
        out.append({
            "member": member,
            "combo": combo,
            "P":  _to_float(cell(col_p)),
            "V2": _to_float(cell(col_v2)),
            "V3": _to_float(cell(col_v3)),
            "T":  _to_float(cell(col_t)),
            "M2": _to_float(cell(col_m2)),
            "M3": _to_float(cell(col_m3)),
        })
    return out


def parse_concreto_texto(texto: str) -> dict:
    """Parser de un export ETABS de CONCRETO en CSV/TSV/texto.
    Devuelve {secciones:{member:{...}}, fuerzas:[{...}], avisos:[]}.
    Secciones y fuerzas pueden venir en el mismo texto (una tabla debajo de la
    otra) o por separado; el parser intenta ambas."""
    out = {"secciones": {}, "fuerzas": [], "avisos": []}
    rows = _split_rows(texto)
    if not rows:
        out["avisos"].append("Texto vacio o ilegible.")
        return out
    out["secciones"] = _parse_secciones(rows)
    out["fuerzas"] = _parse_fuerzas(rows)
    if not out["secciones"]:
        out["avisos"].append("No se reconocio la tabla de SECCIONES "
                             "(Frame/Section/t2/t3/fc).")
    if not out["fuerzas"]:
        out["avisos"].append("No se reconocio la tabla de FUERZAS "
                             "(Frame/OutputCase/P/V2/V3/T/M2/M3).")
    return out


def parse_reacciones_texto(texto: str) -> dict:
    """Parser de un export ETABS de JOINT REACTIONS (reacciones en apoyos).
    Devuelve {reacciones:[{joint,combo,FX,FY,FZ,MX,MY,MZ}], avisos:[]}.
    Reconoce Joint/Label/UniqueName + OutputCase/Combo + FX/FY/FZ (o F1/F2/F3)
    + MX/MY/MZ (o M1/M2/M3). Valores crudos en la unidad de ETABS.
    Uso: placa base §J8 -> Pu = |FZ| (axial de columna), Vu = hypot(FX,FY)."""
    out = {"reacciones": [], "avisos": []}
    rows = _split_rows(texto)
    if not rows:
        out["avisos"].append("Texto vacio o ilegible.")
        return out
    hi, hit = _find_header(
        rows, requeridas=["outputcase"],
        opcionales=["joint", "label", "uniquename", "unique", "combo", "loadcase",
                    "casecombo", "fx", "fy", "fz", "f1", "f2", "f3",
                    "mx", "my", "mz", "m1", "m2", "m3"],
    )
    if hi is None:
        hi, hit = _find_header(
            rows, requeridas=["combo"],
            opcionales=["joint", "label", "uniquename", "unique",
                        "fx", "fy", "fz", "f1", "f2", "f3",
                        "mx", "my", "mz", "m1", "m2", "m3"],
        )
    if hi is None:
        out["avisos"].append("No se reconocio la tabla de REACCIONES "
                             "(Joint/OutputCase/FX/FY/FZ).")
        return out

    col_j  = _buscar_col(hit, "joint", "label", "uniquename", "unique")
    col_c  = _buscar_col(hit, "outputcase", "combo", "loadcase", "casecombo")
    col_fx = _buscar_col(hit, "fx", "f1")
    col_fy = _buscar_col(hit, "fy", "f2")
    col_fz = _buscar_col(hit, "fz", "f3")
    col_mx = _buscar_col(hit, "mx", "m1")
    col_my = _buscar_col(hit, "my", "m2")
    col_mz = _buscar_col(hit, "mz", "m3")
    if col_c is None or col_fz is None:
        out["avisos"].append("Faltan columnas clave (OutputCase y FZ/F3).")
        return out

    for row in rows[hi + 1:]:
        if not any(row):
            continue
        def cell(idx):
            return row[idx] if (idx is not None and idx < len(row)) else None
        combo = (cell(col_c) or "").strip()
        if not combo:
            continue
        out["reacciones"].append({
            "joint": (cell(col_j) or "").strip(),
            "combo": combo,
            "FX": _to_float(cell(col_fx)), "FY": _to_float(cell(col_fy)),
            "FZ": _to_float(cell(col_fz)),
            "MX": _to_float(cell(col_mx)), "MY": _to_float(cell(col_my)),
            "MZ": _to_float(cell(col_mz)),
        })
    if not out["reacciones"]:
        out["avisos"].append("Cabecera de reacciones reconocida pero sin filas de datos.")
    return out


def parse_concreto_bytes(raw: bytes) -> dict:
    """Parser de export .xlsx ETABS (multi-hoja). Une secciones y fuerzas
    encontradas en cualquier hoja. Tolerante a nombres de hoja."""
    out = {"secciones": {}, "fuerzas": [], "avisos": [], "formato": "xlsx"}
    try:
        from openpyxl import load_workbook
    except ImportError:
        out["avisos"].append("openpyxl no instalado; exporta como CSV.")
        return out
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        out["avisos"].append(f"No se pudo abrir el .xlsx: {e}")
        return out

    for ws in wb.worksheets:
        buf = io.StringIO()
        w = csv.writer(buf)
        for fila in ws.iter_rows(values_only=True):
            if fila is None:
                continue
            w.writerow(["" if c is None else c for c in fila])
        texto = buf.getvalue()
        if not texto.strip():
            continue
        rows = _split_rows(texto)
        sec = _parse_secciones(rows)
        fz = _parse_fuerzas(rows)
        if sec:
            out["secciones"].update(sec)
        if fz:
            out["fuerzas"].extend(fz)
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass

    if not out["secciones"]:
        out["avisos"].append("No se reconocio la tabla de SECCIONES en el .xlsx.")
    if not out["fuerzas"]:
        out["avisos"].append("No se reconocio la tabla de FUERZAS en el .xlsx.")
    return out
