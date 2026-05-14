import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

wb = openpyxl.load_workbook(r"D:\OneDrive\Bots\Estimbot\MasterFiles\Fichas revisadas 26 abril.xlsx", data_only=True)
print("Sheets:", wb.sheetnames)
for sh in wb.sheetnames[:5]:
    ws = wb[sh]
    print(f"\n--- Sheet: {sh} ({ws.max_row} rows x {ws.max_column} cols) ---")
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(row)
