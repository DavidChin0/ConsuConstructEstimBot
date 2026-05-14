import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None

EXCEL = r"D:\OneDrive\Bots\Estimbot\MasterFiles\INPUTS\FULL_breakdown\MatricesImport.xlsx"
wb = openpyxl.load_workbook(EXCEL, data_only=True)
print("Sheets:", wb.sheetnames)
for sh in wb.sheetnames[:3]:
    ws = wb[sh]
    print(f"\n--- Sheet: {sh} ({ws.max_row} rows x {ws.max_column} cols) ---")
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        print(row)
