"""
seed_recursos.py — Importa Materiales y Manode Obra.xlsx al catálogo de recursos
Tipos: DIS→DISEÑO, EQ→EQUIPO, HER→HERRAMIENTA, MA→MATERIAL, MO→MANO_OBRA, SC→SUBCONTRATO
"""
import os, sys
os.chdir(os.path.dirname(__file__))
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from db import SessionLocal, engine
from models import Base, Recurso

XLSX_PATH = r"D:\OneDrive\Bots\Estimbot\Despieces\Materiales y Manode Obra.xlsx"

TIPO_MAP = {
    "DIS": "DISEÑO",
    "EQ":  "EQUIPO",
    "HER": "HERRAMIENTA",
    "MA":  "MATERIAL",
    "MO":  "MANO_OBRA",
    "SC":  "SUBCONTRATO",
}

Base.metadata.create_all(bind=engine)
db = SessionLocal()

try:
    existing = db.query(Recurso).count()
    if existing:
        print(f"Limpiando {existing} recursos anteriores...")
        db.query(Recurso).delete()
        db.commit()

    print(f"Leyendo {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["MA-MO"]

    importados = 0
    omitidos = 0
    por_tipo = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        clave = row[0]
        if not clave:
            continue
        clave = str(clave).strip()
        prefix = clave.split("-")[0].upper()
        tipo = TIPO_MAP.get(prefix)
        if not tipo:
            omitidos += 1
            continue

        descripcion = str(row[1]).strip() if row[1] else clave
        unidad = str(row[2]).strip() if row[2] else "global"
        precio = float(row[4]) if row[4] is not None else 0.0

        r = Recurso(
            clave=clave,
            descripcion=descripcion,
            unidad=unidad,
            tipo=tipo,
            precio_unitario=precio,
        )
        db.add(r)
        importados += 1
        por_tipo[tipo] = por_tipo.get(tipo, 0) + 1

    db.commit()
    print(f"\nSeed recursos completado:")
    print(f"  Total importados: {importados}")
    for tipo, cnt in sorted(por_tipo.items()):
        print(f"  {tipo}: {cnt}")
    print(f"  Omitidos: {omitidos}")

except Exception as e:
    db.rollback()
    print(f"ERROR: {e}")
    raise
finally:
    db.close()
