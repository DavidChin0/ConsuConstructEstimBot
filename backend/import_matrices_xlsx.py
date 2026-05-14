"""Importa matrices e insumos desde XLSX al proyecto 'test'.
Reusa la lógica de import_matrices_csv pero parsea xlsx con openpyxl."""
import sys
import os
from decimal import Decimal
import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from db import SessionLocal, engine
from models import Base, Capitulo, Partida, InsumoPartida, Recurso, DIVISIONES_CSI
from import_matrices_csv import get_or_create_obra_test, get_tipo_from_clave


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrices = []
    current = None
    for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None or (isinstance(c, str) and c.strip() == '') for c in row):
            continue
        a = row[0]
        if a is None:
            continue
        a_str = str(a).strip()
        if a_str == 'C' or a_str == 'Clave':
            continue
        if a_str == 'Simple':
            if current is None:
                continue
            def f(idx):
                try:
                    return float(row[idx]) if row[idx] is not None and str(row[idx]).strip() != '' else 0
                except (TypeError, ValueError):
                    return 0
            current['recursos'].append({
                'clave': str(row[1]).strip() if row[1] else '',
                'descripcion': str(row[2]).strip() if row[2] else '',
                'unidad': str(row[3]).strip() if row[3] else 'global',
                'rendimiento': f(4),
                'costo_unit': f(5),
                'total': f(6),
            })
            continue
        if a_str and a_str[0].isdigit() and ' ' in a_str:
            if current:
                matrices.append(current)
            try:
                cant = float(row[3]) if row[3] is not None else 0
            except (TypeError, ValueError):
                cant = 0
            current = {
                'csi': a_str,
                'clave': str(row[1]).strip() if row[1] else '',
                'descripcion': str(row[2]).strip() if row[2] else '',
                'unidad': str(row[3]).strip() if row[3] else 'global',
                'cantidad': cant,
                'recursos': [],
            }
    if current:
        matrices.append(current)
    return matrices


def import_matrices_from_xlsx(xlsx_path):
    db = SessionLocal()
    try:
        matrices = parse_xlsx(xlsx_path)
        print(f"\n[OK] {len(matrices)} matrices encontradas en XLSX")

        obra = get_or_create_obra_test(db)
        imported = 0
        failed = 0
        partida_counter = {}

        for matrix in matrices:
            try:
                csi = matrix['csi']
                cap_code = csi.split()[0]
                try:
                    cap_orden = int(cap_code)
                except ValueError:
                    cap_orden = 999

                capitulo = db.query(Capitulo).filter(
                    Capitulo.presupuesto_id == obra.id,
                    Capitulo.clave == cap_code
                ).first()

                if not capitulo:
                    cap_name = DIVISIONES_CSI.get(cap_code, f"Capitulo {cap_code}")
                    capitulo = Capitulo(
                        presupuesto_id=obra.id,
                        clave=cap_code,
                        nombre=cap_name,
                        orden=cap_orden
                    )
                    db.add(capitulo)
                    db.flush()
                else:
                    if capitulo.orden != cap_orden:
                        capitulo.orden = cap_orden

                partida_counter.setdefault(capitulo.id, 0)
                partida_counter[capitulo.id] += 1
                part_orden = partida_counter[capitulo.id]

                partida = db.query(Partida).filter(
                    Partida.capitulo_id == capitulo.id,
                    Partida.clave_csi == csi
                ).first()

                if not partida:
                    partida = Partida(
                        capitulo_id=capitulo.id,
                        clave_csi=csi,
                        descripcion=matrix['descripcion'],
                        unidad=matrix['unidad'],
                        cantidad=Decimal(str(matrix['cantidad'])),
                        costo_mo=Decimal('0'),
                        costo_ma=Decimal('0'),
                        unitario_matriz=Decimal('0'),
                        costo_base=Decimal('0'),
                        precio_unitario=Decimal('0'),
                        total=Decimal('0'),
                        orden=part_orden
                    )
                    db.add(partida)
                    db.flush()
                else:
                    partida.orden = part_orden
                    partida.descripcion = matrix['descripcion']
                    partida.unidad = matrix['unidad']
                    partida.cantidad = Decimal(str(matrix['cantidad']))

                # Limpiar insumos previos
                db.query(InsumoPartida).filter(
                    InsumoPartida.partida_id == partida.id
                ).delete()

                mo_total = Decimal('0')
                ma_total = Decimal('0')
                for i, recurso in enumerate(matrix['recursos'], 1):
                    clave = recurso['clave']
                    tipo = get_tipo_from_clave(clave)
                    res_obj = db.query(Recurso).filter(
                        Recurso.clave == clave
                    ).first()
                    recurso_id = res_obj.id if res_obj else None

                    insumo = InsumoPartida(
                        partida_id=partida.id,
                        recurso_id=recurso_id,
                        clave=clave,
                        descripcion=recurso['descripcion'],
                        unidad=recurso['unidad'],
                        tipo=tipo,
                        cantidad=Decimal(str(recurso['rendimiento'])),
                        costo_unit=Decimal(str(recurso['costo_unit'])),
                        total=Decimal(str(recurso['total'])),
                        orden=i
                    )
                    db.add(insumo)
                    if tipo == 'MANO_OBRA':
                        mo_total += Decimal(str(recurso['total']))
                    else:
                        ma_total += Decimal(str(recurso['total']))

                partida.costo_mo = mo_total
                partida.costo_ma = ma_total
                partida.unitario_matriz = Decimal('0')
                partida.costo_base = mo_total + ma_total
                partida.precio_unitario = partida.costo_base
                if partida.cantidad > 0:
                    partida.total = partida.costo_base * partida.cantidad

                imported += 1
                print(f"  [OK] {csi:14} | {len(matrix['recursos']):2} insumos | {matrix['descripcion'][:40]}")
            except Exception as e:
                failed += 1
                print(f"  [ERR] {matrix['csi']}: {e}")

        db.commit()
        print(f"\n{'='*60}")
        print(f"Importadas: {imported} | Errores: {failed}")
        print(f"Obra: 'test' ({obra.id})")
        print('='*60)
        return failed == 0
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        return False
    finally:
        db.close()


if __name__ == "__main__":
    Base.metadata.create_all(bind=engine)
    xlsx = sys.argv[1] if len(sys.argv) > 1 else r"D:\OneDrive\Bots\Estimbot\MasterFiles\INPUTS\FULL_breakdown\MatricesImport.xlsx"
    if not os.path.exists(xlsx):
        print(f"No existe: {xlsx}")
        sys.exit(1)
    print(f"Importando desde: {xlsx}")
    ok = import_matrices_from_xlsx(xlsx)
    sys.exit(0 if ok else 1)
