"""Verifica matriz por matriz que el XLSX coincida exactamente con la DB del proyecto 'test'."""
import sys
import os
from decimal import Decimal
import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from db import SessionLocal
from models import Presupuesto, Capitulo, Partida, InsumoPartida

XLSX = r"D:\OneDrive\Bots\Estimbot\MasterFiles\INPUTS\FULL_breakdown\MatricesImport.xlsx"
TOL = Decimal('0.000001')


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrices = []
    current = None
    for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None or (isinstance(c, str) and c.strip() == '') for c in row):
            continue
        a = row[0]
        if a is None:
            continue
        a_str = str(a).strip()
        # Encabezado interno
        if a_str == 'C':
            continue
        if a_str == 'Simple':
            if current is None:
                continue
            try:
                rend = float(row[4]) if row[4] is not None else 0
            except (TypeError, ValueError):
                rend = 0
            try:
                costo = float(row[5]) if row[5] is not None else 0
            except (TypeError, ValueError):
                costo = 0
            try:
                total = float(row[6]) if row[6] is not None else 0
            except (TypeError, ValueError):
                total = 0
            current['recursos'].append({
                'clave': str(row[1]).strip() if row[1] else '',
                'descripcion': str(row[2]).strip() if row[2] else '',
                'unidad': str(row[3]).strip() if row[3] else '',
                'rendimiento': rend,
                'costo_unit': costo,
                'total': total,
            })
            continue
        # Posible fila de matriz: empieza con dígito y tiene espacio
        if a_str and a_str[0].isdigit() and ' ' in a_str:
            if current:
                matrices.append(current)
            try:
                cant = float(row[3]) if row[3] is not None else 0
            except (TypeError, ValueError):
                cant = 0
            current = {
                'csi': a_str,
                'clave': str(row[1]).strip() if row[1] else '',
                'descripcion': str(row[2]).strip() if row[2] else '',
                'unidad': str(row[3]).strip() if row[3] else '',
                'cantidad': cant,
                'recursos': [],
            }
    if current:
        matrices.append(current)
    return matrices


def main():
    matrices = parse_xlsx(XLSX)
    print(f"XLSX: {len(matrices)} matrices")

    # Detectar duplicados de CSI
    seen = {}
    dups = []
    for m in matrices:
        seen.setdefault(m['csi'], 0)
        seen[m['csi']] += 1
    for k, v in seen.items():
        if v > 1:
            dups.append((k, v))
    if dups:
        print(f"AVISO: CSIs duplicados en XLSX: {dups[:10]}")

    db = SessionLocal()
    obra = db.query(Presupuesto).filter(Presupuesto.nombre == 'test').first()
    if not obra:
        print("ERROR: obra 'test' no existe")
        return 1

    matrices_ok = 0
    diffs = []  # (csi, tipo, detalle)

    csi_set_xlsx = set()
    for m in matrices:
        csi = m['csi']
        csi_set_xlsx.add(csi)
        cap_code = csi.split()[0]
        capitulo = db.query(Capitulo).filter(
            Capitulo.presupuesto_id == obra.id,
            Capitulo.clave == cap_code
        ).first()
        if not capitulo:
            diffs.append((csi, 'CAP_FALTA', cap_code))
            continue
        partida = db.query(Partida).filter(
            Partida.capitulo_id == capitulo.id,
            Partida.clave_csi == csi
        ).first()
        if not partida:
            diffs.append((csi, 'MAT_FALTA', ''))
            continue

        insumos_db = db.query(InsumoPartida).filter(
            InsumoPartida.partida_id == partida.id
        ).all()
        xlsx_claves = [r['clave'] for r in m['recursos']]
        db_claves = [i.clave for i in insumos_db]

        extra = set(db_claves) - set(xlsx_claves)
        falt = set(xlsx_claves) - set(db_claves)
        for c in extra:
            diffs.append((csi, 'INSUMO_EXTRA_DB', c))
        for c in falt:
            diffs.append((csi, 'INSUMO_FALTA_DB', c))

        db_map = {i.clave: i for i in insumos_db}
        rend_diff_local = False
        for r in m['recursos']:
            ins = db_map.get(r['clave'])
            if not ins:
                continue
            xlsx_rend = Decimal(str(r['rendimiento']))
            db_rend = Decimal(str(ins.cantidad))
            if abs(xlsx_rend - db_rend) > TOL:
                diffs.append((csi, 'REND_DIFF', f"{r['clave']} xlsx={xlsx_rend} db={db_rend}"))
                rend_diff_local = True

        if not extra and not falt and not rend_diff_local:
            matrices_ok += 1

    todas_partidas = db.query(Partida).join(Capitulo).filter(
        Capitulo.presupuesto_id == obra.id
    ).all()
    huerf = [p.clave_csi for p in todas_partidas if p.clave_csi not in csi_set_xlsx]

    print(f"\nOK: {matrices_ok}/{len(matrices)}")
    print(f"Discrepancias: {len(diffs)}")
    by_type = {}
    for c, t, d in diffs:
        by_type.setdefault(t, []).append((c, d))
    for t, items in by_type.items():
        print(f"\n[{t}] x {len(items)}")
        for c, d in items[:20]:
            print(f"  - {c}: {d}")
        if len(items) > 20:
            print(f"  ... y {len(items)-20} mas")

    print(f"\nPartidas en DB sin XLSX: {len(huerf)}")
    for c in huerf[:20]:
        print(f"  - {c}")

    db.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
