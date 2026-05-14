"""
seed_bd.py — Carga 'Base de Datos Import.xlsx' como Template CC 2026
Columnas: A=CSI, B=TypeMark, C=Descripcion, D=Unidad, E=Cantidad(formula), F=MO, G=Materiales
Colores: amarillo=FFFFFF00, verde=FF92D050, azul=FF00B0F0, blanco=00000000, rosa=resto
"""
import os, sys, re
os.chdir(os.path.dirname(__file__))
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from db import SessionLocal, engine
from models import Base, Presupuesto, ConfigPresupuesto, Capitulo, Partida, DIVISIONES_CSI

XLSX_PATH = r"D:\OneDrive\Bots\Estimbot\MasterFiles\Base de Datos Import.xlsx"
TEMPLATE_NOMBRE = "Template CC 2026"

COLOR_MAP = {
    "FFFFFF00": "amarillo",
    "FF92D050": "verde",
    "FF00B0F0": "azul",
    "00000000": "blanco",
}

def get_color(cell):
    try:
        rgb = cell.fill.fgColor.rgb
        return COLOR_MAP.get(rgb, "rosa")
    except Exception:
        return "rosa"

def parse_cantidad(val):
    """Devuelve (cantidad_float, es_formula, formula_ref)."""
    if val is None:
        return 0.0, False, None
    if isinstance(val, str) and val.startswith("="):
        formula_ref = val
        # Fórmula aritmética simple (ej: =95.35+42.14) → evaluar
        expr = val[1:].replace(",", ".")
        if re.match(r'^[\d\.\+\-\*\/\(\) ]+$', expr):
            try:
                return float(eval(expr)), True, formula_ref
            except Exception:
                pass
        # Referencia a otra celda (=$E$66) → cantidad 0, se resuelve al clonar
        return 0.0, True, formula_ref
    try:
        return float(val), False, None
    except (TypeError, ValueError):
        return 0.0, False, None

def to_float(v):
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


Base.metadata.create_all(bind=engine)

if not os.path.exists(XLSX_PATH):
    print(f"[seed_bd] Archivo no encontrado, omitiendo seed: {XLSX_PATH}")
    sys.exit(0)

db = SessionLocal()

try:
    existing = db.query(Presupuesto).filter(Presupuesto.es_template == True).first()
    if existing:
        print(f"Borrando template anterior: {existing.nombre}")
        db.delete(existing)
        db.commit()

    print(f"Leyendo {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb.active
    print(f"  Sheet: {ws.title} | Filas: {ws.max_row}")

    # Row 1 = meta (sobrecosto), Row 2 = headers, Row 3+ = datos
    rows_raw = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False))
    print(f"  Filas a procesar: {len(rows_raw)}")

    template = Presupuesto(nombre=TEMPLATE_NOMBRE, es_template=True, moneda="HNL")
    db.add(template)
    db.flush()
    cfg = ConfigPresupuesto(presupuesto_id=template.id, sobrecosto=20, iva=15)
    db.add(cfg)
    db.flush()

    capitulos_map = {}
    orden_cap = 0
    orden_partida = {}
    actividades = 0
    formulas = 0
    omitidas = 0

    for row in rows_raw:
        clave_csi = row[0].value
        if not clave_csi or str(clave_csi).strip() == "":
            omitidas += 1
            continue

        clave_csi = str(clave_csi).strip()
        div = clave_csi[:2] if len(clave_csi) >= 2 else "00"

        if div not in DIVISIONES_CSI:
            div_int = int(div) if div.isdigit() else 0
            div = min(DIVISIONES_CSI.keys(), key=lambda d: abs(int(d) - div_int))

        if div not in capitulos_map:
            cap = Capitulo(
                presupuesto_id=template.id,
                clave=div,
                nombre=DIVISIONES_CSI[div],
                orden=orden_cap,
            )
            db.add(cap)
            db.flush()
            capitulos_map[div] = cap
            orden_partida[cap.id] = 0
            orden_cap += 1

        cap = capitulos_map[div]

        # Descripcion y unidad
        descripcion = str(row[2].value).strip().replace("_x000D_", "").strip() if row[2].value else clave_csi
        unidad = str(row[3].value).strip() if row[3].value else "global"

        # Cantidad / fórmula (col E)
        cantidad_val, es_formula, formula_ref = parse_cantidad(row[4].value)

        # Costos (col F=MO, G=Materiales)
        costo_mo = to_float(row[5].value)
        costo_ma = to_float(row[6].value)
        costo_base = costo_mo + costo_ma
        precio_unitario = costo_base * 1.20

        # Type Mark (col B)
        type_mark = str(row[1].value).strip() if row[1].value else None

        # Color de celda
        color_tipo = get_color(row[0])

        partida = Partida(
            capitulo_id=cap.id,
            clave_csi=clave_csi,
            descripcion=descripcion,
            unidad=unidad,
            cantidad=0,
            revit_q=cantidad_val if not es_formula else 0,
            factor_e=1,
            factor_f=1,
            color_tipo=color_tipo,
            costo_mo=costo_mo,
            costo_ma=costo_ma,
            unitario_matriz=0,
            costo_base=costo_base,
            precio_unitario=precio_unitario,
            total=0,
            es_formula=es_formula,
            formula_ref=formula_ref,
            type_mark=type_mark,
            omniclass_num=None,
            assembly_num=None,
            orden=orden_partida[cap.id],
        )
        db.add(partida)
        orden_partida[cap.id] += 1
        actividades += 1
        if es_formula:
            formulas += 1

    db.commit()

    print(f"\nSeed completado:")
    print(f"  Divisiones CSI:      {len(capitulos_map)}")
    print(f"  Actividades totales: {actividades}")
    print(f"  Con fórmulas:        {formulas}")
    print(f"  Filas omitidas:      {omitidas}")
    print(f"\nTemplate '{TEMPLATE_NOMBRE}' listo. ID: {template.id}")

    # Mostrar divisiones cargadas
    print("\nDivisiones cargadas:")
    for div, cap in sorted(capitulos_map.items()):
        cur_ord = orden_partida[cap.id]
        print(f"  {div} — {DIVISIONES_CSI[div]}: {cur_ord} partidas")

except Exception as e:
    db.rollback()
    print(f"ERROR: {e}")
    raise
finally:
    db.close()
