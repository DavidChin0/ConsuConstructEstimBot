"""
Updater — importa o actualiza fichas en una base de datos desde un Excel.
Formato Excel esperado (igual al de 'Fichas revisadas 26 abril.xlsx'):
  Col A: "CSI / Recurso" (header) | CSI code (ficha) | "Recurso" (insumo)
  Col B: Type Mark / Clave
  Col C: Descripción
  Col D: Unidad
  Col E: Cantidad
  Col F: Costo Unit.
  Col G: Total

GET  /updater/files          → lista archivos xlsx en la carpeta Updater
POST /updater/import         → body: {filename, version} → parsea y actualiza/agrega fichas al JSON
"""
# MIGRATION TOOL — Solo para import inicial desde Excel/OPUS.
# No usar en operación normal. Fuente activa: estimacion.db
import json, os, re, shutil, tempfile
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel
from db import get_db
from models import Partida, Capitulo, ConfigPresupuesto, InsumoPartida, DIVISIONES_CSI, new_uuid
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from csi_utils import infer_csi

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

router = APIRouter(prefix="/updater", tags=["updater"])

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
UPDATER_PATH = os.environ.get("ESTIMASTRUCT_UPDATER_PATH", os.path.join(REPO_ROOT, "MasterFiles", "Updater"))
BASE_PATH = os.path.join(REPO_ROOT, "development", "Template2_Updated")

_CSI_RE = re.compile(r"^\d{2}\s\d{2}\s\d{2}")
_TM_RE  = re.compile(r"^[A-Z][A-Z0-9]{0,7}[\-\d]")   # type mark: letters then dash/digit
_SKIP_A = {"CSI / Recurso", "CSI/Recurso", "CSI"}


_MAX_BAKS = 4

def _ficha_path(version: str) -> str:
    return os.path.join(BASE_PATH, version, "fichas", f"fichas_{version}.json")

def _bak_path(version: str, n: int) -> str:
    return os.path.join(BASE_PATH, version, "fichas", f"fichas_{version}.bak{n}.json")


def _backup(version: str):
    path = _ficha_path(version)
    if not os.path.exists(path):
        return
    for i in range(_MAX_BAKS, 1, -1):
        src = _bak_path(version, i - 1)
        dst = _bak_path(version, i)
        if os.path.exists(src):
            shutil.copy2(src, dst)
    shutil.copy2(path, _bak_path(version, 1))


@router.get("/files")
def list_updater_files():
    if not os.path.exists(UPDATER_PATH):
        return []
    out = []
    for name in sorted(os.listdir(UPDATER_PATH)):
        if name.lower().endswith((".xlsx", ".xls")):
            st = os.stat(os.path.join(UPDATER_PATH, name))
            out.append({"name": name, "size": st.st_size, "mtime": st.st_mtime})
    return out


def _cell(row, idx) -> str:
    v = row[idx] if idx < len(row) else None
    return str(v).strip() if v is not None else ""


def _num(row, idx) -> float:
    v = row[idx] if idx < len(row) else None
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _parse_excel(filepath: str) -> list:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado")

    # Copy to temp file to bypass Excel's file lock
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        shutil.copy2(filepath, tmp_path)
        return _parse_workbook(tmp_path)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def _parse_workbook(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    fichas: list  = []
    current: dict = None

    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None for v in row):
            continue

        col_a = _cell(row, 0)
        col_b = _cell(row, 1)
        col_c = _cell(row, 2)
        col_d = _cell(row, 3)

        # Skip header rows
        if col_a in _SKIP_A or col_b in ("Type Mark / Clave", "Type Mark"):
            continue

        # Ficha header: CSI present
        if _CSI_RE.match(col_a):
            if current:
                fichas.append(current)
            # Handle merged "TM-01 Descripción..." in col_b when col_c is empty
            parts = col_b.split(None, 1)
            if len(parts) == 2 and re.match(r'^[A-Z0-9\-\.]+$', parts[0]) and not col_c:
                codigo = parts[0]
                desc = parts[1]
            else:
                codigo = col_b
                desc = col_c or col_b
            current = {
                "csi": col_a,
                "codigo": (codigo or "").strip(),
                "descripcion": (desc or "").replace("_x000D_", "").strip(),
                "unidad": col_d or "global",
                "precio_unitario": 0.0,
                "insumos": [],
            }
            continue

        # Ficha header: no CSI — infer chapter from type mark prefix / description
        if not col_a and col_b and _TM_RE.match(col_b):
            if current:
                fichas.append(current)
            parts = col_b.split(None, 1)
            if len(parts) == 2 and re.match(r'^[A-Z0-9\-\.]+$', parts[0]) and not col_c:
                codigo = parts[0]
                desc = parts[1]
            else:
                codigo = col_b
                desc = col_c or col_b
            current = {
                "csi": infer_csi(codigo, desc),
                "codigo": (codigo or "").strip(),
                "descripcion": (desc or "").replace("_x000D_", "").strip(),
                "unidad": col_d or "global",
                "precio_unitario": 0.0,
                "insumos": [],
            }
            continue

        # Insumo row
        if col_a.lower() == "recurso" and current is not None:
            cant = _num(row, 4)
            pu   = _num(row, 5)
            current["insumos"].append({
                "codigo":       col_b,
                "descripcion":  col_c,
                "unidad":       col_d or "global",
                "cantidad":     cant,
                "precioUnitario": pu,
            })

    if current:
        fichas.append(current)

    # Calculate precio_unitario from insumos sum
    for fi in fichas:
        fi["precio_unitario"] = round(
            sum(i["cantidad"] * i["precioUnitario"] for i in fi["insumos"]), 4
        )

    return fichas


class ImportRequest(BaseModel):
    filename: str
    version:  str


@router.post("/import")
def import_fichas(req: ImportRequest):
    filepath = os.path.join(UPDATER_PATH, req.filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, f"Archivo '{req.filename}' no encontrado en carpeta Updater")

    json_path = _ficha_path(req.version)
    if not os.path.exists(json_path):
        raise HTTPException(404, f"Versión '{req.version}' no encontrada")

    try:
        new_fichas = _parse_excel(filepath)
    except Exception as e:
        raise HTTPException(422, f"Error al parsear Excel: {e}")

    if not new_fichas:
        raise HTTPException(422, "No se encontraron fichas válidas en el archivo. Verifica el formato.")

    seen_codes = set()
    dup_codes = []
    for fi in new_fichas:
        code = (fi.get("codigo") or "").strip()
        if not code:
            continue
        if code in seen_codes:
            dup_codes.append(code)
        seen_codes.add(code)
    if dup_codes:
        unique_dup_codes = sorted(set(dup_codes))
        raise HTTPException(422, f"Type Mark duplicado en el Excel importado: {', '.join(unique_dup_codes)}")

    _backup(req.version)

    with open(json_path, encoding="utf-8") as f:
        existing = json.load(f)

    idx_by_codigo = {fi["codigo"]: i for i, fi in enumerate(existing)}

    added    = []
    updated  = []

    for fi in new_fichas:
        codigo = (fi.get("codigo") or "").strip()
        desc = (fi.get("descripcion") or "").replace("_x000D_", "").strip()
        fi["codigo"] = codigo
        fi["descripcion"] = desc
        if codigo in idx_by_codigo:
            existing[idx_by_codigo[codigo]] = fi
            updated.append(codigo)
        else:
            existing.append(fi)
            idx_by_codigo[codigo] = len(existing) - 1
            added.append(codigo)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    # Propagate new + updated fichas to all presupuestos using this version
    db_added = _sync_new_fichas_to_db(new_fichas, added + updated, req.version, db)

    return {
        "ok":                True,
        "version":           req.version,
        "fichas_en_archivo": len(new_fichas),
        "agregadas":         added,
        "actualizadas":      updated,
        "total_en_json":     len(existing),
        "partidas_en_db":    db_added,
    }


def _tipo_from_clave(clave: str) -> str:
    c = (clave or "").upper()
    if c.startswith("MO-") or c.startswith("MO."):   return "MANO_OBRA"
    if c.startswith("HE-") or c.startswith("EQ-"):   return "EQUIPO"
    if c.startswith("SC-") or c.startswith("SUB-"):  return "SUBCONTRATO"
    if c.startswith("DIS-"):                          return "DISEÑO"
    return "MATERIAL"


def _sync_new_fichas_to_db(fichas: list, only_codigos: list, version: str, db: Session) -> int:
    """
    For each ficha in `only_codigos`, upsert partida + insumos in every presupuesto
    that uses `version`. Updates prices on existing partidas; creates new ones.
    """
    if not only_codigos:
        return 0

    ficha_map = {f["codigo"]: f for f in fichas if f.get("codigo") in only_codigos}
    if not ficha_map:
        return 0

    configs = db.query(ConfigPresupuesto).filter(
        ConfigPresupuesto.template_version == version
    ).all()

    total_added = 0

    for cfg in configs:
        pid = cfg.presupuesto_id
        sc  = float(cfg.sobrecosto or 20)

        # Build chapter map for this presupuesto
        caps = {c.clave: c for c in db.query(Capitulo).filter(Capitulo.presupuesto_id == pid).all()}
        max_cap_ord = max((c.orden or 0 for c in caps.values()), default=-1) + 1

        # Existing partidas by type_mark
        existing_tm = {}
        for cap in caps.values():
            for p in db.query(Partida).filter(Partida.capitulo_id == cap.id).all():
                if p.type_mark:
                    existing_tm[p.type_mark] = p

        for codigo, fi in ficha_map.items():
            csi = (fi.get("csi") or infer_csi(codigo, fi.get("descripcion", "")) or "00 00 00").strip()
            div = csi[:2] if len(csi) >= 2 and csi[:2].isdigit() else "00"

            # Get or create chapter
            if div not in caps:
                nombre_div = DIVISIONES_CSI.get(div, f"División {div}")
                new_cap = Capitulo(
                    presupuesto_id=pid,
                    clave=div,
                    nombre=nombre_div,
                    orden=max_cap_ord,
                )
                db.add(new_cap)
                db.flush()
                caps[div] = new_cap
                max_cap_ord += 1

            cap = caps[div]

            # Compute costs from insumos
            mo, ma = 0.0, 0.0
            for ins in fi.get("insumos", []):
                qty = float(ins.get("cantidad", 0))
                pu  = float(ins.get("precioUnitario", 0))
                tot = qty * pu
                if ins.get("codigo", "").upper().startswith("MO"):
                    mo += tot
                else:
                    ma += tot
            base = round(mo + ma, 4)
            pu_final = round(base * (1 + sc / 100), 4)

            desc = (fi.get("descripcion") or "").replace("_x000D_", "").strip()
            if not desc:
                desc = codigo

            if codigo in existing_tm:
                # Update prices on existing partida
                p = existing_tm[codigo]
                p.costo_mo = mo
                p.costo_ma = ma
                p.costo_base = base
                p.precio_unitario = pu_final
                p.total = round(float(p.cantidad or 0) * pu_final, 4)
                # Replace insumos
                for ip in list(p.insumos):
                    db.delete(ip)
                db.flush()
            else:
                # Create new partida
                ord_p = db.query(Partida).filter(Partida.capitulo_id == cap.id).count()
                p = Partida(
                    capitulo_id=cap.id,
                    clave_csi=csi,
                    descripcion=desc,
                    unidad=fi.get("unidad", "global"),
                    cantidad=0, revit_q=0,
                    factor_e=1.0, factor_f=1.0,
                    color_tipo=fi.get("color_tipo", "rosa"),
                    costo_mo=mo, costo_ma=ma,
                    unitario_matriz=0,
                    costo_base=base,
                    precio_unitario=pu_final,
                    total=0,
                    es_formula=False,
                    type_mark=codigo,
                    orden=ord_p,
                )
                db.add(p)
                db.flush()
                total_added += 1

            # Add insumos
            for idx, ins in enumerate(fi.get("insumos", [])):
                qty = float(ins.get("cantidad", 0))
                pu_ins = float(ins.get("precioUnitario", 0))
                db.add(InsumoPartida(
                    partida_id  = p.id,
                    recurso_id  = None,
                    clave       = ins.get("codigo", ""),
                    descripcion = ins.get("descripcion", ""),
                    unidad      = ins.get("unidad", "global"),
                    tipo        = _tipo_from_clave(ins.get("codigo", "")),
                    cantidad    = qty,
                    costo_unit  = pu_ins,
                    total       = round(qty * pu_ins, 4),
                    orden       = idx,
                ))

    db.commit()
    return total_added


_OPUS_XLSX  = r"D:\OneDrive\Bots\Estimbot\MasterFiles\BaseDatosOpus2026.xlsx"
_CELL_COLOR = {
    "FFFFFF00": "amarillo",
    "FF92D050": "verde",
    "FF00B0F0": "verde",   # azul → verde (mismo color por diseño)
}


@router.post("/sync-colors")
def sync_colors(db: Session = Depends(get_db)):
    """Copia colores de BaseDatosOpus2026.xlsx a fichas JSON y Partidas en DB."""
    if not _HAS_OPENPYXL:
        raise HTTPException(422, "openpyxl no instalado")
    if not os.path.exists(_OPUS_XLSX):
        raise HTTPException(404, f"No encontrado: {_OPUS_XLSX}")

    # Copy to temp (bypass Excel lock)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        shutil.copy2(_OPUS_XLSX, tmp_path)
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        try: os.unlink(tmp_path)
        except OSError: pass

    ws = wb.active
    color_by_tm: dict = {}

    for row in ws.iter_rows(min_row=3):
        col_b = row[1].value if len(row) > 1 else None
        if col_b is None:
            continue
        tm = str(col_b).strip()
        if not tm:
            continue
        try:
            rgb   = row[0].fill.fgColor.rgb
            color = _CELL_COLOR.get(rgb, None)
        except Exception:
            color = None
        if color:
            color_by_tm[tm] = color

    # Update fichas JSON for all versions
    json_updated: dict = {}
    if os.path.exists(BASE_PATH):
        for d in sorted(os.listdir(BASE_PATH)):
            fp = _ficha_path(d)
            if not os.path.exists(fp):
                continue
            with open(fp, encoding="utf-8") as f:
                fichas = json.load(f)
            changed = 0
            for fi in fichas:
                tm = fi.get("codigo", "")
                if tm in color_by_tm:
                    fi["color_tipo"] = color_by_tm[tm]
                    changed += 1
            if changed:
                with open(fp, "w", encoding="utf-8") as f:
                    json.dump(fichas, f, ensure_ascii=False, indent=2)
                json_updated[d] = changed

    # Update existing Partidas in DB
    db_updated = 0
    for tm, color in color_by_tm.items():
        rows = db.query(Partida).filter(Partida.type_mark == tm).all()
        for p in rows:
            if p.color_tipo != color:
                p.color_tipo = color
                db_updated += 1
    db.commit()

    return {
        "ok":                   True,
        "colores_extraidos":    len(color_by_tm),
        "json_actualizados":    json_updated,
        "partidas_actualizadas": db_updated,
    }
