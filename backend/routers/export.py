from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
import io, sys, os, re
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from db import get_db
from models import Presupuesto, Capitulo, Partida, InsumoPartida

router = APIRouter(tags=["export"])

# Layout 'main' — Opción 1 (BaseDatos-style + Type Mark)
MAIN_HEADERS = [
    "CSI", "Type Mark", "Descripción",
    "Cantidad", "Fórmula", "Cantidad Calc.",
    "Unidad", "Mano de Obra", "INSUMOS",
    "PRECIO UNITARIO", "Total"
]
PROPOSAL_HEADERS = ["CSI", "Descripción", "Cantidad", "PRECIO UNITARIO", "Total"]

COLOR_TIPO_FILL = {
    "amarillo": "FFF2C94C",
    "verde":    "FF92D050",
    "azul":     "FF00B0F0",
    "rosa":     "FFE91E8C",
    "blanco":   None,
}

# Paleta ConsuConstruct
HDR_FILL = "FFF5C518"   # amarillo casco
HDR_TEXT = "FF1A1A1A"   # negro
DIV_FILL = "FF3A3A3A"   # gris concreto oscuro
DIV_TEXT = "FFF5C518"
SUBT_TEXT = "FFD4A815"
TOT_TEXT = "FFE94560"


@router.get("/presupuestos/{pid}/export")
def exportar(pid: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.config),
        joinedload(Presupuesto.capitulos).joinedload(Capitulo.partidas)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    sobrecosto = float(p.config.sobrecosto) / 100 if (p.config and p.config.sobrecosto is not None) else 0.20

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "main"
    ws_proposal = wb.create_sheet("proposal")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===================== MAIN =====================
    # Título
    ws_main.merge_cells("A1:K1")
    ws_main["A1"] = p.nombre
    ws_main["A1"].font = Font(bold=True, size=14, color=HDR_TEXT)
    ws_main["A1"].fill = PatternFill("solid", fgColor=HDR_FILL)
    ws_main["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_main.row_dimensions[1].height = 24

    ws_main.merge_cells("A2:K2")
    ws_main["A2"] = f"Cliente: {p.cliente or '—'}   |   Moneda: {p.moneda}   |   Sobrecosto: {sobrecosto*100:.1f}%"
    ws_main["A2"].font = Font(size=10, italic=True, color="666666")
    ws_main["A2"].alignment = Alignment(horizontal="center")

    # Headers (fila 4)
    hdr_row = 4
    for ci, txt in enumerate(MAIN_HEADERS, 1):
        cell = ws_main.cell(row=hdr_row, column=ci, value=txt)
        cell.font = Font(bold=True, color=HDR_TEXT, size=10)
        cell.fill = PatternFill("solid", fgColor=HDR_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws_main.row_dimensions[hdr_row].height = 30

    row = hdr_row + 1

    # Mapa pid → fila en main (para que proposal pueda referenciar)
    main_row_map = {}
    # Partidas amarillas que necesitan fórmula E al final: [(row, factor_e)]
    yellow_rows = []

    for cap in sorted(p.capitulos, key=lambda c: c.orden):
        # División header
        ws_main.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        hcell = ws_main.cell(row=row, column=1, value=f"{cap.clave} — {cap.nombre}")
        hcell.font = Font(bold=True, color=DIV_TEXT, size=11)
        hcell.fill = PatternFill("solid", fgColor=DIV_FILL)
        hcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_main.row_dimensions[row].height = 20
        row += 1

        cap_first = row

        for pa in sorted(cap.partidas, key=lambda x: x.orden):
            fill_color = COLOR_TIPO_FILL.get(pa.color_tipo or "blanco")
            fill = PatternFill("solid", fgColor=fill_color) if fill_color else None

            r = row  # fila absoluta de esta partida
            main_row_map[pa.id] = r

            # A=CSI  B=Type Mark  C=Descripción
            ws_main.cell(row=r, column=1, value=pa.clave_csi)
            ws_main.cell(row=r, column=2, value=pa.type_mark or "")
            ws_main.cell(row=r, column=3, value=pa.descripcion)

            # D = Cantidad bruta (revit_q)
            ws_main.cell(row=r, column=4, value=float(pa.revit_q or 0))

            # E: amarillas obtienen fórmula revit_q × factor_e (desde DB). Se rellena al final.
            if (pa.color_tipo or "blanco") == "amarillo":
                yellow_rows.append((r, float(pa.factor_e or 1)))
            # else: E queda vacío

            # F = redondeo
            ws_main.cell(row=r, column=6,
                value=f'=IF(E{r}=0,"",IF(E{r}>5,ROUNDUP(E{r}+1,0),ROUNDUP(E{r},0)))')

            # G = Unidad
            ws_main.cell(row=r, column=7, value=pa.unidad or "")

            # H = MO unit, I = MA unit
            ws_main.cell(row=r, column=8, value=float(pa.costo_mo or 0))
            ws_main.cell(row=r, column=9, value=float(pa.costo_ma or 0))

            # J = P.U. = (H + I) * (1 + sobrecosto)
            ws_main.cell(row=r, column=10, value=f"=(H{r}+I{r})*(1+{sobrecosto})")

            # K = Total = F * J  (con guard cuando F está vacío)
            ws_main.cell(row=r, column=11,
                value=f'=IF(OR(F{r}="",F{r}=0),0,F{r}*J{r})')

            # Estilos
            for ci in range(1, 12):
                c = ws_main.cell(row=r, column=ci)
                c.border = border
                if fill:
                    c.fill = fill
                if ci in (4, 5, 6, 8, 9, 10, 11):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '#,##0.00' if ci >= 8 else '#,##0.0000'
                if ci == 1:
                    c.font = Font(size=10, color="666666")
                if ci == 2:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            row += 1

        cap_last = row - 1
        # Subtotal del capítulo
        ws_main.cell(row=row, column=10, value="Subtotal").font = Font(bold=True, color=SUBT_TEXT)
        ws_main.cell(row=row, column=10).alignment = Alignment(horizontal="right")
        if cap_last >= cap_first:
            ws_main.cell(row=row, column=11, value=f"=SUM(K{cap_first}:K{cap_last})")
        else:
            ws_main.cell(row=row, column=11, value=0)
        ws_main.cell(row=row, column=11).font = Font(bold=True, color=SUBT_TEXT)
        ws_main.cell(row=row, column=11).number_format = '#,##0.00'
        row += 2  # blank line

    # Total de obra
    total_row = row + 1
    # Sumar todos los subtotales — más simple: SUM toda la columna K hasta total_row-2 dividido por 2 (porque cada partida y cada subtotal se sumarían). Usar SUMIF a partidas mejor.
    # Práctico: sumar todos los subtotales (que tienen "Subtotal" en col J).
    ws_main.cell(row=total_row, column=10, value="TOTAL OBRA").font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_main.cell(row=total_row, column=10).alignment = Alignment(horizontal="right")
    ws_main.cell(row=total_row, column=11, value=f'=SUMIF(J5:J{total_row-1},"Subtotal",K5:K{total_row-1})')
    ws_main.cell(row=total_row, column=11).font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_main.cell(row=total_row, column=11).number_format = '#,##0.00'

    # Segundo pass: fórmulas E para amarillas — factor_e desde estimacion.db
    for r, factor_e in yellow_rows:
        ws_main.cell(row=r, column=5, value=f"=$D${r}*{factor_e:.6f}")

    # Anchos
    widths_main = [13, 11, 50, 11, 11, 13, 9, 12, 12, 13, 14]
    for ci, w in enumerate(widths_main, 1):
        ws_main.column_dimensions[get_column_letter(ci)].width = w

    ws_main.freeze_panes = "A5"

    # ===================== PROPOSAL =====================
    # Encabezado de empresa (estilo Template.proposal)
    company_lines = [
        ("Consultorías de Construcción S. de R.L.", True, 14),
        ("Nosotros lo Construimos", False, 11),
        ("Ing. David A. Chinchilla  |  CICH 8222", False, 10),
        ("Tegucigalpa, MDC, Honduras", False, 10),
        ("+504 9662-8408   |   davidchinchilla@consuconstruct.com", False, 10),
        ("https://consuconstruct.com", False, 10),
    ]
    for i, (txt, bold, size) in enumerate(company_lines, 1):
        ws_proposal.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
        c = ws_proposal.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=size, color=HDR_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if i == 1:
            c.fill = PatternFill("solid", fgColor=HDR_FILL)
            ws_proposal.row_dimensions[i].height = 22

    # Título del presupuesto
    title_row = len(company_lines) + 2
    ws_proposal.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)
    tc = ws_proposal.cell(row=title_row, column=1, value=f"ESTIMACION DE OBRA — {p.nombre}")
    tc.font = Font(bold=True, size=12, color=HDR_TEXT)
    tc.fill = PatternFill("solid", fgColor=HDR_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_proposal.row_dimensions[title_row].height = 22

    # Datos del cliente
    info_row = title_row + 1
    ws_proposal.cell(row=info_row, column=1, value="CLIENTE:").font = Font(bold=True, size=10)
    ws_proposal.merge_cells(start_row=info_row, start_column=2, end_row=info_row, end_column=5)
    ws_proposal.cell(row=info_row, column=2, value=p.cliente or "—").font = Font(size=10)

    info_row += 1
    ws_proposal.cell(row=info_row, column=1, value="MONEDA:").font = Font(bold=True, size=10)
    ws_proposal.cell(row=info_row, column=2, value=p.moneda).font = Font(size=10)

    # Headers de tabla
    tbl_hdr_row = info_row + 2
    for ci, txt in enumerate(PROPOSAL_HEADERS, 1):
        c = ws_proposal.cell(row=tbl_hdr_row, column=ci, value=txt)
        c.font = Font(bold=True, color=HDR_TEXT, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws_proposal.row_dimensions[tbl_hdr_row].height = 24

    # Filas: una por capítulo + sus partidas
    pr = tbl_hdr_row + 1
    for cap in sorted(p.capitulos, key=lambda c: c.orden):
        ws_proposal.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=5)
        hc = ws_proposal.cell(row=pr, column=1, value=f"{cap.clave} — {cap.nombre}")
        hc.font = Font(bold=True, color=DIV_TEXT, size=11)
        hc.fill = PatternFill("solid", fgColor=DIV_FILL)
        hc.alignment = Alignment(horizontal="left", indent=1)
        pr += 1

        for pa in sorted(cap.partidas, key=lambda x: x.orden):
            mr = main_row_map.get(pa.id)
            ws_proposal.cell(row=pr, column=1, value=pa.clave_csi).font = Font(size=10, color="666666")
            ws_proposal.cell(row=pr, column=2, value=pa.descripcion).font = Font(size=10)
            if mr:
                ws_proposal.cell(row=pr, column=3, value=f"=main!F{mr}")
                ws_proposal.cell(row=pr, column=4, value=f"=main!J{mr}")
                ws_proposal.cell(row=pr, column=5, value=f"=main!K{mr}")
            else:
                ws_proposal.cell(row=pr, column=3, value=0)
                ws_proposal.cell(row=pr, column=4, value=0)
                ws_proposal.cell(row=pr, column=5, value=0)
            for ci in (3, 4, 5):
                cc = ws_proposal.cell(row=pr, column=ci)
                cc.alignment = Alignment(horizontal="right")
                cc.number_format = '#,##0.00'
            for ci in range(1, 6):
                ws_proposal.cell(row=pr, column=ci).border = border
            pr += 1

    # Total final en proposal
    pr += 1
    ws_proposal.cell(row=pr, column=4, value="TOTAL").font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_proposal.cell(row=pr, column=4).alignment = Alignment(horizontal="right")
    last_total_row = ws_proposal.max_row  # before this insert
    ws_proposal.cell(row=pr, column=5, value=f"=SUM(E{tbl_hdr_row+1}:E{pr-1})")
    ws_proposal.cell(row=pr, column=5).font = Font(bold=True, size=12, color=TOT_TEXT)
    ws_proposal.cell(row=pr, column=5).number_format = '#,##0.00'

    widths_prop = [14, 55, 14, 14, 16]
    for ci, w in enumerate(widths_prop, 1):
        ws_proposal.column_dimensions[get_column_letter(ci)].width = w

    ws_proposal.freeze_panes = f"A{tbl_hdr_row+1}"

    # ===================== Output =====================
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = (p.nombre or "obra").replace(" ", "_").replace("/", "-")[:40]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_{nombre}.xlsx"'},
    )


# ───────────────────────────────────────────────────────────────────────────
#  EXPORT BASE DE DATOS COMPLETA — un renglón por (matriz, insumo)
# ───────────────────────────────────────────────────────────────────────────

DB_HEADERS = ["CSI / Recurso", "Type Mark / Clave", "Descripción", "Unidad",
              "Cantidad / Rendimiento", "MO / Costo Unit.", "Insumos / Total"]

MATRIZ_TEXT_COLOR = "FF1F6FB7"   # azul para fila matriz (estilo OPUS)


@router.get("/presupuestos/{pid}/export-db")
def exportar_db(pid: str, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no disponible")

    p = db.query(Presupuesto).options(
        joinedload(Presupuesto.capitulos)
        .joinedload(Capitulo.partidas)
        .joinedload(Partida.insumos)
    ).filter(Presupuesto.id == pid).first()
    if not p:
        raise HTTPException(404, "Presupuesto no encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "BD"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DB_HEADERS))
    t = ws.cell(row=1, column=1, value=f"{p.nombre} — Base de Datos Completa")
    t.font = Font(bold=True, size=13, color=HDR_TEXT)
    t.fill = PatternFill("solid", fgColor=HDR_FILL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for ci, txt in enumerate(DB_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=txt)
        c.font = Font(bold=True, color=HDR_TEXT, size=10)
        c.fill = PatternFill("solid", fgColor=HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[2].height = 26

    def csi_natural_key(pa):
        s = pa.clave_csi or ""
        parts = []
        for tok in re.findall(r"\d+|\D+", s):
            parts.append((0, int(tok)) if tok.isdigit() else (1, tok.lower()))
        return parts

    row = 3
    for cap in sorted(p.capitulos, key=lambda c: (c.orden if c.orden is not None else 999, c.clave or "")):
        # Header de capítulo
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DB_HEADERS))
        hc = ws.cell(row=row, column=1, value=f"{cap.clave} — {cap.nombre}")
        hc.font = Font(bold=True, color=DIV_TEXT, size=11)
        hc.fill = PatternFill("solid", fgColor=DIV_FILL)
        hc.alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        for pa in sorted(cap.partidas, key=csi_natural_key):
            color_fill = COLOR_TIPO_FILL.get(pa.color_tipo or "blanco")
            fill = PatternFill("solid", fgColor=color_fill) if color_fill else None

            # ── Fila MATRIZ (valores duros: cantidad, MO total, INSUMOS total) ──
            mo_total  = sum(float(i.total or 0) for i in pa.insumos if i.tipo == "MANO_OBRA")
            ins_total = sum(float(i.total or 0) for i in pa.insumos if i.tipo != "MANO_OBRA")
            matriz_vals = [
                pa.clave_csi or "",
                pa.type_mark or "",
                pa.descripcion or "",
                pa.unidad or "",
                float(pa.cantidad or 0),
                mo_total,
                ins_total,
            ]
            for ci, v in enumerate(matriz_vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.border = border
                if fill:
                    c.fill = fill
                c.font = Font(bold=True, color=MATRIZ_TEXT_COLOR, size=10)
                if ci in (5, 6, 7) and isinstance(v, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '#,##0.0000' if ci == 5 else '#,##0.00'
            row += 1

            # ── Filas RECURSO: insumos no-MO primero, MO después; alfabético por clave ──
            insumos = sorted(
                pa.insumos,
                key=lambda i: (1 if i.tipo == "MANO_OBRA" else 0, (i.clave or "").lower())
            )
            for ins in insumos:
                ins_vals = [
                    "Recurso",
                    ins.clave or "",
                    ins.descripcion or "",
                    ins.unidad or "",
                    float(ins.cantidad or 0),
                    float(ins.costo_unit or 0),
                    float(ins.total or 0),
                ]
                for ci, v in enumerate(ins_vals, 1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.border = border
                    c.font = Font(size=10)
                    if ci == 1:
                        c.font = Font(size=9, color="888888", italic=True)
                    if ci in (5, 6, 7) and isinstance(v, (int, float)):
                        c.alignment = Alignment(horizontal="right", vertical="center")
                        c.number_format = '#,##0.0000' if ci == 5 else '#,##0.00'
                row += 1

            # Fila vacía entre matrices
            row += 1

    # Anchos de columna
    widths = [13, 14, 50, 10, 14, 14, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = (p.nombre or "obra").replace(" ", "_").replace("/", "-")[:40]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Estimastruct_BD_{nombre}.xlsx"'},
    )
